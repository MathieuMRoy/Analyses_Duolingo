"""
Module d'orchestration pour la génération du rapport Excel final.
Prend les statistiques calculées et les formats pour générer le workbook.
"""

from pathlib import Path
import shutil
import pandas as pd
from openpyxl import load_workbook

from .alternative_data import build_alternative_data_raw_df
from .columns import (
    ALT_DATA_RAW_SHEET,
    ALT_DATA_SHEET,
    BAD_SHEET_NAMES,
    CHART_DATA_SHEET,
    DCF_SHEET,
    GLOSSAIRE_RAW_SHEET,
    GLOSSAIRE_SHEET,
    QUARTERLY_RAW_SHEET,
    QUARTERLY_SHEET,
    SIGNALS_RAW_SHEET,
    SIGNALS_SHEET,
    SUMMARY_COLUMN_ALIASES,
    SUMMARY_SHEET,
    TRENDS_SHEET,
)
from .config import DAILY_LOG_FILE, GOOGLE_DRIVE_REPORT_DIR, RAPPORT_EXCEL_FILE, REPORT_DIR
from .excel_dashboard import refresh_trends_dashboard
from .financial_signals import build_financial_signal_sheet_df
from .quarterly_nowcast import build_quarterly_nowcast_raw_df
from .reporting.sheets.dcf_valuation_sheet import render_dcf_valuation_sheet
from .reporting.sheets.alternative_data_sheet import render_alternative_data_sheet
from .reporting.sheets.briefing_ai_sheet import render_briefing_ai_sheet
from .reporting.sheets.financial_nowcast_sheet import render_financial_nowcast_sheet
from .reporting.sheets.kpi_dictionary_sheet import build_kpi_dictionary_df
from .reporting.sheets.monthly_trends_sheet import add_monthly_trends_chart, build_monthly_trends_frames
from .reporting.sheets.quarterly_nowcast_sheet import render_quarterly_nowcast_sheet
from .reporting.sheets.summary_sheet import build_summary_today_df, merge_summary_history
from .reporting.render_helpers import build_render_helpers
from .reporting.styles import build_style_context
from .reporting.workbook_layout import apply_standard_workbook_layout
from .reporting.workbook_postprocess import apply_standard_table_style
from .stats import _build_summary_history_from_log, _load_daily_log_df, _normalize_summary_df


def _load_summary_sheet(report_path: Path) -> pd.DataFrame | None:
    try:
        workbook = pd.ExcelFile(report_path)
    except Exception:
        return None

    for sheet_name in [SUMMARY_SHEET, *BAD_SHEET_NAMES]:
        if sheet_name not in workbook.sheet_names:
            continue
        try:
            df_summary = pd.read_excel(report_path, sheet_name=sheet_name)
            if df_summary is not None and not df_summary.empty:
                return df_summary
        except Exception:
            continue

    for sheet_name in workbook.sheet_names:
        try:
            df_preview = pd.read_excel(report_path, sheet_name=sheet_name, nrows=5)
        except Exception:
            continue

        normalized_headers = {
            SUMMARY_COLUMN_ALIASES.get(str(column).strip())
            for column in df_preview.columns
        }
        if {"Date", "Panel Total"}.issubset(normalized_headers):
            try:
                df_summary = pd.read_excel(report_path, sheet_name=sheet_name)
                if df_summary is not None and not df_summary.empty:
                    return df_summary
            except Exception:
                continue

    return None


def _charger_resume_historique() -> pd.DataFrame | None:
    frames: list[pd.DataFrame] = []

    try:
        df_log = _load_daily_log_df()
        df_from_log = _build_summary_history_from_log(df_log)
        if df_from_log is not None and not df_from_log.empty:
            frames.append(df_from_log)
    except Exception:
        pass

    if RAPPORT_EXCEL_FILE.exists():
        try:
            df_resume = _load_summary_sheet(RAPPORT_EXCEL_FILE)
            if df_resume is not None and not df_resume.empty:
                frames.append(df_resume)
        except Exception:
            pass

    try:
        daily_reports = sorted(REPORT_DIR.glob("rapport_*.xlsx"))
    except Exception:
        daily_reports = []

    for report_path in daily_reports:
        if report_path == RAPPORT_EXCEL_FILE:
            continue
        try:
            df_tmp = _load_summary_sheet(report_path)
            if df_tmp is not None and not df_tmp.empty:
                frames.append(df_tmp)
        except Exception:
            continue

    if not frames:
        return None

    df_resume = pd.concat(frames, ignore_index=True)
    return _normalize_summary_df(df_resume)


def _copier_rapport_vers_google_drive(report_path: Path) -> Path | None:
    if not GOOGLE_DRIVE_REPORT_DIR:
        return None

    destination_dir = Path(GOOGLE_DRIVE_REPORT_DIR).expanduser()
    try:
        destination_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        print(f"  ⚠️ Impossible de créer le dossier Google Drive : {destination_dir} ({exc})")
        return None

    destination_file = destination_dir / report_path.name
    try:
        shutil.copy2(report_path, destination_file)
    except Exception as exc:
        print(f"  ⚠️ Copie Google Drive impossible : {destination_file} ({exc})")
        return None

    return destination_file


def sauvegarder_rapport_excel(
    stats: dict,
    ia_report: str = None,
    financial_signals: dict | None = None,
    quarterly_nowcast: dict | None = None,
    alternative_data: dict | None = None,
    dcf_valuation: dict | None = None,
) -> None:
    """
    Exporte les données d'engagement et les statistiques dans un fichier Excel (.xlsx)
    avec un design premium et l'analyse de l'IA.
    """
    print("  📂 Génération du rapport Excel Premium...")

    if not DAILY_LOG_FILE.exists():
        return

    try:
        df = _load_daily_log_df()
        df = df[~df["Username"].str.contains("Aggregated", na=False)]
        df = df[df["Cohort"] != "Global"]
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
        df = df.sort_values(["Date", "Username"]).drop_duplicates(subset=["Date", "Username"], keep="last")

        df_stats = build_summary_today_df(stats)
        df_resume = merge_summary_history(_charger_resume_historique(), df_stats, _normalize_summary_df)

        writer_mode = "a" if RAPPORT_EXCEL_FILE.exists() else "w"
        writer_kwargs = {"if_sheet_exists": "replace"} if writer_mode == "a" else {}

        with pd.ExcelWriter(RAPPORT_EXCEL_FILE, engine="openpyxl", mode=writer_mode, **writer_kwargs) as writer:
            df_resume.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)

            if financial_signals:
                signal_sheet_df = build_financial_signal_sheet_df(financial_signals)
                pd.DataFrame().to_excel(writer, sheet_name=SIGNALS_SHEET, index=False)
                signal_sheet_df.to_excel(writer, sheet_name=SIGNALS_RAW_SHEET, index=False)
            if quarterly_nowcast:
                quarterly_sheet_df = build_quarterly_nowcast_raw_df(quarterly_nowcast)
                pd.DataFrame().to_excel(writer, sheet_name=QUARTERLY_SHEET, index=False)
                quarterly_sheet_df.to_excel(writer, sheet_name=QUARTERLY_RAW_SHEET, index=False)
            if alternative_data is not None:
                alternative_data_df = build_alternative_data_raw_df(alternative_data)
                pd.DataFrame().to_excel(writer, sheet_name=ALT_DATA_SHEET, index=False)
                alternative_data_df.to_excel(writer, sheet_name=ALT_DATA_RAW_SHEET, index=False)
            if dcf_valuation:
                pd.DataFrame().to_excel(writer, sheet_name=DCF_SHEET, index=False)

            df_glossaire = build_kpi_dictionary_df()
            pd.DataFrame().to_excel(writer, sheet_name=GLOSSAIRE_SHEET, index=False)
            df_glossaire.to_excel(writer, sheet_name=GLOSSAIRE_RAW_SHEET, index=False)

        monthly_stats, df_daily_conv = build_monthly_trends_frames(df)

        with pd.ExcelWriter(RAPPORT_EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            if not monthly_stats.empty:
                monthly_stats.to_excel(writer, sheet_name=TRENDS_SHEET, index=False)

            if not df_daily_conv.empty:
                df_daily_conv.to_excel(writer, sheet_name=CHART_DATA_SHEET, index=False)

        wb = load_workbook(RAPPORT_EXCEL_FILE)
        apply_standard_workbook_layout(
            wb,
            include_dcf=DCF_SHEET in wb.sheetnames,
            hide_chart_data=False,
        )

        style_ctx = build_style_context()
        render_helpers = build_render_helpers()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            if sheet_name == SIGNALS_SHEET and financial_signals:
                render_financial_nowcast_sheet(ws, financial_signals, ia_report, style_ctx, render_helpers)
                ws.freeze_panes = "A5"
                continue

            if sheet_name == QUARTERLY_SHEET and quarterly_nowcast:
                render_quarterly_nowcast_sheet(
                    ws,
                    quarterly_nowcast,
                    wb,
                    QUARTERLY_RAW_SHEET,
                    style_ctx,
                    ia_report,
                    render_helpers,
                )
                ws.freeze_panes = "A5"
                continue

            if sheet_name == ALT_DATA_SHEET and alternative_data is not None:
                render_alternative_data_sheet(ws, alternative_data, style_ctx)
                ws.freeze_panes = "A8"
                continue

            if sheet_name == DCF_SHEET and dcf_valuation:
                render_dcf_valuation_sheet(ws, dcf_valuation, style_ctx)
                ws.freeze_panes = "A5"
                continue

            if sheet_name == GLOSSAIRE_SHEET:
                render_briefing_ai_sheet(
                    ws,
                    wb,
                    GLOSSAIRE_RAW_SHEET,
                    style_ctx,
                    render_helpers,
                    ia_report=ia_report,
                    signal_package=financial_signals,
                    quarterly_nowcast=quarterly_nowcast,
                    alternative_data=alternative_data,
                    dcf_valuation=dcf_valuation,
                )
                ws.freeze_panes = "A5"
                continue

            if sheet_name == TRENDS_SHEET:
                try:
                    add_monthly_trends_chart(ws, wb, CHART_DATA_SHEET)
                except Exception as e:
                    print(f"  ⚠️ Erreur ajout graphique : {e}")

            apply_standard_table_style(
                ws,
                sheet_name,
                style_ctx,
                summary_sheet_name=SUMMARY_SHEET,
                glossary_sheet_name=GLOSSAIRE_SHEET,
                chart_data_sheet_name=CHART_DATA_SHEET,
            )

        apply_standard_workbook_layout(
            wb,
            include_dcf=DCF_SHEET in wb.sheetnames,
            hide_chart_data=False,
        )

        wb.save(RAPPORT_EXCEL_FILE)
        refresh_trends_dashboard(RAPPORT_EXCEL_FILE)
        wb = load_workbook(RAPPORT_EXCEL_FILE)
        apply_standard_workbook_layout(
            wb,
            include_dcf=DCF_SHEET in wb.sheetnames,
            hide_chart_data=True,
        )
        wb.save(RAPPORT_EXCEL_FILE)
        google_drive_file = _copier_rapport_vers_google_drive(RAPPORT_EXCEL_FILE)

        print("  ✅ Rapport Excel Premium sauvegardé dans :")
        print(f"     → {RAPPORT_EXCEL_FILE}")
        if google_drive_file is not None:
            print(f"     → Copie Google Drive : {google_drive_file}")

    except Exception as e:
        import traceback
        print(f"  ❌ Erreur lors de la sauvegarde Excel : {e}")
        traceback.print_exc()
