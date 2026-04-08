"""Read-only ADK tools for the Duolingo investor workbook.

These tools sit *around* the existing deterministic pipeline. They do not
replace the model or workbook logic; they expose the current state in a way an
ADK multi-agent layer can query safely.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from ..alternative_data import (
    _build_latest_signal_rows,
    _build_weekly_summary,
    _load_history_df,
)
from ..columns import (
    ALT_DATA_RAW_SHEET,
    ALT_DATA_SHEET,
    CHART_DATA_SHEET,
    DCF_SHEET,
    GLOSSAIRE_RAW_SHEET,
    GLOSSAIRE_SHEET,
    QUARTERLY_RAW_SHEET,
    QUARTERLY_SHEET,
    SIGNALS_RAW_SHEET,
    SIGNALS_SHEET,
    SUMMARY_SHEET,
    TRENDS_SHEET,
)
from ..config import RAPPORT_EXCEL_FILE, now_toronto
from ..financial_signals import (
    FINANCIAL_SIGNALS_HISTORY_FILE,
    FINANCIAL_SIGNALS_JSON_FILE,
    build_financial_signal_package,
)
from ..quarterly_nowcast import (
    QUARTERLY_LOCKED_ARCHIVE_FILE,
    QUARTERLY_NOWCAST_JSON_FILE,
    QUARTERLY_SNAPSHOTS_FILE,
    build_quarterly_nowcast_package,
)
from ..valuation_dcf import DCF_VALUATION_JSON_FILE, build_dcf_valuation_package


REQUIRED_VISIBLE_SHEETS = [
    SUMMARY_SHEET,
    SIGNALS_SHEET,
    QUARTERLY_SHEET,
    ALT_DATA_SHEET,
    TRENDS_SHEET,
    DCF_SHEET,
    GLOSSAIRE_SHEET,
]

EXPECTED_HIDDEN_RAW_SHEETS = [
    SIGNALS_RAW_SHEET,
    QUARTERLY_RAW_SHEET,
    ALT_DATA_RAW_SHEET,
    GLOSSAIRE_RAW_SHEET,
    CHART_DATA_SHEET,
]

EXPECTED_SOCIAL_SIGNAL_KEYS = {
    "instagram_followers",
    "tiktok_followers",
    "youtube_subscribers",
}


def _today_iso() -> str:
    return now_toronto().strftime("%Y-%m-%d")


def _coerce_reference_date(reference_date: str | None) -> str | None:
    raw = (reference_date or "").strip()
    return raw or None


def _safe_load_json(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    return payload if isinstance(payload, dict) else None


def _pct_text(value: object, digits: int = 1) -> str:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return "N/D"
    return f"{numeric * 100:.{digits}f}%"


def _money_text(value: object, digits: int = 1) -> str:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return "N/D"
    return f"{numeric:.{digits}f} M$"


def _load_financial_signal_package(reference_date: str | None = None) -> dict[str, Any] | None:
    reference_date = _coerce_reference_date(reference_date)
    cached = _safe_load_json(FINANCIAL_SIGNALS_JSON_FILE)
    cached_date = (cached or {}).get("metadata", {}).get("as_of_date")
    if cached and (reference_date is None or str(cached_date) == reference_date):
        return cached
    return build_financial_signal_package(reference_date)


def _load_quarterly_nowcast_package(reference_date: str | None = None) -> dict[str, Any] | None:
    reference_date = _coerce_reference_date(reference_date)
    cached = _safe_load_json(QUARTERLY_NOWCAST_JSON_FILE)
    cached_date = (cached or {}).get("metadata", {}).get("as_of_date")
    if cached and (reference_date is None or str(cached_date) == reference_date):
        return cached
    return build_quarterly_nowcast_package(reference_date)


def _load_dcf_package(reference_date: str | None = None) -> dict[str, Any] | None:
    reference_date = _coerce_reference_date(reference_date)
    cached = _safe_load_json(DCF_VALUATION_JSON_FILE)
    cached_date = (cached or {}).get("metadata", {}).get("as_of_date")
    if cached and (reference_date is None or str(cached_date) == reference_date):
        return cached

    nowcast = _load_quarterly_nowcast_package(reference_date)
    if not nowcast:
        return None
    return build_dcf_valuation_package(nowcast, reference_date=reference_date)


def _load_alternative_data_package() -> dict[str, Any]:
    history_df = _load_history_df()
    rows = _build_latest_signal_rows(history_df)
    weekly_summary_columns, weekly_summary_rows = _build_weekly_summary(history_df)

    latest_snapshot_date = max((row.get("snapshot_date") for row in rows), default=None)
    latest_week_label = rows[0].get("week_label") if rows else None
    signals_up = sum(
        1 for row in rows if isinstance(row.get("wow_change_pct"), float) and row.get("wow_change_pct", 0.0) > 0
    )
    signals_down = sum(
        1 for row in rows if isinstance(row.get("wow_change_pct"), float) and row.get("wow_change_pct", 0.0) < 0
    )

    return {
        "metadata": {
            "has_data": bool(rows),
            "latest_snapshot_date": latest_snapshot_date,
            "latest_week_label": latest_week_label,
            "signal_count": len(rows),
            "signals_up": signals_up,
            "signals_down": signals_down,
        },
        "rows": rows,
        "weekly_summary_columns": weekly_summary_columns,
        "weekly_summary_rows": weekly_summary_rows,
        "history_rows": history_df.to_dict("records") if not history_df.empty else [],
    }


def _select_quarter_snapshot(package: dict[str, Any], quarter: str | None = None) -> dict[str, Any] | None:
    if not package:
        return None

    target_quarter = (quarter or "").strip() or str(package.get("metadata", {}).get("default_selected_quarter") or "")
    current_snapshot = package.get("current_quarter") or {}
    snapshots = package.get("historical_snapshots") or []

    if target_quarter:
        for snapshot in snapshots:
            if str(snapshot.get("quarter")) == target_quarter:
                return snapshot

    return current_snapshot or (snapshots[-1] if snapshots else None)


def _compute_dcf_readout(package: dict[str, Any]) -> dict[str, Any]:
    assumptions = package.get("assumptions", {}) or {}
    anchors = package.get("anchors", {}) or {}
    market = package.get("market_context", {}) or {}

    growth_rate = assumptions.get("growth_rate")
    wacc = assumptions.get("wacc")
    terminal_growth = assumptions.get("terminal_growth")
    fcf_base = anchors.get("free_cash_flow_base_less_sbc_musd")
    if fcf_base in (None, 0):
        fcf_base = anchors.get("free_cash_flow_base_musd")
    net_cash = anchors.get("net_cash_musd")
    diluted_shares = anchors.get("diluted_shares_m")

    try:
        growth_rate = float(growth_rate)
        wacc = float(wacc)
        terminal_growth = float(terminal_growth)
        fcf_base = float(fcf_base)
        net_cash = float(net_cash)
        diluted_shares = float(diluted_shares)
    except (TypeError, ValueError):
        return {
            "price_target": None,
            "current_share_price": market.get("current_share_price"),
            "upside_downside_pct": None,
        }

    if diluted_shares <= 0 or wacc <= terminal_growth:
        return {
            "price_target": None,
            "current_share_price": market.get("current_share_price"),
            "upside_downside_pct": None,
        }

    yearly_growth = [
        growth_rate,
        max(terminal_growth, growth_rate - ((growth_rate - terminal_growth) * 1 / 4)),
        max(terminal_growth, growth_rate - ((growth_rate - terminal_growth) * 2 / 4)),
        max(terminal_growth, growth_rate - ((growth_rate - terminal_growth) * 3 / 4)),
        terminal_growth,
    ]

    fcf_values: list[float] = []
    previous_fcf = fcf_base
    for index, year_growth in enumerate(yearly_growth, start=1):
        current_fcf = previous_fcf * (1 + year_growth)
        fcf_values.append(current_fcf / ((1 + wacc) ** index))
        previous_fcf = current_fcf

    terminal_fcf = previous_fcf * (1 + terminal_growth)
    terminal_value = terminal_fcf / (wacc - terminal_growth) / ((1 + wacc) ** 5)
    enterprise_value = sum(fcf_values) + terminal_value
    equity_value = enterprise_value + net_cash
    price_target = equity_value / diluted_shares

    current_share_price = market.get("current_share_price")
    try:
        current_share_price = float(current_share_price)
    except (TypeError, ValueError):
        current_share_price = None

    upside_downside_pct = None
    if current_share_price not in (None, 0):
        upside_downside_pct = price_target / current_share_price - 1.0

    return {
        "price_target": round(price_target, 2),
        "current_share_price": round(current_share_price, 2) if current_share_price is not None else None,
        "upside_downside_pct": round(upside_downside_pct, 4) if upside_downside_pct is not None else None,
    }


def get_workbook_overview() -> dict[str, Any]:
    """Return the current workbook structure, visible tabs, and hidden raw sheets."""
    path = RAPPORT_EXCEL_FILE
    if not path.exists():
        return {
            "available": False,
            "path": str(path),
            "message": "Le workbook principal n'existe pas encore.",
        }

    wb = load_workbook(path, read_only=True, data_only=True)
    visible_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    hidden_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]

    summary_rows = wb[SUMMARY_SHEET].max_row - 1 if SUMMARY_SHEET in wb.sheetnames else 0
    return {
        "available": True,
        "path": str(path),
        "sheet_count": len(wb.sheetnames),
        "visible_sheets": visible_sheets,
        "hidden_sheets": hidden_sheets,
        "summary_history_rows": max(summary_rows, 0),
    }


def get_daily_signals_context(reference_date: str = "") -> dict[str, Any]:
    """Return the latest daily panel and 7d/30d business signal context."""
    package = _load_financial_signal_package(reference_date)
    if not package:
        return {"available": False, "message": "Aucun paquet de signaux financiers disponible."}

    metadata = package.get("metadata", {}) or {}
    panel = package.get("panel", {}) or {}
    business = package.get("business_signals", {}) or {}
    daily = package.get("daily_comparison", {}) or {}
    horizon = package.get("horizon_context", {}) or {}
    proxy = package.get("financial_proxy_signals", {}) or {}

    return {
        "available": True,
        "as_of_date": metadata.get("as_of_date"),
        "phase": metadata.get("phase"),
        "coverage_ratio_pct": _pct_text(panel.get("coverage_ratio")),
        "panel_observed": panel.get("observed_users_today"),
        "panel_target": panel.get("target_panel_size"),
        "signal_bias": proxy.get("signal_bias"),
        "confidence_level": proxy.get("confidence_level"),
        "paid_rate_pct": _pct_text(business.get("paid_rate")),
        "super_rate_pct": _pct_text(business.get("super_rate")),
        "active_rate_pct": _pct_text(business.get("active_rate")),
        "avg_streak_days": business.get("avg_streak"),
        "xp_delta_mean": business.get("xp_delta_mean"),
        "churn_rate_pct": _pct_text(business.get("churn_rate")),
        "reactivation_rate_pct": _pct_text(business.get("reactivation_rate")),
        "high_value_retention_pct": _pct_text(business.get("high_value_retention_rate")),
        "premium_net_adds_today": daily.get("premium_net_adds_today"),
        "reactivated_users_today": daily.get("reactivated_users_today"),
        "churned_users_today": daily.get("churned_users_today"),
        "seven_day_summary": (horizon.get("seven_day") or {}).get("summary_text"),
        "thirty_day_summary": (horizon.get("thirty_day") or {}).get("summary_text"),
        "main_drivers": proxy.get("main_drivers") or [],
        "main_risks": proxy.get("main_risks") or [],
    }


def get_quarterly_nowcast_context(quarter: str = "", reference_date: str = "") -> dict[str, Any]:
    """Return the selected quarter nowcast context, estimates, and confidence drivers."""
    package = _load_quarterly_nowcast_package(reference_date)
    if not package:
        return {"available": False, "message": "Aucun paquet de nowcast trimestriel disponible."}

    snapshot = _select_quarter_snapshot(package, quarter)
    if not snapshot:
        return {"available": False, "message": "Aucun snapshot trimestriel exploitable."}

    model_output = package.get("model_output", {}) or {}
    selected_quarter = str(snapshot.get("quarter") or quarter or package.get("metadata", {}).get("current_quarter"))
    is_current = selected_quarter == str(package.get("metadata", {}).get("current_quarter"))

    return {
        "available": True,
        "selected_quarter": selected_quarter,
        "current_quarter": package.get("metadata", {}).get("current_quarter"),
        "is_current_quarter": is_current,
        "snapshot_status": snapshot.get("snapshot_status_label"),
        "snapshot_locked": bool(snapshot.get("snapshot_locked")),
        "snapshot_as_of_date": snapshot.get("snapshot_as_of_date"),
        "observed_days": snapshot.get("observed_days"),
        "avg_coverage_ratio_pct": _pct_text(snapshot.get("avg_coverage_ratio")),
        "quarter_signal_bias": snapshot.get("quarter_signal_bias"),
        "confidence_level": snapshot.get("confidence_level"),
        "confidence_context_text": snapshot.get("confidence_context_text") or model_output.get("confidence_context_text"),
        "revenue_guidance_reference_musd": _money_text(snapshot.get("revenue_guidance_reference_musd")),
        "estimated_revenue_musd": _money_text(snapshot.get("estimated_revenue_musd")),
        "estimated_ebitda_musd": _money_text(snapshot.get("estimated_ebitda_musd")),
        "estimated_eps": (
            f"{float(snapshot.get('estimated_eps')):.2f} $"
            if snapshot.get("estimated_eps") not in (None, "")
            else "N/D"
        ),
        "revenue_beat_probability_pct": _pct_text(snapshot.get("revenue_beat_probability_proxy")),
        "ebitda_beat_probability_pct": _pct_text(snapshot.get("ebitda_beat_probability_proxy")),
        "guidance_raise_probability_pct": _pct_text(snapshot.get("guidance_raise_probability_proxy")),
        "estimated_next_q_guidance_musd": _money_text(snapshot.get("estimated_next_q_guidance_musd")),
        "revenue_estimation_basis": snapshot.get("revenue_estimation_basis"),
        "main_drivers": snapshot.get("main_drivers") or [],
        "main_risks": snapshot.get("main_risks") or [],
        "available_quarters": package.get("metadata", {}).get("available_quarters") or [],
    }


def get_alternative_data_context(weeks: int = 4) -> dict[str, Any]:
    """Return the latest alternative data snapshot plus the recent weekly average table."""
    package = _load_alternative_data_package()
    metadata = package.get("metadata", {}) or {}
    rows = package.get("rows") or []
    weekly_rows = package.get("weekly_summary_rows") or []

    recent_weeks = weekly_rows[-max(int(weeks or 1), 1):]
    social_signals = {
        row.get("signal_key")
        for row in rows
        if row.get("signal_key") in EXPECTED_SOCIAL_SIGNAL_KEYS
    }

    return {
        "available": bool(rows),
        "latest_snapshot_date": metadata.get("latest_snapshot_date"),
        "latest_week_label": metadata.get("latest_week_label"),
        "signal_count": metadata.get("signal_count", 0),
        "signals_up": metadata.get("signals_up", 0),
        "signals_down": metadata.get("signals_down", 0),
        "latest_rows": rows,
        "weekly_summary_columns": package.get("weekly_summary_columns") or [],
        "recent_weekly_summary_rows": recent_weeks,
        "missing_social_signals": sorted(EXPECTED_SOCIAL_SIGNAL_KEYS - social_signals),
    }


def get_dcf_context(reference_date: str = "") -> dict[str, Any]:
    """Return the DCF context with an explicit price target and upside/downside readout."""
    package = _load_dcf_package(reference_date)
    if not package:
        return {"available": False, "message": "Aucun paquet DCF disponible."}

    computed = _compute_dcf_readout(package)
    assumptions = package.get("assumptions", {}) or {}
    anchors = package.get("anchors", {}) or {}
    market = package.get("market_context", {}) or {}

    return {
        "available": True,
        "as_of_date": package.get("metadata", {}).get("as_of_date"),
        "quarter": package.get("metadata", {}).get("quarter"),
        "price_target": computed.get("price_target"),
        "current_share_price": computed.get("current_share_price"),
        "upside_downside_pct": computed.get("upside_downside_pct"),
        "growth_rate_pct": _pct_text(assumptions.get("growth_rate")),
        "wacc_pct": _pct_text(assumptions.get("wacc")),
        "terminal_growth_pct": _pct_text(assumptions.get("terminal_growth")),
        "revenue_ttm_estimated_musd": _money_text(anchors.get("revenue_ttm_estimated_musd")),
        "ebitda_ttm_estimated_musd": _money_text(anchors.get("ebitda_ttm_estimated_musd")),
        "free_cash_flow_base_musd": _money_text(anchors.get("free_cash_flow_base_less_sbc_musd") or anchors.get("free_cash_flow_base_musd")),
        "net_cash_musd": _money_text(anchors.get("net_cash_musd")),
        "diluted_shares_m": anchors.get("diluted_shares_m"),
        "market_source": market.get("market_source"),
        "market_timestamp": market.get("market_timestamp"),
        "assumption_notes": package.get("assumption_notes") or [],
    }


def run_report_quality_checks(reference_date: str = "") -> dict[str, Any]:
    """Run lightweight QA checks on workbook structure, state persistence, and missing signals."""
    findings: list[dict[str, Any]] = []
    workbook_overview = get_workbook_overview()

    if not workbook_overview.get("available"):
        findings.append(
            {
                "severity": "high",
                "title": "Workbook principal absent",
                "detail": "Le rapport historique principal n'existe pas encore sur disque.",
            }
        )
    else:
        visible_sheets = set(workbook_overview.get("visible_sheets") or [])
        hidden_sheets = set(workbook_overview.get("hidden_sheets") or [])
        missing_visible = [sheet for sheet in REQUIRED_VISIBLE_SHEETS if sheet not in visible_sheets]
        for sheet in missing_visible:
            findings.append(
                {
                    "severity": "high",
                    "title": "Onglet visible manquant",
                    "detail": f"L'onglet requis '{sheet}' n'apparait pas dans le workbook visible.",
                }
            )

        raw_not_hidden = [sheet for sheet in EXPECTED_HIDDEN_RAW_SHEETS if sheet not in hidden_sheets]
        for sheet in raw_not_hidden:
            findings.append(
                {
                    "severity": "medium",
                    "title": "Feuille raw non masquée",
                    "detail": f"La feuille technique '{sheet}' devrait rester cachée.",
                }
            )

        if int(workbook_overview.get("summary_history_rows") or 0) < 10:
            findings.append(
                {
                    "severity": "medium",
                    "title": "Historique quotidien mince",
                    "detail": "Suivi Quotidien contient moins de 10 lignes d'historique, ce qui fragilise les lectures de tendance.",
                }
            )

    signal_package = _load_financial_signal_package(reference_date)
    if not signal_package:
        findings.append(
            {
                "severity": "high",
                "title": "Signaux financiers indisponibles",
                "detail": "Aucun paquet de signaux financiers n'a pu être chargé.",
            }
        )
    elif not FINANCIAL_SIGNALS_HISTORY_FILE.exists():
        findings.append(
            {
                "severity": "medium",
                "title": "Historique des signaux absent",
                "detail": "financial_signals_history.csv est manquant; les lectures 7j/30j seront moins robustes.",
            }
        )

    nowcast_package = _load_quarterly_nowcast_package(reference_date)
    if not nowcast_package:
        findings.append(
            {
                "severity": "high",
                "title": "Nowcast trimestriel indisponible",
                "detail": "Aucun paquet de nowcast trimestriel n'a pu être chargé.",
            }
        )
    else:
        current_quarter = str(nowcast_package.get("metadata", {}).get("current_quarter") or "")
        for snapshot in nowcast_package.get("historical_snapshots") or []:
            quarter = str(snapshot.get("quarter") or "")
            if quarter and current_quarter and quarter < current_quarter and not bool(snapshot.get("snapshot_locked")):
                findings.append(
                    {
                        "severity": "medium",
                        "title": "Trimestre historique non figé",
                        "detail": f"Le trimestre {quarter} reste marqué 'En cours' alors qu'un trimestre plus récent existe déjà.",
                    }
                )

        if not QUARTERLY_LOCKED_ARCHIVE_FILE.exists():
            findings.append(
                {
                    "severity": "medium",
                    "title": "Archive trimestrielle absente",
                    "detail": "quarterly_nowcast_locked_archive.csv n'existe pas encore; la mémoire des trimestres clos reste fragile.",
                }
            )
        if QUARTERLY_SNAPSHOTS_FILE.exists():
            try:
                snapshots_df = pd.read_csv(QUARTERLY_SNAPSHOTS_FILE)
            except Exception:
                snapshots_df = pd.DataFrame()
            if not snapshots_df.empty and "observed_days" in snapshots_df.columns:
                bad_rows = snapshots_df[
                    pd.to_numeric(snapshots_df["observed_days"], errors="coerce").fillna(0) <= 0
                ]
                if not bad_rows.empty:
                    findings.append(
                        {
                            "severity": "medium",
                            "title": "Snapshot sans jours observés",
                            "detail": "Au moins un snapshot trimestriel a 0 jour observé, ce qui indique un état corrompu ou incomplet.",
                        }
                    )

        model_output = nowcast_package.get("model_output", {}) or {}
        revenue_prob = model_output.get("revenue_beat_probability")
        revenue_estimate = model_output.get("estimated_revenue_musd")
        if revenue_prob not in (None, "") and revenue_estimate in (None, 0, 0.0):
            findings.append(
                {
                    "severity": "medium",
                    "title": "Probabilité sans estimation",
                    "detail": "Le nowcast expose une probabilité revenus mais pas de montant de revenus estimé exploitable.",
                }
            )

    alt_package = _load_alternative_data_package()
    signal_count = int((alt_package.get("metadata") or {}).get("signal_count") or 0)
    if signal_count < 5:
        findings.append(
            {
                "severity": "medium",
                "title": "Alternative Data trop maigre",
                "detail": f"L'onglet Alternative Data n'a que {signal_count} signaux stables; il gagnerait à rester au-dessus de 5.",
            }
        )
    missing_social = alt_package.get("missing_social_signals") or []
    if missing_social:
        findings.append(
            {
                "severity": "low",
                "title": "Followers sociaux incomplets",
                "detail": f"Les signaux sociaux manquants aujourd'hui sont : {', '.join(missing_social)}.",
            }
        )

    dcf_package = _load_dcf_package(reference_date)
    if dcf_package:
        dcf_context = get_dcf_context(reference_date)
        if dcf_context.get("current_share_price") in (None, "N/D"):
            findings.append(
                {
                    "severity": "low",
                    "title": "Cours actuel absent dans la DCF",
                    "detail": "La DCF n'a pas de cours actuel exploitable, donc l'upside/downside reste partiel.",
                }
            )

    severity_order = {"high": 0, "medium": 1, "low": 2}
    findings = sorted(findings, key=lambda item: (severity_order.get(item["severity"], 9), item["title"]))
    status = "ok" if not findings else ("warning" if all(item["severity"] != "high" for item in findings) else "alert")

    return {
        "checked_at": _today_iso(),
        "status": status,
        "finding_count": len(findings),
        "findings": findings,
        "workbook_overview": workbook_overview,
    }


def get_full_report_context(reference_date: str = "", quarter: str = "") -> dict[str, Any]:
    """Return a full investor-ready context bundle for the latest Duolingo report."""
    return {
        "generated_at": _today_iso(),
        "daily_signals": get_daily_signals_context(reference_date),
        "quarterly_nowcast": get_quarterly_nowcast_context(quarter=quarter, reference_date=reference_date),
        "alternative_data": get_alternative_data_context(),
        "dcf_valuation": get_dcf_context(reference_date),
        "quality_checks": run_report_quality_checks(reference_date),
        "workbook_overview": get_workbook_overview(),
    }


__all__ = [
    "get_workbook_overview",
    "get_daily_signals_context",
    "get_quarterly_nowcast_context",
    "get_alternative_data_context",
    "get_dcf_context",
    "run_report_quality_checks",
    "get_full_report_context",
]
