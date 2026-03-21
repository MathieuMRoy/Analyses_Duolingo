"""Helpers de post-traitement du workbook Excel."""

from datetime import datetime
import numbers

from openpyxl.styles import Alignment, Font


def remove_sheets(wb, sheet_names: list[str] | tuple[str, ...]) -> None:
    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]


def hide_sheets(wb, sheet_names: list[str] | tuple[str, ...]) -> None:
    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_state = "hidden"


def reorder_sheets(wb, ordered_sheet_names: list[str] | tuple[str, ...]) -> None:
    wb._sheets = sorted(
        wb._sheets,
        key=lambda sheet: ordered_sheet_names.index(sheet.title)
        if sheet.title in ordered_sheet_names
        else len(ordered_sheet_names),
    )


def apply_standard_table_style(
    ws,
    sheet_name: str,
    styles: dict[str, object],
    *,
    summary_sheet_name: str,
    glossary_sheet_name: str,
    chart_data_sheet_name: str,
) -> None:
    base_font_name = styles["BASE_FONT_NAME"]
    header_fill = styles["header_fill"]
    header_font = styles["header_font"]
    zebra_fill = styles["zebra_fill"]
    success_fill = styles["success_fill"]
    warning_fill = styles["warning_fill"]
    alert_fill = styles["alert_fill"]
    base_font = styles["base_font"]
    center_align = styles["center_align"]
    left_align = styles["left_align"]
    thin_border = styles["thin_border"]

    ws.sheet_view.showGridLines = False

    if ws.max_row < 1:
        return

    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        is_zebra = row_idx % 2 == 0
        for cell in row:
            cell.border = thin_border
            header_value = ws.cell(1, cell.column).value
            header_key = str(header_value).strip().lower() if header_value is not None else ""

            is_delta = ("Δ" in str(header_value)) or ("delta" in header_key) or ("evol" in header_key)
            is_percent = ("%" in str(header_value)) or any(
                key in header_key for key in ["taux", "attrition", "score", "pénétration", "penetration"]
            )
            is_xp = "xp" in header_key
            is_streak = ("série" in header_key) or ("serie" in header_key) or ("streak" in header_key)
            is_panel = any(key in header_key for key in ["panel", "total", "profils", "actifs", "reactiv", "abandons"])

            # Integer count columns override percentage detection
            if is_panel:
                is_percent = False

            if isinstance(cell.value, numbers.Number) and cell.value != cell.value:
                cell.value = None

            if isinstance(cell.value, datetime):
                cell.alignment = center_align
                cell.font = base_font
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(cell.value, numbers.Number):
                cell.alignment = center_align
                cell.font = base_font

                if is_delta and is_percent:
                    if sheet_name == chart_data_sheet_name:
                        cell.number_format = '+0.0"%" ;-0.0"%" ;0.0"%"'
                    else:
                        cell.number_format = "+0.0%;-0.0%;0.0%"
                elif is_percent:
                    if sheet_name == chart_data_sheet_name:
                        cell.number_format = '0.0"%"'
                    else:
                        cell.number_format = "0.0%"
                elif is_delta and is_xp:
                    cell.number_format = '+#,##0" XP";-#,##0" XP";0" XP"'
                elif is_xp:
                    cell.number_format = '#,##0" XP"'
                elif is_delta:
                    cell.number_format = "+0.0;-0.0;0.0"
                elif is_streak:
                    cell.number_format = "0.0"
                elif is_panel:
                    cell.number_format = "#,##0"

                if is_delta:
                    if cell.value > 0:
                        cell.font = Font(name=base_font_name, size=11, color="008000", bold=True)
                    elif cell.value < 0:
                        cell.font = Font(name=base_font_name, size=11, color="C00000", bold=True)
            elif cell.value is None and is_delta:
                cell.value = "N/A"
                cell.alignment = center_align
                cell.font = base_font
            else:
                cell.alignment = left_align
                cell.font = base_font
                if "définition" in header_key or "definition" in header_key:
                    cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)

            if is_zebra:
                cell.fill = zebra_fill

            if sheet_name == summary_sheet_name and any(key in header_key for key in ["taux", "attrition", "churn"]) and is_percent:
                try:
                    metric_value = float(cell.value)
                    if metric_value <= 0.02:
                        cell.fill = success_fill
                    elif metric_value <= 0.05:
                        cell.fill = warning_fill
                    else:
                        cell.fill = alert_fill
                except Exception:
                    pass
            elif sheet_name == summary_sheet_name and "reactiv" in header_key:
                try:
                    if float(cell.value) > 0:
                        cell.fill = success_fill
                        cell.font = Font(name=base_font_name, size=11, color="008000", bold=True)
                except Exception:
                    pass
            elif sheet_name == summary_sheet_name and "progression" in header_key:
                try:
                    if float(cell.value) > 0:
                        cell.fill = success_fill
                        cell.font = Font(name=base_font_name, size=11, color="008000", bold=True)
                except Exception:
                    pass

    for col in ws.columns:
        max_length = 0
        first_cell = col[0]
        if not hasattr(first_cell, "column_letter"):
            continue
        column = first_cell.column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass

        adjusted_width = max_length + 4
        max_width = 70 if sheet_name == glossary_sheet_name else 60
        ws.column_dimensions[column].width = min(adjusted_width, max_width)

    if ws.max_row > 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
