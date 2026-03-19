"""Style context partage pour le rendu Excel."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def build_style_context() -> dict[str, object]:
    duo_green = "58CC02"
    duo_blue = "1CB0F6"
    navy = "1F4E78"
    light_grey = "F2F2F2"
    green_soft = "E2F0D9"
    amber_soft = "FFF2CC"
    red_soft = "FFC7CE"
    white = "FFFFFF"
    base_font_name = "Calibri"

    return {
        "DUO_GREEN": duo_green,
        "DUO_BLUE": duo_blue,
        "NAVY": navy,
        "LIGHT_GREY": light_grey,
        "GREEN_SOFT": green_soft,
        "AMBER_SOFT": amber_soft,
        "RED_SOFT": red_soft,
        "WHITE": white,
        "BASE_FONT_NAME": base_font_name,
        "header_fill": PatternFill(start_color=navy, end_color=navy, fill_type="solid"),
        "header_font": Font(name=base_font_name, color=white, bold=True, size=11),
        "zebra_fill": PatternFill(start_color=light_grey, end_color=light_grey, fill_type="solid"),
        "success_fill": PatternFill(start_color=green_soft, end_color=green_soft, fill_type="solid"),
        "warning_fill": PatternFill(start_color=amber_soft, end_color=amber_soft, fill_type="solid"),
        "alert_fill": PatternFill(start_color=red_soft, end_color=red_soft, fill_type="solid"),
        "white_fill": PatternFill(start_color=white, end_color=white, fill_type="solid"),
        "base_font": Font(name=base_font_name, size=11, color="000000"),
        "center_align": Alignment(horizontal="center", vertical="center"),
        "left_align": Alignment(horizontal="left", vertical="center", indent=1),
        "thin_border": Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        ),
    }
