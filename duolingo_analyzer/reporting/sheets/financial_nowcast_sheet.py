"""Renderer de la feuille Signaux Financiers."""

from __future__ import annotations

import numbers
import re

from openpyxl.styles import Alignment, Font, PatternFill


def render_financial_nowcast_sheet(
    ws,
    signal_package: dict,
    ia_report: str | None,
    styles: dict[str, object],
    helpers: dict[str, object],
) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    for column_letter, width in {
        "A": 16,
        "B": 16,
        "C": 16,
        "D": 15,
        "E": 15,
        "F": 15,
        "G": 15,
        "H": 15,
    }.items():
        ws.column_dimensions[column_letter].width = width

    metadata = signal_package.get("metadata", {})
    panel = signal_package.get("panel", {})
    business = signal_package.get("business_signals", {})
    daily = signal_package.get("daily_comparison", {})
    horizon = signal_package.get("horizon_context", {})
    proxy = signal_package.get("financial_proxy_signals", {})
    assumptions = signal_package.get("assumptions", [])

    ai_sections = {
        "RESUME": None,
        "TENDANCES": None,
        "ATTENTION": None,
        "CONSEILS": None,
    }
    if ia_report:
        for key in ai_sections.keys():
            match = re.search(rf"\[{key}\](.*?)(?=\[|$)", ia_report, re.DOTALL)
            if match:
                ai_sections[key] = match.group(1).strip()

    duo_green = styles["DUO_GREEN"]
    navy = styles["NAVY"]
    light_grey = styles["LIGHT_GREY"]
    white = styles["WHITE"]
    base_font_name = styles["BASE_FONT_NAME"]
    zebra_fill = styles["zebra_fill"]
    white_fill = styles["white_fill"]
    center_align = styles["center_align"]
    left_align = styles["left_align"]
    thin_border = styles["thin_border"]

    label_signal_bias = helpers["label_signal_bias"]
    label_confidence = helpers["label_confidence"]
    pretty_fr_number = helpers["pretty_fr_number"]
    pretty_ratio_pct = helpers["pretty_ratio_pct"]
    pretty_score = helpers["pretty_score"]
    pretty_delta_pts = helpers["pretty_delta_pts"]
    compact_summary_text = helpers["compact_summary_text"]
    compact_bullet_text = helpers["compact_bullet_text"]

    canvas = "F5F7FA"
    surface = "FFFFFF"
    slate = "E5EDF4"
    deep_teal = "234E52"
    sea = "2E7D74"
    calm_blue = "2F6EA8"
    soft_blue = "EEF5FB"
    soft_green = "EEF7E8"
    soft_red = "FCEBEC"
    soft_slate = "F3F6F9"
    muted = "5C6775"
    ink = "16324F"
    warm = "C48A3A"
    risk = "C65A46"

    def write_box(
        range_ref: str,
        value: str,
        *,
        fill: str = surface,
        font_color: str = ink,
        size: int = 11,
        bold: bool = False,
        align=None,
    ) -> None:
        ws.merge_cells(range_ref)
        cell = ws[range_ref.split(":")[0]]
        cell.value = value
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.font = Font(name=base_font_name, size=size, bold=bold, color=font_color)
        cell.alignment = align or center_align
        cell.border = thin_border

    def write_strip(
        title_range: str,
        value_range: str,
        title: str,
        value: str,
        *,
        title_fill: str,
        value_fill: str,
        value_color: str = ink,
        value_size: int = 16,
    ) -> None:
        write_box(title_range, title, fill=title_fill, font_color=white, size=10, bold=True)
        write_box(value_range, value, fill=value_fill, font_color=value_color, size=value_size, bold=True)

    bias_label = label_signal_bias(proxy.get("signal_bias"))
    confidence_label = label_confidence(proxy.get("confidence_level"))
    coverage_ratio = panel.get("coverage_ratio")
    observed_users = panel.get("observed_users_today")
    target_users = panel.get("target_panel_size")
    premium_net_adds_today = daily.get("premium_net_adds_today")

    drivers = proxy.get("main_drivers") or ["Aucun driver majeur identifie pour l'instant."]
    risks = proxy.get("main_risks") or ["Aucun risque majeur identifie pour l'instant."]

    lead_text = (
        f"Signal {bias_label.lower()} avec confiance {confidence_label.lower()}. "
        f"La monetisation ressort a {pretty_score(proxy.get('monetization_momentum_index'))}, "
        f"tandis que la qualite d'engagement atteint {pretty_score(proxy.get('engagement_quality_index'))}."
    )
    if ai_sections["RESUME"]:
        lead_text = ai_sections["RESUME"]
    elif isinstance(premium_net_adds_today, numbers.Number):
        if float(premium_net_adds_today) > 0:
            lead_text = (
                f"Signal {bias_label.lower()} avec confiance {confidence_label.lower()}. "
                f"La monetisation s'ameliore aujourd'hui avec {pretty_fr_number(premium_net_adds_today, 0)} abonnements nets observables vs hier."
            )
        elif float(premium_net_adds_today) < 0:
            lead_text = (
                f"Signal {bias_label.lower()} avec confiance {confidence_label.lower()}. "
                f"La monetisation se tasse aujourd'hui avec {pretty_fr_number(abs(float(premium_net_adds_today)), 0)} pertes nettes d'abonnements observables vs hier."
            )
    lead_text = compact_summary_text(lead_text, max_sentences=2, max_chars=220, separator="\n")

    trend_text = compact_bullet_text(
        ai_sections["TENDANCES"] or "\n".join(f"- {item}" for item in drivers),
        max_items=2,
        max_chars=220,
    )
    attention_text = compact_bullet_text(
        ai_sections["ATTENTION"] or "\n".join(f"- {item}" for item in risks),
        max_items=2,
        max_chars=220,
    )

    closing_text = compact_summary_text(
        ai_sections["CONSEILS"]
        or "Le signal du jour doit surtout etre lu comme un briefing de tendance : on surveille la qualite d'engagement, la pression de churn et la dynamique premium.",
        max_sentences=2,
        max_chars=180,
        separator="\n",
    )
    week_text = compact_summary_text(
        (horizon.get("seven_day") or {}).get("summary_text"),
        max_sentences=2,
        max_chars=220,
        separator="\n",
    )
    month_text = compact_summary_text(
        (horizon.get("thirty_day") or {}).get("summary_text"),
        max_sentences=2,
        max_chars=220,
        separator="\n",
    )

    bias_fill = {
        "Favorable": duo_green,
        "Neutre": warm,
        "Defavorable": risk,
    }.get(bias_label, deep_teal)
    confidence_fill = {
        "Elevee": duo_green,
        "Moyenne": warm,
        "Faible": "FF8A65",
    }.get(confidence_label, calm_blue)

    summary_line = (
        f"Date de reference : {metadata.get('as_of_date', 'N/D')} | "
        f"Panel observe : {pretty_fr_number(observed_users, 0)} / {pretty_fr_number(target_users, 0)} "
        f"({pretty_ratio_pct(coverage_ratio, 1)}) | "
        f"Signal : {bias_label} | Confiance : {confidence_label}"
    )

    write_box("A1:H2", "SIGNAUX FINANCIERS", fill=deep_teal, font_color=white, size=18, bold=True)
    write_box(
        "A3:H3",
        summary_line,
        fill=canvas,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_box("A5:C5", "Briefing du jour", fill=navy, font_color=white, size=11, bold=True)
    write_box(
        "A6:C10",
        lead_text,
        fill=surface,
        font_color=ink,
        size=12,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_strip("D5:E5", "D6:E7", "Signal du jour", bias_label, title_fill=bias_fill, value_fill=surface, value_size=18)
    write_strip("D8:E8", "D9:E10", "Confiance", confidence_label, title_fill=confidence_fill, value_fill=soft_slate, value_size=16)

    write_strip(
        "F5:H5",
        "F6:H7",
        "Empreinte du panel",
        pretty_ratio_pct(coverage_ratio, 1),
        title_fill=calm_blue,
        value_fill=surface,
        value_size=20,
    )
    write_box(
        "F8:H8",
        f"{pretty_fr_number(observed_users, 0)} profils observes",
        fill=soft_blue,
        font_color=ink,
        size=10,
        bold=True,
    )
    write_box(
        "F9:H9",
        f"Cible suivie : {pretty_fr_number(target_users, 0)} profils",
        fill=soft_blue,
        font_color=muted,
        size=9,
    )
    write_box(
        "F10:H10",
        "Lecture quotidienne du panel, pas une projection trimestrielle.",
        fill=canvas,
        font_color=muted,
        size=9,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_box("A12:H12", "Tableau de bord du jour", fill=calm_blue, font_color=white, size=11, bold=True)

    write_strip(
        "A13:B13",
        "A14:B15",
        "Monetisation",
        pretty_score(proxy.get("monetization_momentum_index")),
        title_fill=sea,
        value_fill=soft_green,
    )
    write_box(
        "A16:B16",
        "Indice composite de monetisation",
        fill=canvas,
        font_color=muted,
        size=9,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_strip(
        "C13:D13",
        "C14:D15",
        "Engagement",
        pretty_score(proxy.get("engagement_quality_index")),
        title_fill=calm_blue,
        value_fill=soft_blue,
    )
    write_box(
        "C16:D16",
        "Qualite d'engagement recente",
        fill=canvas,
        font_color=muted,
        size=9,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_strip(
        "E13:F13",
        "E14:F15",
        "Premium 14j",
        pretty_delta_pts(proxy.get("premium_momentum_14d")),
        title_fill=navy,
        value_fill=soft_slate,
    )
    write_box(
        "E16:F16",
        "Variation recente du taux Super",
        fill=canvas,
        font_color=muted,
        size=9,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    churn_value = proxy.get("churn_trend_14d")
    churn_accent = duo_green if isinstance(churn_value, numbers.Number) and churn_value <= 0 else risk
    churn_fill = soft_green if isinstance(churn_value, numbers.Number) and churn_value <= 0 else soft_red
    write_strip(
        "G13:H13",
        "G14:H15",
        "Churn 14j",
        pretty_delta_pts(churn_value),
        title_fill=churn_accent,
        value_fill=churn_fill,
    )
    write_box(
        "G16:H16",
        "Pression recente sur l'abandon",
        fill=canvas,
        font_color=muted,
        size=9,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_box("A18:D18", "Ce qui s'ameliore", fill=duo_green, font_color=white, size=11, bold=True)
    write_box("E18:H18", "Ce qui se degrade", fill=risk, font_color=white, size=11, bold=True)
    write_box(
        "A19:D23",
        trend_text,
        fill=soft_green,
        font_color=ink,
        size=10,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )
    write_box(
        "E19:H23",
        attention_text,
        fill=soft_red,
        font_color=ink,
        size=10,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_box("A25:D25", "Lecture 7 jours", fill=navy, font_color=white, size=11, bold=True)
    write_box("E25:H25", "Lecture 30 jours", fill=calm_blue, font_color=white, size=11, bold=True)
    write_box(
        "A26:D28",
        week_text,
        fill=soft_slate,
        font_color=ink,
        size=10,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )
    write_box(
        "E26:H28",
        month_text,
        fill=soft_blue,
        font_color=ink,
        size=10,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_box("A30:H30", "Conclusion operationnelle", fill=deep_teal, font_color=white, size=11, bold=True)
    write_box(
        "A31:H33",
        closing_text,
        fill=canvas,
        font_color=ink,
        size=10,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_box("A35:H35", "Diagnostics du modele", fill=navy, font_color=white, size=11, bold=True)
    model_rows = [
        ("Reactivation 7j", pretty_delta_pts(proxy.get("reactivation_trend_7d"))),
        ("Retention high-value", pretty_delta_pts(proxy.get("high_value_retention_trend"))),
        ("Super rate", pretty_ratio_pct(business.get("super_rate"), 1)),
        ("Progression Debutants -> Standard", pretty_ratio_pct(business.get("debutants_to_standard_rate"), 1)),
        ("Abandon Debutants", pretty_ratio_pct(business.get("debutants_abandon_rate"), 1)),
        ("Abandon Standard", pretty_ratio_pct(business.get("standard_abandon_rate"), 1)),
        ("Abandon Super-Actifs", pretty_ratio_pct(business.get("super_actifs_abandon_rate"), 1)),
        (
            "Readiness",
            "Lecture quotidienne prete; les probabilites trimestrielles se lisent desormais dans l'onglet Nowcast Trimestriel.",
        ),
    ]

    row_cursor = 36
    for label, value in model_rows:
        row_fill = zebra_fill if row_cursor % 2 == 1 else white_fill
        ws[f"A{row_cursor}"] = label
        ws[f"A{row_cursor}"].fill = row_fill
        ws[f"A{row_cursor}"].font = Font(name=base_font_name, size=10, bold=True, color=ink)
        ws[f"A{row_cursor}"].alignment = left_align
        ws[f"A{row_cursor}"].border = thin_border
        ws.merge_cells(f"B{row_cursor}:H{row_cursor}")
        value_cell = ws[f"B{row_cursor}"]
        value_cell.value = value
        value_cell.fill = row_fill
        value_cell.font = Font(name=base_font_name, size=10, color=ink)
        value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        value_cell.border = thin_border
        for merged_col in range(3, 9):
            ws.cell(row=row_cursor, column=merged_col).border = thin_border
            ws.cell(row=row_cursor, column=merged_col).fill = row_fill
        row_cursor += 1

    hypotheses_row = row_cursor + 1
    write_box(f"A{hypotheses_row}:H{hypotheses_row}", "Hypotheses & notes", fill=navy, font_color=white, size=11, bold=True)
    assumptions_text = "\n".join(f"- {item}" for item in assumptions) if assumptions else "- Aucune hypothese specifique."
    write_box(
        f"A{hypotheses_row + 1}:H{hypotheses_row + 4}",
        assumptions_text,
        fill=light_grey,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    ws.row_dimensions[3].height = 24
    ws.row_dimensions[5].height = 24
    for row_idx in range(6, 11):
        ws.row_dimensions[row_idx].height = 30 if row_idx != 6 else 40
    ws.row_dimensions[12].height = 24
    ws.row_dimensions[13].height = 22
    ws.row_dimensions[14].height = 26
    ws.row_dimensions[15].height = 26
    ws.row_dimensions[16].height = 22
    ws.row_dimensions[18].height = 24
    for row_idx in range(19, 24):
        ws.row_dimensions[row_idx].height = 34
    ws.row_dimensions[25].height = 24
    for row_idx in range(26, 29):
        ws.row_dimensions[row_idx].height = 32
    ws.row_dimensions[30].height = 24
    for row_idx in range(31, 34):
        ws.row_dimensions[row_idx].height = 30
    ws.row_dimensions[35].height = 24
    for row_idx in range(36, row_cursor):
        ws.row_dimensions[row_idx].height = 22
    for row_idx in range(hypotheses_row + 1, hypotheses_row + 5):
        ws.row_dimensions[row_idx].height = 22
