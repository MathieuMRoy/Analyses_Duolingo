"""Construction et rendu de la feuille dictionnaire des KPIs."""

from __future__ import annotations

import math

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


def build_kpi_dictionary_df() -> pd.DataFrame:
    daily_rows = [
        {
            "Section": "Lecture quotidienne",
            "KPI": "Moyenne Streak (Jours)",
            "Lecture utile": "Longueur moyenne de la série de jours consécutifs d'utilisation. Mesure la fidélité du panel.",
            "Methode / calcul": "Moyenne simple du streak observé sur le panel du jour.",
        },
        {
            "Section": "Lecture quotidienne",
            "KPI": "Apprentissage (XP/j)",
            "Lecture utile": "Gain moyen de points d'expérience depuis la veille. Mesure l'effort d'apprentissage quotidien.",
            "Methode / calcul": "Delta de TotalXP entre hier et aujourd'hui, borné à 0 minimum, puis moyenne sur le panel.",
        },
        {
            "Section": "Lecture quotidienne",
            "KPI": "Taux Abonn. Super",
            "Lecture utile": "Part observable du panel premium via le signal hasPlus. Tant que MAX n'est pas détecté de façon fiable, une partie de MAX peut rester incluse ici.",
            "Methode / calcul": "Profils premium observables / profils observés dans le panel.",
        },
        {
            "Section": "Lecture quotidienne",
            "KPI": "Taux d'Abandon Global",
            "Lecture utile": "Part des utilisateurs actifs hier qui ne le sont plus aujourd'hui.",
            "Methode / calcul": "Utilisateurs avec streak > 0 hier et streak = 0 aujourd'hui / utilisateurs actifs hier.",
        },
        {
            "Section": "Lecture quotidienne",
            "KPI": "Reactivations vs Veille",
            "Lecture utile": "Nombre d'utilisateurs inactifs hier redevenus actifs aujourd'hui.",
            "Methode / calcul": "Comptage des profils avec streak = 0 hier puis streak > 0 aujourd'hui.",
        },
        {
            "Section": "Lecture quotidienne",
            "KPI": "Score d'Engagement",
            "Lecture utile": "Part des profils observés qui restent actifs aujourd'hui.",
            "Methode / calcul": "Profils avec streak > 0 / panel observé du jour.",
        },
    ]

    transition_rows = [
        {
            "Section": "Transitions et churn",
            "KPI": "Progression Debutants vers Standard",
            "Lecture utile": "Part des Débutants actifs hier qui ont progressé vers Standard aujourd'hui sans abandonner.",
            "Methode / calcul": "Débutants actifs hier devenus Standard aujourd'hui / Débutants actifs hier.",
        },
        {
            "Section": "Transitions et churn",
            "KPI": "Abandon Debutants",
            "Lecture utile": "Vrai abandon chez les Débutants, mesuré sur la cohorte d'hier et non sur la cohorte du jour.",
            "Methode / calcul": "Débutants actifs hier devenus inactifs aujourd'hui / Débutants actifs hier.",
        },
        {
            "Section": "Transitions et churn",
            "KPI": "Abandon Standard",
            "Lecture utile": "Vrai abandon chez les Standard, mesuré sur la cohorte d'hier.",
            "Methode / calcul": "Standard actifs hier devenus inactifs aujourd'hui / Standard actifs hier.",
        },
        {
            "Section": "Transitions et churn",
            "KPI": "Abandon Super-Actifs",
            "Lecture utile": "Vrai abandon chez les Super-Actifs, mesuré sur la cohorte d'hier.",
            "Methode / calcul": "Super-Actifs actifs hier devenus inactifs aujourd'hui / Super-Actifs actifs hier.",
        },
    ]

    quarterly_rows = [
        {
            "Section": "Nowcast trimestriel",
            "KPI": "Confiance (trimestriel)",
            "Lecture utile": "Niveau de fiabilité du nowcast trimestriel.",
            "Methode / calcul": "Élevée si la couverture moyenne est >= 75% et si les jours observés sont >= 20 ; moyenne si la couverture est >= 45% et si les jours observés sont >= 10 ; sinon faible.",
        },
        {
            "Section": "Nowcast trimestriel",
            "KPI": "Score trimestre",
            "Lecture utile": "Score synthétique de 0 à 100 du trimestre, utilisé pour lire le biais global du nowcast.",
            "Methode / calcul": "Score composé à partir de la monétisation, de l'engagement, du momentum Super, du churn, des réactivations et de la rétention high-value, puis ajusté par un breadth factor qui pénalise les trimestres encore jeunes.",
        },
        {
            "Section": "Nowcast trimestriel",
            "KPI": "Prob. beat revenus",
            "Lecture utile": "Probabilité implicite de battre les revenus du trimestre. En façade, c'est la lecture principale du nowcast revenus.",
            "Methode / calcul": "On part d'un score revenus, converti en probabilité implicite. Cette probabilité reste guidance-first : elle se lit face au benchmark management du trimestre tant qu'un consensus complet n'est pas disponible.",
        },
        {
            "Section": "Nowcast trimestriel",
            "KPI": "Prob. beat EBITDA",
            "Lecture utile": "Probabilité implicite de battre l'EBITDA du trimestre.",
            "Methode / calcul": "Transformation score -> probabilité appliquée à un score EBITDA fondé sur la monétisation, l'engagement, le churn, les réactivations et la rétention high-value.",
        },
        {
            "Section": "Nowcast trimestriel",
            "KPI": "Prob. guidance raise",
            "Lecture utile": "Probabilité implicite d'un relèvement de guidance pour le trimestre suivant.",
            "Methode / calcul": "Transformation score -> probabilité appliquée à un score guidance surtout piloté par la monétisation, le momentum Super, l'engagement, le churn et les réactivations.",
        },
        {
            "Section": "Nowcast trimestriel",
            "KPI": "Estimation revenus trimestrielle",
            "Lecture utile": "Estimation interne des revenus du trimestre, affichée en grand dans le nowcast.",
            "Methode / calcul": "Si une guidance revenus de référence existe, estimation = guidance x (1 + beat implicite). Le beat implicite part du beat historique médian et est ajusté par la probabilité revenus, avec des bornes de prudence. Sinon, fallback sur le revenu réel du trimestre précédent avec une croissance QoQ implicite bornée.",
        },
        {
            "Section": "Nowcast trimestriel",
            "KPI": "Est. X M$ vs guidance Y M$",
            "Lecture utile": "Lecture directe de l'estimation revenus par rapport au benchmark management du trimestre.",
            "Methode / calcul": "Le chiffre de gauche est l'estimation du modèle. Le chiffre de droite est la guidance revenus du management pour le trimestre suivi, lorsqu'elle est disponible dans l'historique trimestriel.",
        },
        {
            "Section": "Nowcast trimestriel",
            "KPI": "EBITDA estime",
            "Lecture utile": "Estimation interne de l'EBITDA ajusté du trimestre.",
            "Methode / calcul": "EBITDA estimé = revenus estimés x marge EBITDA implicite. La marge implicite part de la marge historique médiane et est ajustée par la probabilité EBITDA, avec des bornes conservatrices.",
        },
        {
            "Section": "Nowcast trimestriel",
            "KPI": "Guide N+1 estime",
            "Lecture utile": "Estimation interne de la guidance revenus du trimestre suivant.",
            "Methode / calcul": "Guide N+1 estimé = revenus estimés x ratio implicite. Le ratio part du ratio historique médian guidance suivante / revenus et est ajusté par la probabilité guidance raise.",
        },
        {
            "Section": "Nowcast trimestriel",
            "KPI": "Historique trimestriel fige",
            "Lecture utile": "Mémoire du nowcast tel qu'il était à la date du snapshot, sans être écrasé une fois le trimestre terminé.",
            "Methode / calcul": "Chaque trimestre garde son snapshot final. Quand un nouveau trimestre commence, le précédent reste figé pour permettre le backtest et la comparaison estimate vs réel.",
        },
        {
            "Section": "Nowcast trimestriel",
            "KPI": "Cadre du modele",
            "Lecture utile": "Bloc qui explique la référence utilisée et le niveau de maturité du modèle trimestriel.",
            "Methode / calcul": "Résumé du statut du snapshot, de la référence guidance revenus, du niveau de supervision disponible et de la prochaine étape de calibration.",
        },
    ]

    return pd.DataFrame(daily_rows + transition_rows + quarterly_rows)


def render_kpi_dictionary_sheet(
    ws,
    wb,
    raw_sheet_name: str,
    styles: dict[str, object],
) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A12"

    for column_letter, width in {
        "A": 15,
        "B": 15,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 20,
        "G": 20,
        "H": 20,
    }.items():
        ws.column_dimensions[column_letter].width = width

    base_font_name = styles["BASE_FONT_NAME"]
    center_align = styles["center_align"]
    left_align = styles["left_align"]
    thin_border = styles["thin_border"]

    ink = "1E2A36"
    muted = "5D6978"
    paper = "FFFFFF"
    canvas = "F6F4EF"
    stone = "EEE8DD"
    slate = "E9EFF4"
    title = "2E4053"
    bronze = "9E6B2D"
    moss = "4F6F52"
    plum = "6C5B7B"
    warm_blue = "345C7C"
    soft_green = "EEF5ED"
    soft_warm = "FBF4E8"
    soft_plum = "F4EFF7"
    soft_blue = "EEF5FB"

    section_styles = {
        "Lecture quotidienne": {"fill": warm_blue, "soft": soft_blue},
        "Transitions et churn": {"fill": moss, "soft": soft_green},
        "Nowcast trimestriel": {"fill": bronze, "soft": soft_warm},
    }

    def write_box(
        range_ref: str,
        value: str,
        *,
        fill: str = paper,
        font_color: str = ink,
        size: int = 11,
        bold: bool = False,
        align=None,
    ) -> None:
        ws.merge_cells(range_ref)
        cell = ws[range_ref.split(":")[0]]
        cell.value = value
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.font = Font(name=base_font_name, size=size, bold=bold, color=font_color)
        cell.alignment = align or center_align
        cell.border = thin_border

    def estimate_lines(*values: str) -> int:
        total = 0
        for value in values:
            text = str(value or "")
            total += max(1, math.ceil(len(text) / 48))
        return total

    rows = build_kpi_dictionary_df().to_dict("records")
    available_sections = ["Toutes"]
    for row in rows:
        section = str(row.get("Section") or "").strip()
        if section and section not in available_sections:
            available_sections.append(section)

    raw_ws = wb[raw_sheet_name] if raw_sheet_name in wb.sheetnames else None
    raw_headers = {}
    if raw_ws and raw_ws.max_row >= 1:
        raw_headers = {
            str(cell.value): idx
            for idx, cell in enumerate(raw_ws[1], start=1)
            if cell.value
        }

    def _formula_text_literal(value: str) -> str:
        return '"' + str(value).replace('"', '""') + '"'

    def _filter_expr(column_label: str, row_number: int) -> str:
        if not raw_ws or column_label not in raw_headers:
            return '""'
        section_col = chr(64 + raw_headers["Section"])
        value_col = chr(64 + raw_headers[column_label])
        criteria = (
            f'IF($C$9="Toutes",LEN(\'{raw_sheet_name}\'!$B$2:$B$999)>0,'
            f'\'{raw_sheet_name}\'!${section_col}$2:${section_col}$999=$C$9)'
        )
        return (
            f'IFERROR(INDEX(FILTER(\'{raw_sheet_name}\'!${value_col}$2:${value_col}$999,{criteria}),'
            f'ROWS($A$12:A{row_number})),"")'
        )

    count_formula = (
        '=IF($C$9="Toutes",'
        f'COUNTA(\'{raw_sheet_name}\'!$B$2:$B$999),'
        f'COUNTIF(\'{raw_sheet_name}\'!$A$2:$A$999,$C$9))'
    )

    write_box("A1:H2", "GUIDE DES KPIs", fill=title, font_color="FFFFFF", size=18, bold=True)
    write_box(
        "A3:H3",
        "Cette page sert de mode d'emploi. Lisez d'abord la colonne de gauche pour comprendre le sens du KPI, puis utilisez la méthode / calcul seulement si vous voulez vérifier le détail.",
        fill=canvas,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_box("A5:C7", "Commencez par les KPI du Résumé Financier Q1 et de Signaux Financiers. Revenez ici seulement lorsqu'un indicateur vous pose question.", fill=soft_blue, font_color=ink, size=11, bold=False, align=Alignment(horizontal="left", vertical="top", wrap_text=True))
    write_box("D5:F7", "Ordre de lecture conseillé : 1. Lecture quotidienne  2. Transitions et churn  3. Nowcast trimestriel.", fill=soft_green, font_color=ink, size=11, bold=False, align=Alignment(horizontal="left", vertical="top", wrap_text=True))
    write_box("G5:H7", "Astuce : si vous voulez aller vite, ne lisez que la colonne 'Lecture utile'.", fill=soft_plum, font_color=ink, size=11, bold=False, align=Alignment(horizontal="left", vertical="top", wrap_text=True))

    write_box("A9:B9", "Vue", fill=stone, font_color=ink, size=11, bold=True)
    write_box("C9", "Toutes", fill=paper, font_color=ink, size=11, bold=True)
    write_box(
        "D9:F9",
        '=IF($C$9="Toutes","Vue complète du dictionnaire",$C$9)',
        fill=soft_blue,
        font_color=ink,
        size=10,
        bold=True,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )
    write_box(
        "G9:H9",
        '="KPI affichés : "&' + count_formula[1:],
        fill=soft_warm,
        font_color=muted,
        size=10,
        bold=True,
    )

    if available_sections:
        validation = DataValidation(
            type="list",
            formula1=_formula_text_literal(",".join(available_sections)),
            allow_blank=False,
        )
        validation.promptTitle = "Section du dictionnaire"
        validation.prompt = "Choisissez la section du dictionnaire à afficher."
        ws.add_data_validation(validation)
        validation.add(ws["C9"])

    write_box(
        "A10:H10",
        '=IF($C$9="Toutes","Vue complète","Section : "&$C$9)',
        fill=plum,
        font_color="FFFFFF",
        size=11,
        bold=True,
        align=Alignment(horizontal="left", vertical="center", indent=1),
    )
    ws.row_dimensions[10].height = 22

    write_box("A11:B11", "KPI", fill=stone, font_color=ink, size=11, bold=True)
    write_box("C11:E11", "Lecture utile", fill=stone, font_color=ink, size=11, bold=True)
    write_box("F11:H11", "Methode / calcul", fill=stone, font_color=ink, size=11, bold=True)

    max_display_rows = len(rows) + 2
    for offset in range(max_display_rows):
        row_number = 12 + offset
        write_box(
            f"A{row_number}:B{row_number}",
            "=" + _filter_expr("KPI", row_number),
            fill=paper,
            font_color=ink,
            size=11,
            bold=True,
            align=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1),
        )
        write_box(
            f"C{row_number}:E{row_number}",
            "=" + _filter_expr("Lecture utile", row_number),
            fill=slate,
            font_color=ink,
            size=10,
            align=Alignment(horizontal="left", vertical="top", wrap_text=True),
        )
        write_box(
            f"F{row_number}:H{row_number}",
            "=" + _filter_expr("Methode / calcul", row_number),
            fill=paper,
            font_color=muted,
            size=10,
            align=Alignment(horizontal="left", vertical="top", wrap_text=True),
        )
        ws.row_dimensions[row_number].height = 52

    footer_row = 12 + max_display_rows + 1
    write_box(
        f"A{footer_row}:H{footer_row}",
        "Le dictionnaire reste pédagogique : il explique d'abord l'usage du KPI, puis sa mécanique. Utilisez le sélecteur de vue pour ne consulter que la famille d'indicateurs qui vous intéresse.",
        fill=canvas,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )
    ws.row_dimensions[footer_row].height = 34
