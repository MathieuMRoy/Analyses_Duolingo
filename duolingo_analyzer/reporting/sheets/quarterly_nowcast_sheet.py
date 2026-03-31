"""Renderer de la feuille Nowcast Trimestriel."""

from __future__ import annotations

import re

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def render_quarterly_nowcast_sheet(
    ws,
    package: dict,
    wb,
    raw_sheet_name: str,
    styles: dict[str, object],
    ia_report: str | None = None,
    helpers: dict[str, object] | None = None,
) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    metadata = package.get("metadata", {})
    readiness = package.get("labels_readiness", {})
    assumptions = package.get("assumptions", [])
    historical = package.get("historical_snapshots", [])[-8:]

    raw_ws = wb[raw_sheet_name] if raw_sheet_name in wb.sheetnames else None
    raw_headers = {}
    if raw_ws and raw_ws.max_row >= 1:
        raw_headers = {
            str(cell.value): get_column_letter(idx)
            for idx, cell in enumerate(raw_ws[1], start=1)
            if cell.value
        }

    available_quarters: list[str] = []
    for quarter in metadata.get("available_quarters") or []:
        quarter_label = str(quarter).strip()
        if quarter_label and quarter_label not in available_quarters:
            available_quarters.append(quarter_label)
    if not available_quarters:
        for snapshot in historical:
            quarter_label = str(snapshot.get("quarter") or "").strip()
            if quarter_label and quarter_label not in available_quarters:
                available_quarters.append(quarter_label)

    default_quarter = str(
        metadata.get("default_selected_quarter")
        or metadata.get("current_quarter")
        or (available_quarters[-1] if available_quarters else "N/D")
    )
    if available_quarters and default_quarter not in available_quarters:
        default_quarter = available_quarters[-1]

    quarter_column = raw_headers.get("quarter")
    raw_sheet_ref = f"'{raw_sheet_name}'"

    navy = styles["NAVY"]
    white = styles["WHITE"]
    base_font_name = styles["BASE_FONT_NAME"]
    center_align = styles["center_align"]
    left_align = styles["left_align"]
    thin_border = styles["thin_border"]

    helpers = helpers or {}
    compact_summary_text = helpers.get(
        "compact_summary_text",
        lambda text, max_sentences=2, max_chars=180, separator="\n": str(text or "N/D"),
    )
    compact_bullet_text = helpers.get(
        "compact_bullet_text",
        lambda text, max_items=2, max_chars=120: str(text or "-"),
    )

    ai_sections = {
        "MODELE": None,
        "MODELE_TENDANCES": None,
        "MODELE_RISQUES": None,
        "TENDANCES": None,
        "ATTENTION": None,
    }
    if ia_report:
        for key in ai_sections.keys():
            match = re.search(rf"\[{key}\](.*?)(?=\[|$)", ia_report, re.DOTALL)
            if match:
                ai_sections[key] = match.group(1).strip()

    canvas = "F4F7FA"
    paper = "FFFFFF"
    slate = "E6EDF4"
    blue_section = "2A5B84"
    ink = "16324F"
    muted = "5E6A79"
    soft_blue = "EEF5FB"
    soft_green = "EEF7E8"
    soft_red = "FCEBEC"
    soft_amber = "FBF3E3"
    soft_slate = "F3F6F9"
    green = "2E7D5A"
    red = "C65A46"
    orange = "C98A3D"

    for column_letter, width in {
        "A": 17,
        "B": 17,
        "C": 17,
        "D": 17,
        "E": 17,
        "F": 17,
        "G": 15,
        "H": 15,
    }.items():
        ws.column_dimensions[column_letter].width = width

    def _formula_text_literal(value: str) -> str:
        return '"' + str(value).replace('"', '""') + '"'

    def _formula_multiline_literal(value: str) -> str:
        lines = [line for line in str(value).splitlines() if line.strip()]
        if not lines:
            return '""'
        parts = [_formula_text_literal(lines[0])]
        for line in lines[1:]:
            parts.append("CHAR(10)")
            parts.append(_formula_text_literal(line))
        return "&".join(parts)

    def _raw_lookup_formula(header: str, fallback: str = '"N/D"', quarter_ref: str = "$C$4") -> str:
        if not raw_ws or not quarter_column or header not in raw_headers:
            return f"={fallback}"
        target_column = raw_headers[header]
        return (
            f'=IFERROR(INDEX('
            f'{raw_sheet_ref}!${target_column}:${target_column},'
            f'MATCH({quarter_ref},{raw_sheet_ref}!${quarter_column}:${quarter_column},0)'
            f'),{fallback})'
        )

    def _raw_lookup_expr(header: str, fallback: str = '"N/D"', quarter_ref: str = "$C$4") -> str:
        formula = _raw_lookup_formula(header, fallback=fallback, quarter_ref=quarter_ref)
        return formula[1:] if formula.startswith("=") else formula

    def _ai_or_lookup_formula(ai_text: str, raw_header: str, fallback: str = '"N/D"') -> str:
        if not ai_text or ai_text in {"-", "N/D"}:
            return _raw_lookup_formula(raw_header, fallback=fallback)
        return (
            f'=IF($C$4={_formula_text_literal(default_quarter)},'
            f'{_formula_multiline_literal(ai_text)},'
            f'{_raw_lookup_expr(raw_header, fallback=fallback)})'
        )

    def write_box(
        range_ref: str,
        value: object,
        *,
        fill: str = paper,
        font_color: str = ink,
        size: int = 11,
        bold: bool = False,
        align=None,
        number_format: str | None = None,
    ) -> None:
        if ":" in range_ref:
            ws.merge_cells(range_ref)
            cell = ws[range_ref.split(":")[0]]
        else:
            cell = ws[range_ref]
        cell.value = value
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.font = Font(name=base_font_name, size=size, bold=bold, color=font_color)
        cell.alignment = align or center_align
        cell.border = thin_border
        if number_format:
            cell.number_format = number_format

    def write_metric_card(
        title_range: str,
        value_range: str,
        note_range: str,
        title: str,
        value_formula: str,
        note_formula: str,
        accent: str,
        *,
        note_fill: str,
        value_fill: str = paper,
        value_color: str = ink,
        value_size: int = 18,
        value_format: str | None = None,
    ) -> None:
        write_box(title_range, title, fill=accent, font_color=white, size=10, bold=True)
        write_box(
            value_range,
            value_formula,
            fill=value_fill,
            font_color=value_color,
            size=value_size,
            bold=True,
            number_format=value_format,
        )
        write_box(
            note_range,
            note_formula,
            fill=note_fill,
            font_color=muted,
            size=9,
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

    summary_formula = (
        '="Date snapshot : "&'
        + _raw_lookup_expr("snapshot_as_of_date")
        + '&" | Trimestre affiche : "&$C$4'
        + '&" | Statut : "&'
        + _raw_lookup_expr("snapshot_status_label")
        + '&" | Jours nowcast : "&'
        + _raw_lookup_expr("observed_days", fallback='"0"')
        + '&" | Couverture moyenne : "&IFERROR(TEXT('
        + _raw_lookup_expr("avg_coverage_ratio", fallback="0")
        + ', "0.0%"), "N/D")'
    )

    guidance_reference_formula = (
        '=IFERROR(IF('
        + _raw_lookup_expr("revenue_guidance_reference_musd", fallback="0")
        + '>0, "Guidance de reference : "&TEXT('
        + _raw_lookup_expr("revenue_guidance_reference_musd", fallback="0")
        + ', "0.0")&" M$"&IF('
        + _raw_lookup_expr("revenue_guidance_reference_quarter", fallback='""')
        + '<>""," ("&'
        + _raw_lookup_expr("revenue_guidance_reference_quarter", fallback='""')
        + '&")",""), "Guidance de reference : N/D"), "Guidance de reference : N/D")'
    )

    model_text = compact_summary_text(
        ai_sections.get("MODELE") or "",
        max_sentences=2,
        max_chars=190,
        separator="\n",
    )
    model_trends = compact_bullet_text(
        ai_sections.get("MODELE_TENDANCES") or ai_sections.get("TENDANCES") or "",
        max_items=2,
        max_chars=115,
    )
    model_risks = compact_bullet_text(
        ai_sections.get("MODELE_RISQUES") or ai_sections.get("ATTENTION") or "",
        max_items=2,
        max_chars=115,
    )

    write_box("A1:H2", "NOWCAST TRIMESTRIEL", fill=navy, font_color=white, size=18, bold=True)
    write_box(
        "A3:H3",
        summary_formula,
        fill=canvas,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_box("A4:B4", "Trimestre affiche", fill=blue_section, font_color=white, size=10, bold=True)
    write_box("C4", default_quarter, fill=paper, font_color=ink, size=11, bold=True)
    write_box(
        "D4:H4",
        "Selectionnez un trimestre pour relire son snapshot fige et ses estimations.",
        fill=canvas,
        font_color=muted,
        size=9,
        align=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )

    if available_quarters:
        validation = DataValidation(
            type="list",
            formula1=_formula_text_literal(",".join(available_quarters)),
            allow_blank=False,
        )
        validation.promptTitle = "Selection du trimestre"
        validation.prompt = "Choisissez le trimestre a afficher."
        ws.add_data_validation(validation)
        validation.add(ws["C4"])

    write_box("A6:F6", "Estimation revenus trimestrielle", fill=navy, font_color=white, size=11, bold=True)
    write_box("G6:H6", "Lecture du trimestre", fill=green, font_color=white, size=11, bold=True)

    write_box(
        "A7:F10",
        '=IFERROR(TEXT(' + _raw_lookup_expr("estimated_revenue_musd", fallback="0") + ', "0.0")&" M$","N/D")',
        fill=paper,
        font_color=ink,
        size=28,
        bold=True,
    )
    write_box(
        "A11:F11",
        _raw_lookup_formula("revenue_note_text"),
        fill=soft_blue,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_box(
        "G7:H8",
        _raw_lookup_formula("quarter_signal_bias"),
        fill=soft_slate,
        font_color=ink,
        size=18,
        bold=True,
    )
    write_box(
        "G9:H9",
        '="Confiance : "&' + _raw_lookup_expr("confidence_level"),
        fill=soft_slate,
        font_color=ink,
        size=10,
        bold=True,
    )
    write_box(
        "G10:H10",
        guidance_reference_formula,
        fill=soft_slate,
        font_color=muted,
        size=9,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )
    write_box(
        "G11:H11",
        '="Snapshot : "&' + _raw_lookup_expr("snapshot_status_label"),
        fill=soft_slate,
        font_color=muted,
        size=9,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_box("A13:H13", "Probabilites implicites", fill=blue_section, font_color=white, size=11, bold=True)
    write_metric_card(
        "A14:B14",
        "A15:B16",
        "A17:B17",
        "Prob. beat revenus",
        _raw_lookup_formula("revenue_beat_probability", fallback="0"),
        _raw_lookup_formula("revenue_note_text"),
        blue_section,
        note_fill=soft_blue,
        value_format="0.0%",
    )
    write_metric_card(
        "C14:D14",
        "C15:D16",
        "C17:D17",
        "Prob. beat EBITDA",
        _raw_lookup_formula("ebitda_beat_probability", fallback="0"),
        _raw_lookup_formula("ebitda_note_text"),
        orange,
        note_fill=soft_amber,
        value_format="0.0%",
    )
    write_metric_card(
        "E14:F14",
        "E15:F16",
        "E17:F17",
        "Prob. guidance raise",
        _raw_lookup_formula("guidance_raise_probability", fallback="0"),
        _raw_lookup_formula("next_guidance_note_text"),
        green,
        note_fill=soft_green,
        value_format="0.0%",
    )
    write_metric_card(
        "G14:H14",
        "G15:H16",
        "G17:H17",
        "EPS estime",
        '=IFERROR(TEXT(' + _raw_lookup_expr("estimated_eps", fallback="0") + ', "0.00")&" $","N/D")',
        _raw_lookup_formula("eps_note_text"),
        green,
        note_fill=soft_green,
    )

    write_box("A19:F19", "Lecture du modele", fill=navy, font_color=white, size=11, bold=True)
    write_box(
        "A20:F23",
        _ai_or_lookup_formula(model_text, "model_summary_text"),
        fill=paper,
        font_color=ink,
        size=11,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )
    write_box("G19:H19", "Pourquoi cette confiance", fill=blue_section, font_color=white, size=10, bold=True)
    write_box(
        "G20:H23",
        _raw_lookup_formula("confidence_context_text"),
        fill=soft_slate,
        font_color=ink,
        size=10,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_box("A25:D25", "Moteurs principaux", fill=green, font_color=white, size=11, bold=True)
    write_box("E25:H25", "Risques principaux", fill=red, font_color=white, size=11, bold=True)
    write_box(
        "A26:D29",
        _ai_or_lookup_formula(model_trends, "main_drivers_text"),
        fill=soft_green,
        font_color=ink,
        size=10,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )
    write_box(
        "E26:H29",
        _ai_or_lookup_formula(model_risks, "main_risks_text"),
        fill=soft_red,
        font_color=ink,
        size=10,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    next_step = readiness.get("next_step") or "Completer les labels trimestriels avant calibration supervisee."
    model_rows = [
        ("Snapshot selectionne", _raw_lookup_formula("snapshot_status_label")),
        ("Date du snapshot", _raw_lookup_formula("snapshot_as_of_date")),
        (
            "Reference guidance revenus",
            '=IFERROR(IF('
            + _raw_lookup_expr("revenue_guidance_reference_musd", fallback="0")
            + '>0, TEXT('
            + _raw_lookup_expr("revenue_guidance_reference_musd", fallback="0")
            + ', "0.0")&" M$"&IF('
            + _raw_lookup_expr("revenue_guidance_reference_quarter", fallback='""')
            + '<>""," ("&'
            + _raw_lookup_expr("revenue_guidance_reference_quarter", fallback='""')
            + '&")",""), "N/D"), "N/D")',
        ),
        ("Modele supervise pret", "Oui" if readiness.get("supervised_ready") else "Non"),
        ("Etape suivante", next_step),
    ]

    write_box("A31:H31", "Cadre du modele", fill=navy, font_color=white, size=11, bold=True)
    row_cursor = 32
    for label, value in model_rows:
        row_fill = soft_slate if row_cursor % 2 == 0 else paper
        ws[f"A{row_cursor}"] = label
        ws[f"A{row_cursor}"].fill = PatternFill(start_color=slate, end_color=slate, fill_type="solid")
        ws[f"A{row_cursor}"].font = Font(name=base_font_name, size=10, bold=True, color=ink)
        ws[f"A{row_cursor}"].alignment = left_align
        ws[f"A{row_cursor}"].border = thin_border
        ws.merge_cells(f"B{row_cursor}:H{row_cursor}")
        value_cell = ws[f"B{row_cursor}"]
        value_cell.value = value
        value_cell.fill = PatternFill(start_color=row_fill, end_color=row_fill, fill_type="solid")
        value_cell.font = Font(name=base_font_name, size=10, color=ink)
        value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        value_cell.border = thin_border
        for merged_col in range(3, 9):
            ws.cell(row=row_cursor, column=merged_col).border = thin_border
            ws.cell(row=row_cursor, column=merged_col).fill = PatternFill(start_color=row_fill, end_color=row_fill, fill_type="solid")
        row_cursor += 1

    history_header_row = row_cursor + 1
    write_box(
        f"A{history_header_row}:H{history_header_row}",
        "Historique trimestriel fige",
        fill=blue_section,
        font_color=white,
        size=11,
        bold=True,
    )
    history_table_row = history_header_row + 1
    history_headers = [
        "Trimestre",
        "Statut",
        "Date snapshot",
        "Score",
        "Beat rev.",
        "Beat EBITDA",
        "Guidance raise",
        "Jours obs.",
    ]
    for col_idx, header in enumerate(history_headers, start=1):
        cell = ws.cell(row=history_table_row, column=col_idx)
        cell.value = header
        cell.fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
        cell.font = Font(name=base_font_name, color=white, bold=True, size=10)
        cell.alignment = center_align
        cell.border = thin_border

    row_cursor = history_table_row + 1
    for idx, snapshot in enumerate(reversed(historical), start=0):
        row_fill = soft_slate if idx % 2 == 0 else paper
        values = [
            snapshot.get("quarter"),
            snapshot.get("snapshot_status_label"),
            snapshot.get("snapshot_as_of_date"),
            snapshot.get("quarter_signal_score"),
            snapshot.get("revenue_beat_probability", snapshot.get("revenue_beat_probability_proxy")),
            snapshot.get("ebitda_beat_probability", snapshot.get("ebitda_beat_probability_proxy")),
            snapshot.get("guidance_raise_probability", snapshot.get("guidance_raise_probability_proxy")),
            snapshot.get("observed_days"),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_cursor, column=col_idx)
            cell.value = value
            cell.fill = PatternFill(start_color=row_fill, end_color=row_fill, fill_type="solid")
            cell.border = thin_border
            cell.font = Font(name=base_font_name, size=10, color=ink)
            cell.alignment = center_align
            if col_idx == 4 and isinstance(value, (int, float)):
                cell.number_format = "0.0"
            elif col_idx in {5, 6, 7} and isinstance(value, (int, float)):
                cell.number_format = "0.0%"
            elif col_idx == 8 and isinstance(value, (int, float)):
                cell.number_format = "#,##0"
        row_cursor += 1

    assumptions_row = row_cursor + 1
    write_box(
        f"A{assumptions_row}:H{assumptions_row}",
        "Hypotheses",
        fill=blue_section,
        font_color=white,
        size=11,
        bold=True,
    )
    assumptions_text = "\n".join(f"- {item}" for item in assumptions) if assumptions else "- Aucune hypothese specifique."
    write_box(
        f"A{assumptions_row + 1}:H{assumptions_row + 4}",
        assumptions_text,
        fill=soft_slate,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[6].height = 24
    ws.row_dimensions[7].height = 34
    ws.row_dimensions[8].height = 34
    ws.row_dimensions[9].height = 22
    ws.row_dimensions[10].height = 22
    ws.row_dimensions[11].height = 22
    ws.row_dimensions[13].height = 24
    ws.row_dimensions[14].height = 22
    ws.row_dimensions[15].height = 26
    ws.row_dimensions[16].height = 26
    ws.row_dimensions[17].height = 24
    ws.row_dimensions[19].height = 24
    for row_idx in range(20, 24):
        ws.row_dimensions[row_idx].height = 34
    ws.row_dimensions[25].height = 24
    for row_idx in range(26, 30):
        ws.row_dimensions[row_idx].height = 28
    for row_idx in range(32, history_header_row):
        ws.row_dimensions[row_idx].height = 22
    ws.row_dimensions[history_header_row].height = 24
    ws.row_dimensions[history_table_row].height = 22
    for row_idx in range(history_table_row + 1, row_cursor):
        ws.row_dimensions[row_idx].height = 21
    for row_idx in range(assumptions_row + 1, assumptions_row + 5):
        ws.row_dimensions[row_idx].height = 22
