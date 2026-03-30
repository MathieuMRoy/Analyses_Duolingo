"""Renderer for the Alternative Data sheet."""

from __future__ import annotations

import numbers

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def render_alternative_data_sheet(
    ws,
    package: dict[str, object],
    styles: dict[str, object],
) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    for column_letter, width in {
        "A": 24,
        "B": 16,
        "C": 14,
        "D": 18,
        "E": 20,
        "F": 28,
    }.items():
        ws.column_dimensions[column_letter].width = width

    base_font_name = styles["BASE_FONT_NAME"]
    center_align = styles["center_align"]
    thin_border = styles["thin_border"]

    ink = "172635"
    muted = "607080"
    title = "1C3C5B"
    canvas = "F7F8FA"
    paper = "FFFFFF"
    line = "E9EDF2"
    soft_green = "EEF7E8"
    soft_red = "FCEBEC"
    soft_neutral = "F3F6F9"
    green = "5E8E2E"
    red = "C93B4D"
    grey = "8A919A"
    white = "FFFFFF"

    metadata = package.get("metadata", {})
    rows = package.get("rows") or []
    weekly_summary_columns = package.get("weekly_summary_columns") or []
    weekly_summary_rows = package.get("weekly_summary_rows") or []

    sheet_end_col = max(6, 2 + len(weekly_summary_columns))
    sheet_end_letter = get_column_letter(sheet_end_col)

    def write_box(
        ref: str,
        value: object,
        *,
        fill: str = paper,
        font_color: str = ink,
        size: int = 11,
        bold: bool = False,
        align=None,
        number_format: str | None = None,
    ) -> None:
        if ":" in ref:
            ws.merge_cells(ref)
            cell = ws[ref.split(":")[0]]
        else:
            cell = ws[ref]
        cell.value = value
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.font = Font(name=base_font_name, size=size, bold=bold, color=font_color)
        cell.alignment = align or center_align
        cell.border = thin_border
        if number_format:
            cell.number_format = number_format

    write_box(f"A1:{sheet_end_letter}2", "ALTERNATIVE DATA", fill=title, font_color=white, size=18, bold=True)
    write_box(
        f"A3:{sheet_end_letter}3",
        (
            "Lecture externe du momentum Duolingo. "
            "Chaque signal gratuit est capture quotidiennement quand la source publique est disponible, "
            "puis compare a sa derniere semaine observable."
        ),
        fill=canvas,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_box("A5:B5", "Snapshot", fill=title, font_color=white, size=10, bold=True)
    write_box("C5:D5", "Semaine", fill=title, font_color=white, size=10, bold=True)
    write_box("E5", "Signaux", fill=title, font_color=white, size=10, bold=True)
    write_box("F5", "Balance", fill=title, font_color=white, size=10, bold=True)

    write_box("A6:B6", metadata.get("latest_snapshot_date") or "N/D", fill=paper, font_color=ink, size=12, bold=True)
    write_box("C6:D6", metadata.get("latest_week_label") or "N/D", fill=paper, font_color=ink, size=12, bold=True)
    write_box("E6", metadata.get("signal_count") or 0, fill=paper, font_color=ink, size=12, bold=True)
    write_box(
        "F6",
        f"+{metadata.get('signals_up') or 0} / -{metadata.get('signals_down') or 0}",
        fill=paper,
        font_color=ink,
        size=12,
        bold=True,
    )

    headers = [
        ("A7", "Alternative Data"),
        ("B7", "Valeur"),
        ("C7", "Var. WoW"),
        ("D7", "Semaine"),
        ("E7", "Source"),
        ("F7", "Commentaire"),
    ]
    for ref, label in headers:
        write_box(ref, label, fill=canvas, font_color=ink, size=10, bold=True)

    if not metadata.get("has_data"):
        write_box(
            "A9:F11",
            (
                "Aucune donnee alternative exploitable pour le moment.\n"
                "Le pipeline tente d'abord de collecter des signaux publics gratuits, puis complete "
                "avec alternative_data_inputs.csv si vous ajoutez des valeurs manuelles.\n"
                "Dès qu'une premiere semaine est disponible, la variation week over week se calcule automatiquement."
            ),
            fill=soft_neutral,
            font_color=muted,
            size=11,
            align=Alignment(horizontal="left", vertical="top", wrap_text=True),
        )
        return

    start_row = 8
    for idx, row in enumerate(rows, start=start_row):
        wow_value = row.get("wow_change_pct")
        if isinstance(wow_value, numbers.Number):
            if wow_value > 0:
                wow_fill = soft_green
                wow_color = green
            elif wow_value < 0:
                wow_fill = soft_red
                wow_color = red
            else:
                wow_fill = soft_neutral
                wow_color = grey
        else:
            wow_fill = soft_neutral
            wow_color = grey

        write_box(f"A{idx}", row.get("signal_label") or "N/D", fill=paper, font_color=ink, size=11, bold=True, align=Alignment(horizontal="left", vertical="center"))
        write_box(f"B{idx}", row.get("value_display") or "N/D", fill=paper, font_color=ink, size=11, bold=False)
        write_box(f"C{idx}", row.get("wow_display") or "N/D", fill=wow_fill, font_color=wow_color, size=11, bold=True)
        write_box(f"D{idx}", row.get("week_label") or "N/D", fill=paper, font_color=muted, size=10)
        write_box(f"E{idx}", row.get("source") or "N/D", fill=paper, font_color=muted, size=10, align=Alignment(horizontal="left", vertical="center", wrap_text=True))
        write_box(f"F{idx}", row.get("notes") or "", fill=paper, font_color=muted, size=10, align=Alignment(horizontal="left", vertical="center", wrap_text=True))
        ws.row_dimensions[idx].height = 24

    last_row = start_row + len(rows)
    write_box(
        f"A{last_row}:F{last_row}",
        "Lecture WoW : comparaison entre la derniere valeur hebdomadaire disponible et la semaine precedente pour chaque signal.",
        fill=canvas,
        font_color=muted,
        size=9,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    if not weekly_summary_rows or not weekly_summary_columns:
        return

    summary_title_row = last_row + 2
    summary_header_row = summary_title_row + 1
    summary_data_row = summary_header_row + 1
    summary_end_letter = get_column_letter(2 + len(weekly_summary_columns))

    write_box(
        f"A{summary_title_row}:{summary_end_letter}{summary_title_row}",
        "HISTORIQUE HEBDOMADAIRE",
        fill=title,
        font_color=white,
        size=12,
        bold=True,
    )

    summary_headers = [
        ("A", "Semaine"),
        ("B", "Jours observes"),
    ]
    for index, column in enumerate(weekly_summary_columns, start=3):
        summary_headers.append((get_column_letter(index), column.get("display_label") or column.get("signal_label") or "Signal"))

    for column_letter, label in summary_headers:
        write_box(
            f"{column_letter}{summary_header_row}",
            label,
            fill=canvas,
            font_color=ink,
            size=10,
            bold=True,
        )

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 20)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 14)
    for index in range(3, 3 + len(weekly_summary_columns)):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, 16)

    for offset, row in enumerate(weekly_summary_rows):
        current_row = summary_data_row + offset
        stripe_fill = paper if offset % 2 == 0 else soft_neutral
        write_box(
            f"A{current_row}",
            row.get("week_label") or "N/D",
            fill=stripe_fill,
            font_color=ink,
            size=10,
            bold=False,
        )
        write_box(
            f"B{current_row}",
            row.get("days_observed") or 0,
            fill=stripe_fill,
            font_color=ink,
            size=10,
            bold=False,
        )
        values = row.get("values") or {}
        for index, column in enumerate(weekly_summary_columns, start=3):
            write_box(
                f"{get_column_letter(index)}{current_row}",
                values.get(column.get("signal_key")) or "N/D",
                fill=stripe_fill,
                font_color=ink,
                size=10,
                bold=False,
            )
        ws.row_dimensions[current_row].height = 22
