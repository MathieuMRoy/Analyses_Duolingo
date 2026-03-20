"""Renderer for the DCF valuation sheet."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill


def render_dcf_valuation_sheet(
    ws,
    package: dict[str, object],
    styles: dict[str, object],
) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    for column_letter, width in {
        "A": 30,
        "B": 16,
        "C": 14,
        "D": 4,
        "E": 22,
        "F": 12,
        "G": 16,
        "H": 15,
        "I": 15,
        "J": 3,
    }.items():
        ws.column_dimensions[column_letter].width = width

    base_font_name = styles["BASE_FONT_NAME"]
    center_align = styles["center_align"]
    thin_border = styles["thin_border"]

    ink = "1A2430"
    muted = "5F6B7A"
    navy = "17324D"
    steel = "274E6B"
    green = "2E7D5A"
    gold = "C7922E"
    canvas = "F5F2EA"
    paper = "FFFFFF"
    soft_gold = "FBF4E8"
    soft_blue = "EEF5FB"
    soft_green = "EDF6EF"
    pale_input = "FFF4CC"
    pale_matrix = "F7F1E4"
    pale_price = "DDF3E3"
    pale_note = "F2F6FA"
    white = "FFFFFF"

    metadata = package.get("metadata", {})
    assumptions = package.get("assumptions", {})
    anchors = package.get("anchors", {})
    notes = package.get("assumption_notes", []) or []

    def write_box(
        ref: str,
        value: object,
        *,
        fill: str = paper,
        font_color: str = ink,
        size: int = 11,
        bold: bool = False,
        align=None,
        number_format: str | None = None,
    ) -> None:
        if ":" in ref:
            ws.merge_cells(ref)
            cell = ws[ref.split(":")[0]]
        else:
            cell = ws[ref]
        cell.value = value
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.font = Font(name=base_font_name, size=size, bold=bold, color=font_color)
        cell.alignment = align or center_align
        cell.border = thin_border
        if number_format:
            cell.number_format = number_format

    def write_label_value(row: int, label: str, value: object, *, fmt: str | None = None, editable: bool = False) -> None:
        write_box(f"A{row}:B{row}", label, fill=paper, font_color=ink, size=11, bold=False, align=Alignment(horizontal="left", vertical="center"))
        write_box(
            f"C{row}",
            value,
            fill=pale_input if editable else paper,
            font_color=ink,
            size=11,
            bold=True if editable else False,
            number_format=fmt,
        )

    def write_anchor_row(row: int, label: str, value: object, *, fmt: str | None = None) -> None:
        write_box(f"E{row}:G{row}", label, fill=soft_blue, font_color=muted, size=10, bold=True, align=Alignment(horizontal="left", vertical="center"))
        write_box(f"H{row}:I{row}", value, fill=paper, font_color=ink, size=12, bold=True, number_format=fmt)

    write_box("A1:I2", "VALORISATION DCF", fill=navy, font_color=white, size=18, bold=True)
    write_box(
        "A3:I3",
        "Lecture de valorisation sous hypotheses. Les cellules jaunes peuvent etre ajustees manuellement pour tester votre scenario central.",
        fill=canvas,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_box("A5:C5", "HYPOTHESES", fill=steel, font_color=white, size=11, bold=True)
    write_box("E5:I5", "POINT DE DEPART", fill=green, font_color=white, size=11, bold=True)

    write_label_value(6, "Croissance annuelle (growth)", assumptions.get("growth_rate"), fmt="0.0%", editable=True)
    write_label_value(7, "Cout du capital (WACC)", assumptions.get("wacc"), fmt="0.0%", editable=True)
    write_label_value(8, "Croissance perpetuelle", assumptions.get("terminal_growth"), fmt="0.0%", editable=True)

    write_anchor_row(6, "Trimestre suivi", metadata.get("quarter"))
    write_anchor_row(7, "Snapshot", metadata.get("as_of_date"))
    write_anchor_row(8, "Revenus trimestriels estimes", anchors.get("estimated_revenue_current_quarter_musd"), fmt='0.0 "M$"')
    write_anchor_row(9, "EBITDA trimestriel estime", anchors.get("estimated_ebitda_current_quarter_musd"), fmt='0.0 "M$"')
    write_anchor_row(10, "Guidance FY de reference", anchors.get("fy_guidance_revenue_musd"), fmt='0.0 "M$"')

    write_box("A11:C11", "DONNEES DE BASE", fill=steel, font_color=white, size=11, bold=True)
    write_box("E11:I11", "PROJECTIONS (FCF)", fill=steel, font_color=white, size=11, bold=True)

    write_label_value(12, "Revenus TTM estimes", anchors.get("revenue_ttm_estimated_musd"), fmt='0.0 "M$"')
    write_label_value(13, "EBITDA TTM estime", anchors.get("ebitda_ttm_estimated_musd"), fmt='0.0 "M$"')
    write_label_value(14, "Cash flow operationnel", anchors.get("operating_cash_flow_musd"), fmt='0.0 "M$"')
    capex_plus = (anchors.get("capitalized_software_musd") or 0) + (anchors.get("capex_musd") or 0)
    write_label_value(15, "CapEx + logiciels", capex_plus, fmt='0.0 "M$"')
    write_label_value(16, "Free cash flow historique", anchors.get("free_cash_flow_historical_musd"), fmt='0.0 "M$"')
    write_label_value(17, "Free cash flow de base", anchors.get("free_cash_flow_base_musd"), fmt='0.0 "M$"')
    write_label_value(18, "Tresorerie + placements", anchors.get("cash_and_investments_musd"), fmt='0.0 "M$"')
    write_label_value(19, "Dette totale", anchors.get("total_debt_musd"), fmt='0.0 "M$"')
    write_label_value(20, "Actions diluees (M)", anchors.get("diluted_shares_m"), fmt='0.000')

    headers = [
        ("E12", "Annee"),
        ("F12", "Croissance"),
        ("G12", "FCF"),
        ("H12", "Facteur PV"),
        ("I12", "PV"),
    ]
    for cell_ref, label in headers:
        write_box(cell_ref, label, fill=soft_gold, font_color=ink, size=10, bold=True)

    growth_formulas = {
        13: "=$C$6",
        14: "=MAX($C$8,$C$6-(($C$6-$C$8)*1/4))",
        15: "=MAX($C$8,$C$6-(($C$6-$C$8)*2/4))",
        16: "=MAX($C$8,$C$6-(($C$6-$C$8)*3/4))",
        17: "=$C$8",
    }
    for offset, row in enumerate(range(13, 18), start=1):
        write_box(f"E{row}", f"Annee {offset}", fill=paper, font_color=ink, size=10, bold=True)
        write_box(f"F{row}", growth_formulas[row], fill=paper, font_color=ink, size=10, bold=True, number_format="0.0%")
        if row == 13:
            fcf_formula = "=IFERROR($C$17*(1+F13),NA())"
        else:
            fcf_formula = f"=IFERROR(G{row-1}*(1+F{row}),NA())"
        write_box(f"G{row}", fcf_formula, fill=paper, font_color=ink, size=10, bold=False, number_format='0.0 "M$"')
        write_box(f"H{row}", f"=(1+$C$7)^{offset}", fill=paper, font_color=muted, size=10, bold=False, number_format="0.000x")
        write_box(f"I{row}", f"=IFERROR(G{row}/H{row},NA())", fill=paper, font_color=ink, size=10, bold=False, number_format='0.0 "M$"')

    write_box("A22:C22", "VALORISATION", fill=steel, font_color=white, size=11, bold=True)
    write_box("E22:I22", "SENSIBILITE WACC / TERMINALE", fill=gold, font_color=white, size=11, bold=True)

    write_label_value(23, "Somme des flux actualises", "=SUM(I13:I17)", fmt='0.0 "M$"')
    write_label_value(24, "Valeur terminale actualisee", "=IF($C$7<=$C$8,NA(),G17*(1+$C$8)/($C$7-$C$8)/H17)", fmt='0.0 "M$"')
    write_label_value(25, "Enterprise value (EV)", "=IFERROR($C$23+$C$24,NA())", fmt='0.0 "M$"')
    write_label_value(26, "Cash net", "=IFERROR($C$18-$C$19,NA())", fmt='0.0 "M$"')
    write_label_value(27, "Equity value", "=IFERROR($C$25+$C$26,NA())", fmt='0.0 "M$"')

    write_box("A29:C31", '=IFERROR("$ "&TEXT($C$27/$C$20,"0.00"),"N/D")', fill=pale_price, font_color=green, size=22, bold=True)
    write_box("A28:C28", "PRIX CIBLE PAR ACTION", fill=green, font_color=white, size=11, bold=True)

    terminal_headers = [0.02, assumptions.get("terminal_growth") or 0.03, 0.04]
    wacc_headers = [
        max(0.01, (assumptions.get("wacc") or 0.1058) - 0.01),
        assumptions.get("wacc") or 0.1058,
        (assumptions.get("wacc") or 0.1058) + 0.01,
    ]
    write_box("F23", "", fill=pale_matrix)
    for idx, terminal in enumerate(terminal_headers, start=7):
        write_box(f"{chr(64+idx)}23", terminal, fill=pale_matrix, font_color=ink, size=10, bold=True, number_format="0.0%")
    for offset, wacc in enumerate(wacc_headers, start=24):
        write_box(f"F{offset}", wacc, fill=pale_matrix, font_color=ink, size=10, bold=True, number_format="0.0%")
        for col_idx, _ in enumerate(terminal_headers, start=7):
            terminal_cell = f"{chr(64+col_idx)}23"
            value_formula = (
                f'=IF($F{offset}<={terminal_cell},NA(),'
                f'($C$23 + (G17*(1+{terminal_cell})/($F{offset}-{terminal_cell})/((1+$F{offset})^5)) + $C$26)/$C$20)'
            )
            write_box(f"{chr(64+col_idx)}{offset}", value_formula, fill=paper, font_color=ink, size=10, bold=False, number_format='$ 0.00')

    write_box("A33:I33", "NOTES DE MODELE", fill=navy, font_color=white, size=11, bold=True)
    joined_notes = " | ".join(str(note) for note in notes if str(note).strip())
    if metadata.get("source_file"):
        source_note = f"Source filing : {metadata.get('source_file')}"
        joined_notes = f"{joined_notes} | {source_note}" if joined_notes else source_note
    write_box(
        "A34:I36",
        joined_notes or "Valorisation basee sur le nowcast trimestriel, l'historique des resultats et les derniers fondamentaux disponibles.",
        fill=pale_note,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    # --- Prominent price target callout on the right side (middle) ---
    for col_letter in ("K", "L", "M", "N"):
        ws.column_dimensions[col_letter].width = 14
    write_box("K12:N12", "PRIX CIBLE PAR ACTION", fill=green, font_color=white, size=13, bold=True)
    write_box(
        "K13:N19",
        '=IFERROR("$ "&TEXT($C$27/$C$20,"0.00"),"N/D")',
        fill=pale_price,
        font_color=green,
        size=36,
        bold=True,
    )
    write_box(
        "K20:N20",
        "Valorisation DCF implicite",
        fill=canvas,
        font_color=muted,
        size=9,
        align=Alignment(horizontal="center", vertical="center"),
    )

    for row in [6, 7, 8, 12, 13, 14, 15, 16, 17, 18, 19, 20, 23, 24, 25, 26, 27]:
        ws.row_dimensions[row].height = 24
    for row in [29, 30, 31]:
        ws.row_dimensions[row].height = 24
    for row in [34, 35, 36]:
        ws.row_dimensions[row].height = 28
