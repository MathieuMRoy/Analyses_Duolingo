"""Renderer for the visible AI briefing sheet."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from ...financial_signals import FINANCIAL_SIGNALS_JSON_FILE
from ...quarterly_nowcast import QUARTERLY_NOWCAST_JSON_FILE
from ...valuation_dcf import DCF_VALUATION_JSON_FILE
from ..render_helpers import build_render_helpers
from .kpi_dictionary_sheet import build_kpi_dictionary_df


def _safe_load_json(path: Path) -> dict[str, object] | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _safe_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _extract_ai_sections(ia_report: str | None) -> dict[str, str]:
    sections = {
        "RESUME": "",
        "TENDANCES": "",
        "ATTENTION": "",
        "CONSEILS": "",
        "MODELE": "",
        "MODELE_TENDANCES": "",
        "MODELE_RISQUES": "",
    }
    if not ia_report:
        return sections

    for key in sections:
        match = re.search(rf"\[{key}\](.*?)(?=\[|$)", ia_report, re.DOTALL)
        if match:
            sections[key] = match.group(1).strip()
    return sections


def _compute_dcf_readout(dcf_valuation: dict[str, object] | None) -> dict[str, float | None]:
    if not dcf_valuation:
        return {"price_target": None, "current_share_price": None, "upside_downside_pct": None}

    assumptions = dcf_valuation.get("assumptions") or {}
    anchors = dcf_valuation.get("anchors") or {}
    market = dcf_valuation.get("market_context") or {}

    growth_rate = _safe_float(assumptions.get("growth_rate"))
    wacc = _safe_float(assumptions.get("wacc"))
    terminal_growth = _safe_float(assumptions.get("terminal_growth"))
    fcf_base = _safe_float(anchors.get("free_cash_flow_base_musd"))
    net_cash = _safe_float(anchors.get("net_cash_musd"))
    diluted_shares = _safe_float(anchors.get("diluted_shares_m"))
    current_share_price = _safe_float(market.get("current_share_price"))

    if None in {growth_rate, wacc, terminal_growth, fcf_base, net_cash, diluted_shares}:
        return {
            "price_target": None,
            "current_share_price": current_share_price,
            "upside_downside_pct": None,
        }
    if diluted_shares <= 0 or wacc <= terminal_growth:
        return {
            "price_target": None,
            "current_share_price": current_share_price,
            "upside_downside_pct": None,
        }

    yearly_growth = [
        growth_rate,
        max(terminal_growth, growth_rate - ((growth_rate - terminal_growth) * 1 / 4)),
        max(terminal_growth, growth_rate - ((growth_rate - terminal_growth) * 2 / 4)),
        max(terminal_growth, growth_rate - ((growth_rate - terminal_growth) * 3 / 4)),
        terminal_growth,
    ]

    discounted_flows: list[float] = []
    previous_fcf = fcf_base
    for index, year_growth in enumerate(yearly_growth, start=1):
        current_fcf = previous_fcf * (1 + year_growth)
        discounted_flows.append(current_fcf / ((1 + wacc) ** index))
        previous_fcf = current_fcf

    terminal_fcf = previous_fcf * (1 + terminal_growth)
    terminal_value = terminal_fcf / (wacc - terminal_growth) / ((1 + wacc) ** 5)
    enterprise_value = sum(discounted_flows) + terminal_value
    equity_value = enterprise_value + net_cash
    price_target = equity_value / diluted_shares

    upside_downside_pct = None
    if current_share_price not in (None, 0):
        upside_downside_pct = price_target / current_share_price - 1.0

    return {
        "price_target": round(price_target, 2),
        "current_share_price": round(current_share_price, 2) if current_share_price is not None else None,
        "upside_downside_pct": round(upside_downside_pct, 4) if upside_downside_pct is not None else None,
    }


def _best_alt_data_line(alternative_data: dict[str, object] | None) -> str:
    rows = (alternative_data or {}).get("rows") or []
    ranked_rows = [
        row
        for row in rows
        if isinstance(row.get("wow_change_pct"), float)
    ]
    if not ranked_rows:
        return "Les signaux externes manquent encore de profondeur hebdomadaire exploitable."

    best_row = max(ranked_rows, key=lambda row: abs(float(row.get("wow_change_pct") or 0.0)))
    signal_label = str(best_row.get("signal_label") or "Signal externe")
    wow_display = str(best_row.get("wow_display") or "N/D")
    return f"Signal externe le plus mobile : {signal_label} ({wow_display} en WoW)."


def render_briefing_ai_sheet(
    ws,
    wb,
    raw_sheet_name: str,
    styles: dict[str, object],
    helpers: dict[str, object] | None = None,
    ia_report: str | None = None,
    signal_package: dict[str, object] | None = None,
    quarterly_nowcast: dict[str, object] | None = None,
    alternative_data: dict[str, object] | None = None,
    dcf_valuation: dict[str, object] | None = None,
) -> None:
    helpers = helpers or build_render_helpers()
    signal_package = signal_package or _safe_load_json(FINANCIAL_SIGNALS_JSON_FILE) or {}
    quarterly_nowcast = quarterly_nowcast or _safe_load_json(QUARTERLY_NOWCAST_JSON_FILE) or {}
    dcf_valuation = dcf_valuation or _safe_load_json(DCF_VALUATION_JSON_FILE) or {}

    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    for column_letter, width in {
        "A": 16,
        "B": 16,
        "C": 16,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 16,
    }.items():
        ws.column_dimensions[column_letter].width = width

    base_font_name = styles["BASE_FONT_NAME"]
    center_align = styles["center_align"]
    thin_border = styles["thin_border"]

    label_signal_bias = helpers["label_signal_bias"]
    label_confidence = helpers["label_confidence"]
    pretty_fr_number = helpers["pretty_fr_number"]
    pretty_ratio_pct = helpers["pretty_ratio_pct"]
    compact_summary_text = helpers["compact_summary_text"]
    compact_bullet_text = helpers["compact_bullet_text"]

    deep_navy = "203A56"
    canvas = "F5F7FA"
    surface = "FFFFFF"
    white = "FFFFFF"
    ink = "17324D"
    muted = "627081"
    slate = "EAF1F7"
    sand = "F7F1E5"
    sea = "2F7A67"
    bronze = "B97A2D"
    plum = "6B5B82"
    soft_green = "EEF7EE"
    soft_red = "FBECEC"
    soft_blue = "EEF5FB"
    soft_plum = "F4EFF8"

    def write_box(
        range_ref: str,
        value: str,
        *,
        fill: str = surface,
        font_color: str = ink,
        size: int = 11,
        bold: bool = False,
        align=None,
    ) -> None:
        ws.merge_cells(range_ref)
        cell = ws[range_ref.split(":")[0]]
        cell.value = value
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.font = Font(name=base_font_name, size=size, bold=bold, color=font_color)
        cell.alignment = align or center_align
        cell.border = thin_border

    ai_sections = _extract_ai_sections(ia_report)
    signal_proxy = (signal_package.get("financial_proxy_signals") or {})
    signal_meta = (signal_package.get("metadata") or {})
    signal_horizon = (signal_package.get("horizon_context") or {})
    quarter_meta = (quarterly_nowcast.get("metadata") or {})
    quarter_model = (quarterly_nowcast.get("model_output") or {})
    quarter_current = (quarterly_nowcast.get("current_quarter") or {})
    dcf_readout = _compute_dcf_readout(dcf_valuation)

    signal_label = label_signal_bias(signal_proxy.get("signal_bias"))
    signal_confidence = label_confidence(signal_proxy.get("confidence_level"))
    quarter_bias = label_signal_bias(quarter_model.get("quarter_signal_bias"))
    quarter_confidence = label_confidence(quarter_model.get("confidence_level"))
    quarter_name = str(
        quarter_current.get("quarter")
        or quarter_meta.get("current_quarter")
        or "N/D"
    )

    daily_text = ai_sections["RESUME"] or (
        f"Signal {signal_label.lower()} avec confiance {signal_confidence.lower()}. "
        f"{(signal_horizon.get('seven_day') or {}).get('summary_text') or 'On lit surtout la direction recente du panel.'}"
    )
    daily_text = compact_summary_text(daily_text, max_sentences=2, max_chars=230, separator="\n")

    quarterly_text = ai_sections["MODELE"] or (
        f"Le nowcast {quarter_name} reste {quarter_bias.lower()} avec confiance {quarter_confidence.lower()}. "
        f"Les revenus estimes ressortent a {pretty_fr_number(quarter_model.get('estimated_revenue_musd'), 1)} M$ "
        f"et la probabilite implicite de beat revenus a {pretty_ratio_pct(quarter_model.get('revenue_beat_probability'), 1)}."
    )
    quarterly_text = compact_summary_text(quarterly_text, max_sentences=2, max_chars=230, separator="\n")

    improving_text = compact_bullet_text(
        ai_sections["TENDANCES"] or "\n".join(f"- {item}" for item in (signal_proxy.get("main_drivers") or [])[:2]),
        max_items=2,
        max_chars=220,
    )
    risk_text = compact_bullet_text(
        ai_sections["ATTENTION"] or "\n".join(f"- {item}" for item in (signal_proxy.get("main_risks") or [])[:2]),
        max_items=2,
        max_chars=220,
    )

    dcf_line = "La valorisation DCF n'est pas encore exploitable."
    if dcf_readout.get("price_target") is not None:
        dcf_line = (
            f"DCF : prix cible {pretty_fr_number(dcf_readout['price_target'], 2)} $ "
            f"vs cours {pretty_fr_number(dcf_readout.get('current_share_price'), 2)} $."
        )
        if dcf_readout.get("upside_downside_pct") is not None:
            dcf_line += f" Ecart implicite {pretty_ratio_pct(dcf_readout.get('upside_downside_pct'), 1)}."

    watch_text = ai_sections["CONSEILS"] or (
        f"{(signal_horizon.get('thirty_day') or {}).get('summary_text') or 'On surveille surtout la tenue de la monetisation et du churn.'} "
        f"{_best_alt_data_line(alternative_data)}"
    )
    watch_text = compact_summary_text(watch_text, max_sentences=2, max_chars=230, separator="\n")

    summary_line = (
        f"Date de reference : {signal_meta.get('as_of_date', quarter_meta.get('as_of_date', 'N/D'))} | "
        f"Signal du jour : {signal_label} | Trimestre suivi : {quarter_name}"
    )

    write_box("A1:H2", "BRIEFING IA", fill=deep_navy, font_color=white, size=18, bold=True)
    write_box(
        "A3:H3",
        summary_line,
        fill=canvas,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

    write_box("A5:D5", "Lecture du jour", fill=sea, font_color=white, size=11, bold=True)
    write_box(
        "A6:D10",
        daily_text,
        fill=surface,
        font_color=ink,
        size=12,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_box("E5:H5", "Lecture trimestrielle", fill=bronze, font_color=white, size=11, bold=True)
    write_box(
        "E6:H10",
        quarterly_text,
        fill=surface,
        font_color=ink,
        size=12,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_box("A12:B12", "Signal du jour", fill=deep_navy, font_color=white, size=10, bold=True)
    write_box("A13:B14", signal_label, fill=soft_blue, font_color=ink, size=16, bold=True)
    write_box("C12:D12", "Confiance", fill=deep_navy, font_color=white, size=10, bold=True)
    write_box("C13:D14", quarter_confidence, fill=soft_plum, font_color=ink, size=16, bold=True)
    write_box("E12:F12", "Prix cible DCF", fill=deep_navy, font_color=white, size=10, bold=True)
    write_box(
        "E13:F14",
        f"{pretty_fr_number(dcf_readout.get('price_target'), 2)} $" if dcf_readout.get("price_target") is not None else "N/D",
        fill=sand,
        font_color=ink,
        size=16,
        bold=True,
    )
    write_box("G12:H12", "Upside implicite", fill=deep_navy, font_color=white, size=10, bold=True)
    upside_value = dcf_readout.get("upside_downside_pct")
    upside_fill = soft_green if isinstance(upside_value, float) and upside_value >= 0 else soft_red
    write_box(
        "G13:H14",
        pretty_ratio_pct(upside_value, 1),
        fill=upside_fill,
        font_color=ink,
        size=16,
        bold=True,
    )

    write_box("A16:D16", "Ce qui s'ameliore", fill=sea, font_color=white, size=11, bold=True)
    write_box(
        "A17:D21",
        improving_text or "Aucune amelioration marquante n'a encore emerge.",
        fill=soft_green,
        font_color=ink,
        size=11,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_box("E16:H16", "Points de vigilance", fill=plum, font_color=white, size=11, bold=True)
    write_box(
        "E17:H21",
        risk_text or "Aucun risque dominant n'est encore ressorti du signal.",
        fill=soft_red,
        font_color=ink,
        size=11,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_box("A23:D23", "Watchlist", fill=bronze, font_color=white, size=11, bold=True)
    write_box(
        "A24:D28",
        watch_text,
        fill=surface,
        font_color=ink,
        size=11,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_box("E23:H23", "Contexte de marche", fill=deep_navy, font_color=white, size=11, bold=True)
    write_box(
        "E24:H28",
        compact_summary_text(dcf_line, max_sentences=2, max_chars=220, separator="\n"),
        fill=slate,
        font_color=ink,
        size=11,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    rows = build_kpi_dictionary_df().to_dict("records")
    kpi_count = len(rows)
    default_kpi = "Prob. beat revenus"
    if not any(str(row.get("KPI")) == default_kpi for row in rows):
        default_kpi = str(rows[0].get("KPI") or "Moyenne Streak (Jours)")

    write_box("A30:H30", "Assistant KPI", fill=deep_navy, font_color=white, size=11, bold=True)
    write_box("A31:B31", "KPI choisi", fill=sand, font_color=ink, size=10, bold=True)
    write_box("C31:E31", default_kpi, fill=surface, font_color=ink, size=11, bold=True)
    write_box(
        "F31:H31",
        f'="Famille : "&IFERROR(INDEX(\'{raw_sheet_name}\'!$A$2:$A$999,MATCH($C$31,\'{raw_sheet_name}\'!$B$2:$B$999,0)),"")',
        fill=soft_blue,
        font_color=muted,
        size=10,
        bold=True,
    )

    validation = DataValidation(
        type="list",
        formula1=f"='{raw_sheet_name}'!$B$2:$B${kpi_count + 1}",
        allow_blank=False,
    )
    validation.promptTitle = "Choix du KPI"
    validation.prompt = "Choisissez le KPI que vous voulez faire expliquer."
    ws.add_data_validation(validation)
    validation.add(ws["C31"])

    write_box("A32:C32", "Ce que cela veut dire", fill=sea, font_color=white, size=10, bold=True)
    write_box(
        "A33:H35",
        f'=IFERROR(INDEX(\'{raw_sheet_name}\'!$C$2:$C$999,MATCH($C$31,\'{raw_sheet_name}\'!$B$2:$B$999,0)),"")',
        fill=soft_blue,
        font_color=ink,
        size=11,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_box("A36:C36", "Comment c'est calcule", fill=plum, font_color=white, size=10, bold=True)
    write_box(
        "A37:H39",
        f'=IFERROR(INDEX(\'{raw_sheet_name}\'!$D$2:$D$999,MATCH($C$31,\'{raw_sheet_name}\'!$B$2:$B$999,0)),"")',
        fill=surface,
        font_color=muted,
        size=11,
        align=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )

    write_box(
        "A41:H41",
        "Le dictionnaire KPI reste disponible en feuille raw cachee. Ici, le but est d'expliquer vite quoi lire, pourquoi, et comment verifier le calcul si necessaire.",
        fill=canvas,
        font_color=muted,
        size=10,
        align=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )

    for row_number, height in {
        3: 26,
        6: 28,
        7: 28,
        8: 28,
        9: 28,
        10: 28,
        13: 24,
        14: 24,
        17: 24,
        18: 24,
        19: 24,
        20: 24,
        21: 24,
        24: 24,
        25: 24,
        26: 24,
        27: 24,
        28: 24,
        33: 28,
        34: 28,
        35: 28,
        37: 28,
        38: 28,
        39: 28,
        41: 30,
    }.items():
        ws.row_dimensions[row_number].height = height
