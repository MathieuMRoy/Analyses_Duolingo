from __future__ import annotations

from math import ceil, floor
from datetime import datetime
from pathlib import Path
import numbers
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .columns import (
    CHART_DATA_SHEET as SHARED_CHART_DATA_SHEET,
    SUMMARY_SHEET as SHARED_SUMMARY_SHEET,
    TRENDS_SHEET as SHARED_TREND_SHEET,
)
from .config import DAILY_LOG_FILE
from .utils import (
    delta_font_color,
    format_integer,
    format_percent,
    format_points,
    format_streak,
    format_xp,
    format_xp_delta,
    parse_bool_fraction,
    parse_float,
    weekly_insight_text,
)

SUMMARY_SHEET = "Suivi Quotidien"
TREND_SHEET = "📈 Tendances Mensuelles"
CHART_DATA_SHEET = "📊 Données Graphique"

SUMMARY_SHEET = SHARED_SUMMARY_SHEET
TREND_SHEET = SHARED_TREND_SHEET
CHART_DATA_SHEET = SHARED_CHART_DATA_SHEET

PERCENT_RATIO_COLUMNS = {
    "Taux Abonn. Super",
    "Taux d'Abandon Global",
    "Score d'Engagement",
    "Abandon Debutants",
    "Abandon Standard",
    "Abandon Super-Actifs",
    "Conv. Super moy.",
    "Taux d'abandon moyen",
    "Score d'eng. moyen",
}


def _load_sheet(report_path: Path, sheet_name: str) -> pd.DataFrame:
    try:
        return pd.read_excel(report_path, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame()


def _coerce_dates(series: pd.Series) -> pd.Series:
    def _parse_date(value: object):
        if value is None:
            return pd.NaT
        if isinstance(value, pd.Timestamp):
            return value
        if isinstance(value, datetime):
            return pd.Timestamp(value)
        if isinstance(value, str):
            raw = value.strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d %b"):
                try:
                    return pd.Timestamp(datetime.strptime(raw, fmt))
                except ValueError:
                    continue
            return pd.NaT
        try:
            return pd.Timestamp(value)
        except Exception:
            return pd.NaT

    return series.apply(_parse_date)


def _first_existing_column(df: pd.DataFrame, names: list[str]) -> str | None:
    for name in names:
        if name in df.columns:
            return name
    return None


def _normalize_percent_series(series: pd.Series) -> pd.Series:
    parsed = series.apply(parse_float)
    return parsed.apply(
        lambda x: (x / 100) if isinstance(x, numbers.Number) and abs(x) > 1 else x
    )


def _prepare_summary_metrics_df(summary_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df.empty or "Date" not in summary_df.columns:
        return pd.DataFrame(
            columns=[
                "Date",
                "Serie Moyenne (Jours)",
                "Apprentissage (XP/j)",
                "Taux Abonn. Super",
                "Taux d'Abandon Global",
                "Reactivations vs Veille",
                "Score d'Engagement",
                "Panel Total",
                "Abandon Debutants",
                "Abandon Standard",
                "Abandon Super-Actifs",
            ]
        )

    df = summary_df.copy()
    df["Date"] = _coerce_dates(df["Date"])
    df = df.dropna(subset=["Date"]).sort_values("Date")
    if df.empty:
        return df

    metric_columns = {
        "Serie Moyenne (Jours)": ["Série Moyenne (Jours)", "Moyenne Streak (J)"],
        "Apprentissage (XP/j)": ["Apprentissage (XP/j)", "Delta XP (Intensité)"],
        "Taux Abonn. Super": ["Taux Abonn. Super", "Conversion Premium"],
        "Taux d'Abandon Global": ["Taux d'Abandon Global", "Taux d'Attrition Global", "Churn Global"],
        "Reactivations vs Veille": ["Reactivations vs Veille"],
        "Score d'Engagement": ["Score d'Engagement", "Score Santé Global"],
        "Panel Total": ["Panel Total", "Total Profils"],
        "Abandon Debutants": ["Abandon Débutants", "Churn Débutants"],
        "Abandon Standard": ["Abandon Standard", "Churn Standard"],
        "Abandon Super-Actifs": ["Abandon Super-Actifs", "Churn Super-Actifs"],
    }

    normalized = pd.DataFrame({"Date": df["Date"]})
    for target, options in metric_columns.items():
        source = _first_existing_column(df, options)
        if source:
            if target in PERCENT_RATIO_COLUMNS:
                normalized[target] = _normalize_percent_series(df[source])
            else:
                normalized[target] = df[source].apply(parse_float)
        else:
            normalized[target] = None

    return normalized.sort_values("Date").reset_index(drop=True)


def _prepare_global_trend_df(summary_df: pd.DataFrame) -> pd.DataFrame:
    normalized = _prepare_summary_metrics_df(summary_df)
    if normalized.empty:
        return pd.DataFrame(columns=["Libelle", "Taux Abonn. Super"])

    df = normalized.dropna(subset=["Taux Abonn. Super"]).copy()
    df["Libelle"] = df["Date"].dt.strftime("%d %b")
    return df[["Libelle", "Taux Abonn. Super"]]


def _prepare_weekly_comparison_df(summary_df: pd.DataFrame) -> pd.DataFrame:
    normalized = _prepare_summary_metrics_df(summary_df)
    if normalized.empty:
        return pd.DataFrame(
            columns=[
                "Semaine",
                "Jours observes",
                "XP moyen",
                "Delta XP vs S-1",
                "Conv. Super moy.",
                "Delta Conv. vs S-1",
                "Taux d'abandon moyen",
                "Delta Abandon vs S-1",
                "Score d'eng. moyen",
                "Delta Score vs S-1",
                "Panel moyen",
            ]
        )

    df = normalized.copy()
    df["WeekStart"] = df["Date"] - pd.to_timedelta(df["Date"].dt.weekday, unit="D")
    df["WeekEnd"] = df["WeekStart"] + pd.Timedelta(days=6)

    weekly = (
        df.groupby(["WeekStart", "WeekEnd"], as_index=False)
        .agg(
            {
                "Date": "nunique",
                "Apprentissage (XP/j)": "mean",
                "Taux Abonn. Super": "mean",
                "Taux d'Abandon Global": "mean",
                "Score d'Engagement": "mean",
                "Panel Total": "mean",
            }
        )
        .sort_values("WeekStart")
    )

    weekly["Semaine"] = (
        weekly["WeekStart"].dt.strftime("%d %b")
        + " - "
        + weekly["WeekEnd"].dt.strftime("%d %b")
    )
    weekly["Delta XP vs S-1"] = weekly["Apprentissage (XP/j)"].diff()
    weekly["Delta Conv. vs S-1"] = weekly["Taux Abonn. Super"].diff()
    weekly["Delta Abandon vs S-1"] = weekly["Taux d'Abandon Global"].diff()
    weekly["Delta Score vs S-1"] = weekly["Score d'Engagement"].diff()

    weekly = weekly.rename(
        columns={
            "Date": "Jours observes",
            "Apprentissage (XP/j)": "XP moyen",
            "Taux Abonn. Super": "Conv. Super moy.",
            "Taux d'Abandon Global": "Taux d'abandon moyen",
            "Score d'Engagement": "Score d'eng. moyen",
            "Panel Total": "Panel moyen",
        }
    )

    return weekly[
        [
            "Semaine",
            "Jours observes",
            "XP moyen",
            "Delta XP vs S-1",
            "Conv. Super moy.",
            "Delta Conv. vs S-1",
            "Taux d'abandon moyen",
            "Delta Abandon vs S-1",
            "Score d'eng. moyen",
            "Delta Score vs S-1",
            "Panel moyen",
        ]
    ].tail(8).reset_index(drop=True)


def _prepare_snapshot_df(chart_df: pd.DataFrame) -> pd.DataFrame:
    if chart_df.empty:
        return pd.DataFrame(columns=["Segment", "Taux"])

    if {"Segment", "Taux (%)"}.issubset(chart_df.columns):
        df = chart_df[["Segment", "Taux (%)"]].copy()
        df["Taux"] = df["Taux (%)"].apply(parse_float)
        df = df.dropna(subset=["Segment", "Taux"])
        if not df.empty:
            return df[["Segment", "Taux"]].reset_index(drop=True)

    if "Date" not in chart_df.columns:
        return pd.DataFrame(columns=["Segment", "Taux"])

    df = chart_df.copy()
    df["Date"] = _coerce_dates(df["Date"])
    df = df.dropna(subset=["Date"]).sort_values("Date")
    if df.empty:
        return pd.DataFrame(columns=["Segment", "Taux"])

    latest = df.iloc[-1]
    mapping = [
        ("Panel global", "Moyenne Panel Global"),
        ("Debutants", "Cohorte Débutants (<1k XP)"),
        ("Standard", "Cohorte Standard (1k-5k XP)"),
        ("Super-Actifs", "Cohorte Super-Actifs (>5k XP)"),
    ]

    rows: list[dict[str, float | str]] = []
    for label, column in mapping:
        value = parse_float(latest.get(column))
        if value is not None:
            rows.append({"Segment": label, "Taux": value})

    return pd.DataFrame(rows)


def _prepare_snapshot_df_from_log() -> pd.DataFrame:
    if not DAILY_LOG_FILE.exists():
        return pd.DataFrame(columns=["Segment", "Taux"])

    try:
        df = pd.read_csv(DAILY_LOG_FILE, usecols=["Date", "Username", "Cohort", "HasPlus"])
    except Exception:
        return pd.DataFrame(columns=["Segment", "Taux"])

    if df.empty:
        return pd.DataFrame(columns=["Segment", "Taux"])

    df = df[~df["Username"].astype(str).str.contains("Aggregated", na=False)].copy()
    df = df[df["Cohort"] != "Global"]
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    df = df.dropna(subset=["Date"])
    df = df.sort_values(["Date", "Username"]).drop_duplicates(subset=["Date", "Username"], keep="last")
    if df.empty:
        return pd.DataFrame(columns=["Segment", "Taux"])

    df["HasPlusValue"] = df["HasPlus"].apply(parse_bool_fraction)
    df = df.dropna(subset=["HasPlusValue"])
    if df.empty:
        return pd.DataFrame(columns=["Segment", "Taux"])

    counts = df.groupby("Date").size()
    valid_dates = counts[counts >= 100].index.tolist()
    latest_date = valid_dates[-1] if valid_dates else df["Date"].max()
    df = df[df["Date"] == latest_date]
    if df.empty:
        return pd.DataFrame(columns=["Segment", "Taux"])

    rows = [{"Segment": "Panel global", "Taux": round(df["HasPlusValue"].mean() * 100, 1)}]

    label_map = {
        "Debutants": "Debutants",
        "Standard": "Standard",
        "Super-Actifs": "Super-Actifs",
    }
    for cohort, label in label_map.items():
        subset = df[df["Cohort"] == cohort]
        if not subset.empty:
            rows.append({"Segment": label, "Taux": round(subset["HasPlusValue"].mean() * 100, 1)})

    return pd.DataFrame(rows)


def _prepare_daily_engagement_snapshot_from_log() -> dict[str, float | str | None]:
    if not DAILY_LOG_FILE.exists():
        return {}

    try:
        df = pd.read_csv(DAILY_LOG_FILE, usecols=["Date", "Username", "Cohort", "Streak"])
    except Exception:
        return {}

    if df.empty:
        return {}

    df = df[~df["Username"].astype(str).str.contains("Aggregated", na=False)].copy()
    df = df[df["Cohort"] != "Global"]
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    df["Streak"] = df["Streak"].apply(parse_float)
    df = df.dropna(subset=["Date", "Streak"])
    df = df.sort_values(["Date", "Username"]).drop_duplicates(subset=["Date", "Username"], keep="last")
    if df.empty:
        return {}

    dates = sorted(df["Date"].dropna().unique().tolist())
    if not dates:
        return {}

    latest_date = dates[-1]
    previous_date = dates[-2] if len(dates) > 1 else None
    current_df = df[df["Date"] == latest_date].copy()
    if current_df.empty:
        return {}

    panel_total = len(current_df)
    active_users = int((current_df["Streak"] > 0).sum())
    engagement_score = (active_users / panel_total) if panel_total else None
    avg_streak = current_df["Streak"].mean() if panel_total else None

    snapshot: dict[str, float | str | None] = {
        "date_label": str(latest_date),
        "panel_total": panel_total,
        "active_users": active_users,
        "engagement_score": engagement_score,
        "avg_streak": avg_streak,
        "reactivated_users": None,
        "abandon_global": None,
        "retention_rate": None,
        "abandon_debutants": None,
        "abandon_standard": None,
        "abandon_super_actifs": None,
    }

    if previous_date is None:
        return snapshot

    previous_df = df[df["Date"] == previous_date][["Username", "Cohort", "Streak"]].copy()
    current_metrics = current_df[["Username", "Cohort", "Streak"]].copy()
    merged = previous_df.merge(
        current_metrics,
        on="Username",
        how="inner",
        suffixes=("_prev", "_curr"),
    )
    if merged.empty:
        return snapshot

    snapshot["reactivated_users"] = int(((merged["Streak_prev"] == 0) & (merged["Streak_curr"] > 0)).sum())

    dropped_mask = (merged["Streak_prev"] > 0) & (merged["Streak_curr"] == 0)
    active_prev = int((merged["Streak_prev"] > 0).sum())
    if active_prev > 0:
        abandon_global = dropped_mask.sum() / active_prev
        snapshot["abandon_global"] = abandon_global
        snapshot["retention_rate"] = 1 - abandon_global

    cohort_col = "Cohort_curr" if "Cohort_curr" in merged.columns else "Cohort_prev"
    cohort_keys = {
        "Debutants": "abandon_debutants",
        "Standard": "abandon_standard",
        "Super-Actifs": "abandon_super_actifs",
    }
    for cohort_name, output_key in cohort_keys.items():
        cohort_df = merged[merged[cohort_col] == cohort_name]
        active_prev_cohort = int((cohort_df["Streak_prev"] > 0).sum())
        if active_prev_cohort > 0:
            snapshot[output_key] = (
                ((cohort_df["Streak_prev"] > 0) & (cohort_df["Streak_curr"] == 0)).sum()
                / active_prev_cohort
            )

    return snapshot


def _prepare_daily_engagement_snapshot(summary_df: pd.DataFrame) -> dict[str, float | str | None]:
    snapshot = _prepare_daily_engagement_snapshot_from_log()
    if snapshot:
        return snapshot

    normalized = _prepare_summary_metrics_df(summary_df)
    if normalized.empty:
        return {}

    latest = normalized.iloc[-1]
    panel_total = parse_float(latest.get("Panel Total"))
    engagement_score = parse_float(latest.get("Score d'Engagement"))
    active_users = None
    if panel_total is not None and engagement_score is not None:
        active_users = int(round(panel_total * engagement_score))

    abandon_global = parse_float(latest.get("Taux d'Abandon Global"))

    return {
        "date_label": latest["Date"].strftime("%Y-%m-%d"),
        "panel_total": panel_total,
        "active_users": active_users,
        "engagement_score": engagement_score,
        "reactivated_users": parse_float(latest.get("Reactivations vs Veille")),
        "avg_streak": parse_float(latest.get("Serie Moyenne (Jours)")),
        "abandon_global": abandon_global,
        "retention_rate": None if abandon_global is None else 1 - abandon_global,
        "abandon_debutants": parse_float(latest.get("Abandon Debutants")),
        "abandon_standard": parse_float(latest.get("Abandon Standard")),
        "abandon_super_actifs": parse_float(latest.get("Abandon Super-Actifs")),
    }


def _sheet_index(wb, reference_name: str, fallback: int) -> int:
    try:
        return wb.sheetnames.index(reference_name)
    except ValueError:
        return fallback


def _build_border() -> Border:
    return Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )


def _write_card(
    ws,
    header_range: str,
    value_range: str,
    title: str,
    value: str,
    header_fill: str,
    value_fill: str,
    white: str,
) -> None:
    ws.merge_cells(header_range)
    ws.merge_cells(value_range)

    thin = _build_border()

    header = ws[header_range.split(":")[0]]
    header.value = title
    header.fill = PatternFill("solid", start_color=header_fill, end_color=header_fill)
    header.font = Font(name="Calibri", size=10, bold=True, color=white)
    header.alignment = Alignment(horizontal="center", vertical="center")
    header.border = thin

    value_cell = ws[value_range.split(":")[0]]
    value_cell.value = value
    value_cell.fill = PatternFill("solid", start_color=value_fill, end_color=value_fill)
    value_cell.font = Font(name="Calibri", size=16, bold=True, color="1F1F1F")
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.border = thin


def _write_section_header(ws, cell_range: str, title: str, fill_color: str, white: str) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = title
    cell.fill = PatternFill("solid", start_color=fill_color, end_color=fill_color)
    cell.font = Font(name="Calibri", size=11, bold=True, color=white)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def refresh_trends_dashboard(report_path: Path) -> None:
    if not report_path.exists():
        raise FileNotFoundError(report_path)

    summary_df = _load_sheet(report_path, SUMMARY_SHEET)
    chart_df = _load_sheet(report_path, CHART_DATA_SHEET)

    global_df = _prepare_global_trend_df(summary_df)
    weekly_df = _prepare_weekly_comparison_df(summary_df)
    snapshot_df = _prepare_snapshot_df(chart_df)
    if snapshot_df.empty:
        snapshot_df = _prepare_snapshot_df_from_log()
    daily_snapshot = _prepare_daily_engagement_snapshot(summary_df)

    wb = load_workbook(report_path)

    if TREND_SHEET in wb.sheetnames:
        wb.remove(wb[TREND_SHEET])
    if CHART_DATA_SHEET in wb.sheetnames:
        wb.remove(wb[CHART_DATA_SHEET])

    chart_ws = wb.create_sheet(CHART_DATA_SHEET)
    trend_index = _sheet_index(wb, "📖 Dictionnaire des KPIs", len(wb.sheetnames))
    trend_ws = wb.create_sheet(TREND_SHEET, trend_index)

    chart_ws.append(["Date", "Taux Abonn. Super"])
    for row in global_df.itertuples(index=False):
        taux_percent = float(row[1]) * 100 if row[1] is not None and not pd.isna(row[1]) else None
        chart_ws.append([row.Libelle, taux_percent])

    chart_ws["D1"] = "Segment"
    chart_ws["E1"] = "Taux (%)"
    for idx, row in enumerate(snapshot_df.itertuples(index=False), start=2):
        chart_ws[f"D{idx}"] = row.Segment
        chart_ws[f"E{idx}"] = row.Taux

    chart_ws.sheet_state = "hidden"

    DUO_GREEN = "58CC02"
    DUO_BLUE = "1CB0F6"
    NAVY = "1F4E78"
    SOFT_BLUE = "EAF6FF"
    SOFT_GREEN = "EEF8E8"
    SOFT_ORANGE = "FFF4E9"
    SOFT_RED = "FFF1F1"
    SOFT_GREY = "F6F7FB"
    WHITE = "FFFFFF"

    trend_ws.sheet_view.showGridLines = False
    for col, width in {
        "A": 16, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14,
        "H": 14, "I": 3, "J": 16, "K": 16, "L": 16, "M": 18, "N": 18,
    }.items():
        trend_ws.column_dimensions[col].width = width

    trend_ws.merge_cells("A1:N2")
    title = trend_ws["A1"]
    title.value = "Monetisation & Tendances"
    title.fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
    title.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    title.alignment = Alignment(horizontal="center", vertical="center")

    trend_ws.merge_cells("A3:N3")
    subtitle = trend_ws["A3"]
    subtitle.value = "Lecture quotidienne de l'engagement et synthese Week over Week de la monetisation."
    subtitle.font = Font(name="Calibri", size=10, italic=True, color="5A5A5A")
    subtitle.alignment = Alignment(horizontal="center", vertical="center")

    _write_section_header(trend_ws, "A5:N5", "Bloc quotidien", NAVY, WHITE)
    _write_card(
        trend_ws,
        "A6:C6",
        "A7:C8",
        "Dernier releve",
        str(daily_snapshot.get("date_label", "N/A")),
        NAVY,
        SOFT_GREY,
        WHITE,
    )
    _write_card(
        trend_ws,
        "D6:F6",
        "D7:F8",
        "Panel observe",
        format_integer(daily_snapshot.get("panel_total")),
        DUO_BLUE,
        SOFT_BLUE,
        WHITE,
    )
    _write_card(
        trend_ws,
        "G6:I6",
        "G7:I8",
        "Utilisateurs actifs",
        format_integer(daily_snapshot.get("active_users")),
        "FF8A34",
        SOFT_ORANGE,
        WHITE,
    )
    _write_card(
        trend_ws,
        "J6:N6",
        "J7:N8",
        "Score d'engagement",
        format_percent(daily_snapshot.get("engagement_score")),
        DUO_GREEN,
        SOFT_GREEN,
        WHITE,
    )
    trend_ws.merge_cells("A9:N9")
    daily_line = trend_ws["A9"]
    daily_line.value = (
        f"Taux d'abandon global : {format_percent(daily_snapshot.get('abandon_global'))}"
        f"   |   Reactivations vs veille : {format_integer(daily_snapshot.get('reactivated_users'))}"
        f"   |   Retention vs hier : {format_percent(daily_snapshot.get('retention_rate'))}"
        f"   |   Serie moyenne : {format_streak(daily_snapshot.get('avg_streak'))}"
    )
    daily_line.font = Font(name="Calibri", size=10, bold=True, color="4A4A4A")
    daily_line.alignment = Alignment(horizontal="center", vertical="center")

    trend_ws.merge_cells("A10:N10")
    cohort_line = trend_ws["A10"]
    cohort_line.value = (
        "Abandon par cohorte : "
        f"Debutants {format_percent(daily_snapshot.get('abandon_debutants'))}"
        f"   |   Standard {format_percent(daily_snapshot.get('abandon_standard'))}"
        f"   |   Super-Actifs {format_percent(daily_snapshot.get('abandon_super_actifs'))}"
    )
    cohort_line.font = Font(name="Calibri", size=9, italic=True, color="666666")
    cohort_line.alignment = Alignment(horizontal="center", vertical="center")

    _write_section_header(trend_ws, "A12:H12", "Evolution recente de la conversion", NAVY, WHITE)
    trend_ws.merge_cells("A13:H13")
    chart_help = trend_ws["A13"]
    chart_help.value = (
        "Courbe quotidienne du taux d'abonnement Super sur le panel suivi. "
        "Elle sert de signal de monetisation dans le temps."
    )
    chart_help.font = Font(name="Calibri", size=9, italic=True, color="666666")
    chart_help.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if len(global_df) >= 2:
        trend_chart = LineChart()
        trend_chart.style = 2
        trend_chart.height = 6.8
        trend_chart.width = 9.6
        trend_chart.legend = None
        trend_chart.title = None
        trend_chart.y_axis.title = None
        trend_chart.x_axis.title = None

        series_percent = global_df["Taux Abonn. Super"] * 100
        min_val = float(series_percent.min())
        max_val = float(series_percent.max())

        trend_chart.y_axis.scaling.min = max(0, floor((min_val - 2) / 5) * 5)
        trend_chart.y_axis.scaling.max = ceil((max_val + 2) / 5) * 5
        trend_chart.x_axis.tickLblPos = "low"
        trend_chart.x_axis.delete = False
        trend_chart.y_axis.delete = False
        trend_chart.x_axis.majorTickMark = "none"
        trend_chart.y_axis.majorTickMark = "none"
        trend_chart.plot_area.layout = Layout(
            manualLayout=ManualLayout(
                x=0.12,
                y=0.08,
                h=0.72,
                w=0.82,
                xMode="edge",
                yMode="edge",
            )
        )

        data = Reference(chart_ws, min_col=2, min_row=1, max_row=len(global_df) + 1)
        categories = Reference(chart_ws, min_col=1, min_row=2, max_row=len(global_df) + 1)
        trend_chart.add_data(data, titles_from_data=True)
        trend_chart.set_categories(categories)

        series = trend_chart.series[0]
        series.graphicalProperties.line.solidFill = DUO_BLUE
        series.graphicalProperties.line.width = 24000
        series.marker.symbol = "circle"
        series.marker.size = 7
        series.marker.graphicalProperties.solidFill = DUO_BLUE
        series.marker.graphicalProperties.line.solidFill = DUO_BLUE

        trend_ws.add_chart(trend_chart, "A15")
    else:
        trend_ws.merge_cells("A15:H20")
        empty_msg = trend_ws["A15"]
        empty_msg.value = "Pas assez d'historique pour afficher une tendance."
        empty_msg.font = Font(name="Calibri", size=11, italic=True, color="666666")
        empty_msg.alignment = Alignment(horizontal="center", vertical="center")

    _write_section_header(trend_ws, "J12:N12", "Lecture par segment", DUO_GREEN, WHITE)
    trend_ws.merge_cells("J13:N13")
    segment_help = trend_ws["J13"]
    segment_help.value = "Taux premium du dernier releve complet par segment, compares au panel global."
    segment_help.font = Font(name="Calibri", size=9, italic=True, color="666666")
    segment_help.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    panel_rate = None
    if not snapshot_df.empty:
        panel_row = snapshot_df[snapshot_df["Segment"] == "Panel global"]
        panel_rate = panel_row["Taux"].iloc[0] if not panel_row.empty else None

    table_headers = ["Segment", "Taux premium", "Ecart vs panel", "Lecture"]
    table_cols = ["J", "K", "L", "M"]
    for col, title_value in zip(table_cols, table_headers):
        cell = trend_ws[f"{col}15"]
        cell.value = title_value
        cell.fill = PatternFill("solid", start_color=SOFT_GREY, end_color=SOFT_GREY)
        cell.font = Font(name="Calibri", size=10, bold=True, color="333333")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _build_border()

    if not snapshot_df.empty:
        display_df = snapshot_df.copy()
        display_df["Gap"] = display_df["Taux"] - panel_rate if panel_rate is not None else None

        panel_df = display_df[display_df["Segment"] == "Panel global"]
        cohorts_df = display_df[display_df["Segment"] != "Panel global"].sort_values("Taux", ascending=False)
        display_df = pd.concat([panel_df, cohorts_df], ignore_index=True)

        fill_map = {
            "Panel global": "EAF6FF",
            "Super-Actifs": "EEF8E8",
            "Standard": "FFF7DB",
            "Debutants": "FFF1F1",
        }

        for row_idx, row in enumerate(display_df.itertuples(index=False), start=16):
            segment = row.Segment
            taux = row.Taux
            gap = row.Gap
            if segment == "Panel global":
                lecture = "Reference"
            elif gap is not None and gap >= 1:
                lecture = "Au-dessus du panel"
            elif gap is not None and gap <= -1:
                lecture = "Sous le panel"
            else:
                lecture = "Proche du panel"

            values = [
                segment,
                f"{float(taux):.1f}%" if taux is not None and not pd.isna(taux) else "N/A",
                f"{float(gap):+.1f} pts" if gap is not None and not pd.isna(gap) else "N/A",
                lecture,
            ]
            for col, value in zip(table_cols, values):
                cell = trend_ws[f"{col}{row_idx}"]
                cell.value = value
                cell.fill = PatternFill(
                    "solid",
                    start_color=fill_map.get(segment, WHITE),
                    end_color=fill_map.get(segment, WHITE),
                )
                cell.font = Font(name="Calibri", size=10, bold=(col == "J"), color="333333")
                if col == "L" and gap is not None:
                    cell.font = Font(name="Calibri", size=10, bold=True, color=delta_font_color(gap))
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = _build_border()
    else:
        trend_ws.merge_cells("J16:N19")
        empty_msg = trend_ws["J16"]
        empty_msg.value = "Pas assez de donnees pour comparer les segments."
        empty_msg.font = Font(name="Calibri", size=11, italic=True, color="666666")
        empty_msg.alignment = Alignment(horizontal="center", vertical="center")

    trend_ws.merge_cells("A28:N28")
    note = trend_ws["A28"]
    note.value = (
        "Lecture: le bloc quotidien suit l'engagement du dernier jour collecte. "
        "La courbe suit la conversion Super dans le temps. "
        "Le tableau de droite compare le dernier releve premium par segment au panel global."
    )
    note.font = Font(name="Calibri", size=10, italic=True, color="666666")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    _write_section_header(trend_ws, "A30:N30", "Lecture Week over Week", NAVY, WHITE)
    trend_ws.merge_cells("A31:N31")
    weekly_help = trend_ws["A31"]
    weekly_help.value = (
        "Cette section compare les moyennes hebdomadaires du resume quotidien. "
        "Les ecarts se lisent vs la semaine precedente."
    )
    weekly_help.font = Font(name="Calibri", size=9, italic=True, color="666666")
    weekly_help.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    latest_week = weekly_df.iloc[-1] if not weekly_df.empty else None
    latest_week_label = latest_week["Semaine"] if latest_week is not None else "N/A"
    latest_week_xp = latest_week["XP moyen"] if latest_week is not None else None
    latest_week_conv = latest_week["Conv. Super moy."] if latest_week is not None else None
    latest_week_abandon = latest_week["Taux d'abandon moyen"] if latest_week is not None else None
    latest_week_panel = latest_week["Panel moyen"] if latest_week is not None else None

    delta_xp = latest_week["Delta XP vs S-1"] if latest_week is not None else None
    delta_conv = latest_week["Delta Conv. vs S-1"] if latest_week is not None else None
    delta_abandon = latest_week["Delta Abandon vs S-1"] if latest_week is not None else None
    delta_score = latest_week["Delta Score vs S-1"] if latest_week is not None else None

    _write_card(trend_ws, "A33:C33", "A34:C35", "Semaine", str(latest_week_label), NAVY, SOFT_GREY, WHITE)
    _write_card(trend_ws, "D33:F33", "D34:F35", "XP moyen / jour", format_xp(latest_week_xp), DUO_BLUE, SOFT_BLUE, WHITE)
    _write_card(trend_ws, "G33:I33", "G34:I35", "Conv. Super moy.", format_percent(latest_week_conv), DUO_GREEN, SOFT_GREEN, WHITE)
    _write_card(trend_ws, "J33:L33", "J34:L35", "Taux d'abandon moy.", format_percent(latest_week_abandon), "FF6B6B", SOFT_RED, WHITE)
    _write_card(trend_ws, "M33:N33", "M34:N35", "Panel moy.", format_integer(latest_week_panel), NAVY, SOFT_GREY, WHITE)

    _write_card(trend_ws, "A37:C37", "A38:C39", "Delta XP vs S-1", format_xp_delta(delta_xp), "FF8A34", SOFT_ORANGE, WHITE)
    trend_ws["A38"].font = Font(name="Calibri", size=16, bold=True, color=delta_font_color(delta_xp))
    _write_card(trend_ws, "D37:F37", "D38:F39", "Delta Conv. vs S-1", format_points(delta_conv), DUO_BLUE, SOFT_BLUE, WHITE)
    trend_ws["D38"].font = Font(name="Calibri", size=16, bold=True, color=delta_font_color(delta_conv))
    _write_card(trend_ws, "G37:I37", "G38:I39", "Delta Abandon vs S-1", format_points(delta_abandon), "FF6B6B", SOFT_RED, WHITE)
    trend_ws["G38"].font = Font(name="Calibri", size=16, bold=True, color=delta_font_color(delta_abandon, positive_is_good=False))
    _write_card(trend_ws, "J37:L37", "J38:L39", "Delta Score vs S-1", format_points(delta_score), DUO_GREEN, SOFT_GREEN, WHITE)
    trend_ws["J38"].font = Font(name="Calibri", size=16, bold=True, color=delta_font_color(delta_score))
    _write_card(trend_ws, "M37:N37", "M38:N39", "Semaines visibles", str(len(weekly_df)) if not weekly_df.empty else "0", NAVY, SOFT_GREY, WHITE)

    trend_ws.merge_cells("A41:N41")
    wow_insight = trend_ws["A41"]
    wow_insight.value = weekly_insight_text(latest_week)
    wow_insight.fill = PatternFill("solid", start_color=SOFT_GREY, end_color=SOFT_GREY)
    wow_insight.font = Font(name="Calibri", size=10, bold=True, color="444444")
    wow_insight.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    trend_ws.merge_cells("A43:N43")
    weekly_table_title = trend_ws["A43"]
    weekly_table_title.value = "Historique hebdomadaire consolide"
    weekly_table_title.fill = PatternFill("solid", start_color=SOFT_GREY, end_color=SOFT_GREY)
    weekly_table_title.font = Font(name="Calibri", size=10, bold=True, color="333333")
    weekly_table_title.alignment = Alignment(horizontal="center", vertical="center")

    weekly_headers = [
        "Semaine",
        "Jours observes",
        "XP moyen",
        "Conv. Super",
        "Taux d'abandon",
        "Score d'eng.",
        "Panel moyen",
    ]
    weekly_cols = ["A", "C", "E", "G", "I", "K", "M"]
    for col, title_value in zip(weekly_cols, weekly_headers):
        start = f"{col}45"
        end = f"{chr(ord(col) + 1)}45"
        trend_ws.merge_cells(f"{start}:{end}")
        cell = trend_ws[start]
        cell.value = title_value
        cell.fill = PatternFill("solid", start_color=SOFT_GREY, end_color=SOFT_GREY)
        cell.font = Font(name="Calibri", size=10, bold=True, color="333333")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _build_border()

    if not weekly_df.empty:
        latest_week_index = len(weekly_df) - 1
        for idx, row in enumerate(weekly_df.itertuples(index=False)):
            offset = 46 + idx
            values = [
                row[0],
                str(int(row[1])) if not pd.isna(row[1]) else "N/A",
                format_xp(row[2]),
                format_percent(row[4]),
                format_percent(row[6]),
                format_percent(row[8]),
                format_integer(row[10]),
            ]
            is_latest_week = idx == latest_week_index
            base_fill = SOFT_BLUE if is_latest_week else (WHITE if offset % 2 else SOFT_GREY)
            for col, value in zip(weekly_cols, values):
                start = f"{col}{offset}"
                end = f"{chr(ord(col) + 1)}{offset}"
                trend_ws.merge_cells(f"{start}:{end}")
                cell = trend_ws[start]
                cell.value = value
                cell.fill = PatternFill("solid", start_color=base_fill, end_color=base_fill)
                cell.font = Font(name="Calibri", size=10, bold=(col == "A" or is_latest_week), color="333333")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = _build_border()
    else:
        trend_ws.merge_cells("A46:N48")
        empty_msg = trend_ws["A46"]
        empty_msg.value = "Le comparatif Week over Week apparaitra ici des qu'au moins une semaine de resume sera disponible."
        empty_msg.font = Font(name="Calibri", size=10, italic=True, color="666666")
        empty_msg.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save(report_path)
