"""
Partie 2 : Statistiques
Calcule les statistiques d'engagement via Pandas.
"""
import csv
from datetime import datetime, timedelta
from pathlib import Path
import shutil
import numbers
import re
import pandas as pd

from .config import DAILY_LOG_FILE, GOOGLE_DRIVE_REPORT_DIR, RAPPORT_EXCEL_FILE, REPORT_DIR, now_toronto
from .excel_dashboard import refresh_trends_dashboard
from .financial_signals import build_financial_signal_sheet_df
from .quarterly_nowcast import build_quarterly_nowcast_raw_df
from .reporting.sheets.dcf_valuation_sheet import render_dcf_valuation_sheet
from .reporting.sheets.financial_nowcast_sheet import render_financial_nowcast_sheet
from .reporting.sheets.kpi_dictionary_sheet import build_kpi_dictionary_df, render_kpi_dictionary_sheet
from .reporting.sheets.monthly_trends_sheet import add_monthly_trends_chart, build_monthly_trends_frames
from .reporting.sheets.quarterly_nowcast_sheet import render_quarterly_nowcast_sheet
from .reporting.sheets.summary_sheet import build_summary_today_df, merge_summary_history
from .reporting.styles import build_style_context
from .reporting.workbook_postprocess import apply_standard_table_style, hide_sheets, remove_sheets, reorder_sheets
from .subscription_detection import (
    apply_max_overrides,
    compute_max_count,
    compute_super_observable_count,
    parse_bool,
    parse_optional_bool,
)

SUMMARY_SHEET = "Suivi Quotidien"
AI_SHEET = "🤖 Analyse Stratégique"
GLOSSAIRE_SHEET = "📖 Dictionnaire des KPIs"
GLOSSAIRE_RAW_SHEET = "Dictionnaire des KPIs - Raw"
TRENDS_SHEET = "📈 Tendances Mensuelles"
CHART_DATA_SHEET = "📊 Données Graphique"

SIGNALS_SHEET = "Signaux Financiers"
SIGNALS_RAW_SHEET = "Signaux Financiers - Raw"
QUARTERLY_SHEET = "Nowcast Trimestriel"
QUARTERLY_RAW_SHEET = "Nowcast Trimestriel - Raw"
DCF_SHEET = "Valorisation DCF"

PERCENT_COLUMNS = {
    "Taux Abonn. Super",
    "Taux d'Abandon Global",
    "Score d'Engagement",
}

BAD_SHEET_NAMES = {
    "ðŸ“Š RÃ©sumÃ© Financier Q1",
    "ðŸ¤– Analyse StratÃ©gique",
    "ðŸ“– Dictionnaire des KPIs",
    "ðŸ“ˆ Tendances Mensuelles",
    "ðŸ“Š DonnÃ©es Graphique",
}

SUMMARY_COLUMNS = [
    "Date",
    "Série Moyenne (Jours)",
    "Évol. vs Veille",
    "Apprentissage (XP/j)",
    "Taux Abonn. Super",
    "Taux d'Abandon Global",
    "Reactivations vs Veille",
    "Score d'Engagement",
    "Panel Total",
]

SUMMARY_COLUMN_ALIASES = {
    "Date": "Date",
    "Moyenne Streak (J)": "Série Moyenne (Jours)",
    "Série Moyenne (Jours)": "Série Moyenne (Jours)",
    "Évolution vs Hier": "Évol. vs Veille",
    "Évol. vs Veille": "Évol. vs Veille",
    "Delta XP (Intensité)": "Apprentissage (XP/j)",
    "Apprentissage (XP/j)": "Apprentissage (XP/j)",
    "Conversion Premium": "Taux Abonn. Super",
    "Taux Abonn. Super": "Taux Abonn. Super",
    "Taux Abonn. Max": "Taux Abonn. Max",
    "Churn Global": "Taux d'Abandon Global",
    "Taux d'Abandon Global": "Taux d'Abandon Global",
    "Taux d'Attrition Global": "Taux d'Abandon Global",
    "Reactivations vs Veille": "Reactivations vs Veille",
    "Score Santé Global": "Score d'Engagement",
    "Score d'Engagement": "Score d'Engagement",
    "Total Profils": "Panel Total",
    "Panel Total": "Panel Total",
}

DAILY_LOG_COLUMNS = ["Date", "Username", "Cohort", "Streak", "TotalXP", "HasPlus", "HasMax"]
SUMMARY_MIN_RELIABLE_DATE = pd.Timestamp("2026-03-16")


def _parse_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, numbers.Number):
        return float(value)
    if isinstance(value, str):
        raw = value.strip()
        if not raw or raw.upper() in {"N/A", "NA"}:
            return None
        cleaned = raw.replace(",", ".")
        cleaned = cleaned.replace("XP", "").replace("xp", "")
        cleaned = cleaned.replace("%", "")
        cleaned = cleaned.replace("+", "")
        cleaned = cleaned.replace("−", "-")
        cleaned = "".join(ch for ch in cleaned if ch.isdigit() or ch in ".-")
        if cleaned in {"", "-", "."}:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _normalize_summary_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    normalized = pd.DataFrame(index=df.index)

    for column in SUMMARY_COLUMNS:
        normalized[column] = pd.Series(index=df.index, dtype="object")

    for source_column in df.columns:
        target_column = SUMMARY_COLUMN_ALIASES.get(str(source_column).strip())
        if not target_column:
            continue
        if target_column not in normalized.columns:
            continue

        source_series = df[source_column]

        if target_column == "Date":
            parsed_series = pd.to_datetime(source_series, errors="coerce")
        else:
            parsed_series = source_series.apply(_parse_float)

            if target_column in PERCENT_COLUMNS:
                parsed_series = parsed_series.apply(
                    lambda x: (x / 100) if isinstance(x, numbers.Number) and abs(x) > 1 else x
                )

        fill_mask = normalized[target_column].isna()
        normalized.loc[fill_mask, target_column] = parsed_series[fill_mask]

    if "Date" in normalized.columns:
        normalized = normalized[normalized["Date"].notna()]
        normalized = normalized[normalized["Date"] >= SUMMARY_MIN_RELIABLE_DATE]
    if "Panel Total" in normalized.columns:
        panel_total_numeric = pd.to_numeric(normalized["Panel Total"], errors="coerce")
        normalized = normalized[panel_total_numeric.fillna(0) > 0]
        normalized["Panel Total"] = panel_total_numeric.loc[normalized.index]
    if "Date" in normalized.columns:
        normalized["_panel_total_rank"] = pd.to_numeric(normalized.get("Panel Total"), errors="coerce").fillna(-1)
        normalized["_completeness_rank"] = normalized.notna().sum(axis=1)
        normalized = (
            normalized.sort_values(["Date", "_panel_total_rank", "_completeness_rank"])
            .drop_duplicates(subset=["Date"], keep="last")
            .drop(columns=["_panel_total_rank", "_completeness_rank"], errors="ignore")
        )

    return normalized.reset_index(drop=True)


def _load_daily_log_df() -> pd.DataFrame:
    rows: list[dict[str, object]] = []

    with open(DAILY_LOG_FILE, "r", encoding="utf-8", newline="") as source:
        reader = csv.reader(source)
        header = next(reader, None)
        if header is None:
            return pd.DataFrame(columns=DAILY_LOG_COLUMNS)

        for row in reader:
            if not row:
                continue
            if row == DAILY_LOG_COLUMNS or row[:6] == DAILY_LOG_COLUMNS[:6]:
                continue

            normalized_row = row[:]
            if len(normalized_row) == 6:
                normalized_row.append("")
            elif len(normalized_row) < 6:
                continue
            elif len(normalized_row) > 7:
                normalized_row = normalized_row[:7]

            if len(normalized_row) < 7:
                normalized_row.extend([""] * (7 - len(normalized_row)))

            rows.append(dict(zip(DAILY_LOG_COLUMNS, normalized_row)))

    df = pd.DataFrame(rows, columns=DAILY_LOG_COLUMNS)
    if df.empty:
        return df

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    df["Streak"] = pd.to_numeric(df["Streak"], errors="coerce").fillna(0)
    df["TotalXP"] = pd.to_numeric(df["TotalXP"], errors="coerce").fillna(0)
    df["HasPlus"] = df["HasPlus"].apply(parse_bool)
    df["HasMax"] = df["HasMax"].apply(parse_optional_bool)
    df = apply_max_overrides(df)
    df = df[df["Date"].notna()]
    return df


def _compute_super_rate_pct(df_jour: pd.DataFrame) -> float | None:
    if df_jour.empty or "HasPlus" not in df_jour.columns:
        return None
    super_count = compute_super_observable_count(df_jour["HasPlus"], df_jour.get("HasMax"))
    return (super_count / len(df_jour)) * 100 if len(df_jour) else None


def _compute_max_rate_pct(df_jour: pd.DataFrame) -> float | None:
    if df_jour.empty or "HasMax" not in df_jour.columns:
        return None
    max_count = compute_max_count(df_jour["HasMax"])
    if max_count is None:
        return None
    return (max_count / len(df_jour)) * 100 if len(df_jour) else None


def _build_summary_history_from_log(df: pd.DataFrame) -> pd.DataFrame | None:
    if df is None or df.empty:
        return None

    df = df.copy()
    df = df[~df["Username"].str.contains("Aggregated", na=False)]
    df = df[df["Cohort"] != "Global"]
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df[df["Date"].notna()]
    if df.empty:
        return None

    df = df.sort_values(["Date", "Username"]).drop_duplicates(subset=["Date", "Username"], keep="last")
    available_dates = sorted(df["Date"].dt.strftime("%Y-%m-%d").unique().tolist())
    if not available_dates:
        return None

    rows: list[dict[str, object]] = []
    for date_str in available_dates:
        date_obj = pd.to_datetime(date_str, errors="coerce")
        prev_date_str = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d") if pd.notna(date_obj) else None

        df_jour = df[df["Date"].dt.strftime("%Y-%m-%d") == date_str].copy()
        df_hier = df[df["Date"].dt.strftime("%Y-%m-%d") == prev_date_str].copy() if prev_date_str else pd.DataFrame()

        moyenne_streak_jour = float(df_jour["Streak"].mean()) if not df_jour.empty else 0.0
        moyenne_streak_hier = float(df_hier["Streak"].mean()) if not df_hier.empty else None
        delta_streak = round(moyenne_streak_jour - moyenne_streak_hier, 1) if moyenne_streak_hier is not None else None

        taux_super = _compute_super_rate_pct(df_jour)
        taux_max = _compute_max_rate_pct(df_jour)

        delta_xp_moyen = 0.0
        taux_churn = 0.0
        reactivations_veille = 0
        if not df_jour.empty and not df_hier.empty:
            merged = df_hier.merge(
                df_jour,
                on="Username",
                suffixes=("_hier", "_jour"),
                how="inner",
            )
            if "TotalXP_hier" in merged.columns and "TotalXP_jour" in merged.columns and not merged.empty:
                merged["Delta_XP"] = merged["TotalXP_jour"] - merged["TotalXP_hier"]
                merged.loc[merged["Delta_XP"] < 0, "Delta_XP"] = 0
                delta_xp_moyen = float(merged["Delta_XP"].mean()) if not merged.empty else 0.0

            tombes_a_zero = merged[(merged["Streak_hier"] > 0) & (merged["Streak_jour"] == 0)]
            actifs_hier = int(len(df_hier[df_hier["Streak"] > 0]))
            taux_churn = ((len(tombes_a_zero) / actifs_hier) * 100) if actifs_hier > 0 else 0.0
            reactivations_veille = int(((merged["Streak_hier"] == 0) & (merged["Streak_jour"] > 0)).sum())

        utilisateurs_actifs = int(len(df_jour[df_jour["Streak"] > 0]))
        score_engagement = ((utilisateurs_actifs / len(df_jour)) * 100) if len(df_jour) > 0 else 0.0

        rows.append(
            {
                "Date": date_obj,
                "Série Moyenne (Jours)": round(moyenne_streak_jour, 1),
                "Évol. vs Veille": delta_streak,
                "Apprentissage (XP/j)": round(delta_xp_moyen, 0),
                "Taux Abonn. Super": round(taux_super / 100, 6) if isinstance(taux_super, numbers.Number) else None,
                "Taux d'Abandon Global": round(taux_churn / 100, 6),
                "Reactivations vs Veille": reactivations_veille,
                "Score d'Engagement": round(score_engagement / 100, 6),
                "Panel Total": int(len(df_jour)),
            }
        )

    return pd.DataFrame(rows, columns=SUMMARY_COLUMNS)


def _load_summary_sheet(report_path: Path) -> pd.DataFrame | None:
    try:
        workbook = pd.ExcelFile(report_path)
    except Exception:
        return None

    for sheet_name in [SUMMARY_SHEET, *BAD_SHEET_NAMES]:
        if sheet_name not in workbook.sheet_names:
            continue
        try:
            df_summary = pd.read_excel(report_path, sheet_name=sheet_name)
            if df_summary is not None and not df_summary.empty:
                return df_summary
        except Exception:
            continue

    for sheet_name in workbook.sheet_names:
        try:
            df_preview = pd.read_excel(report_path, sheet_name=sheet_name, nrows=5)
        except Exception:
            continue

        normalized_headers = {
            SUMMARY_COLUMN_ALIASES.get(str(column).strip())
            for column in df_preview.columns
        }
        if {"Date", "Panel Total"}.issubset(normalized_headers):
            try:
                df_summary = pd.read_excel(report_path, sheet_name=sheet_name)
                if df_summary is not None and not df_summary.empty:
                    return df_summary
            except Exception:
                continue

    return None


def _charger_resume_historique() -> pd.DataFrame | None:
    frames: list[pd.DataFrame] = []

    try:
        df_log = _load_daily_log_df()
        df_from_log = _build_summary_history_from_log(df_log)
        if df_from_log is not None and not df_from_log.empty:
            frames.append(df_from_log)
    except Exception:
        pass

    if RAPPORT_EXCEL_FILE.exists():
        try:
            df_resume = _load_summary_sheet(RAPPORT_EXCEL_FILE)
            if df_resume is not None and not df_resume.empty:
                frames.append(df_resume)
        except Exception:
            pass

    try:
        daily_reports = sorted(REPORT_DIR.glob("rapport_*.xlsx"))
    except Exception:
        daily_reports = []

    for report_path in daily_reports:
        if report_path == RAPPORT_EXCEL_FILE:
            continue
        try:
            df_tmp = _load_summary_sheet(report_path)
            if df_tmp is not None and not df_tmp.empty:
                frames.append(df_tmp)
        except Exception:
            continue

    if not frames:
        return None

    df_resume = pd.concat(frames, ignore_index=True)
    return _normalize_summary_df(df_resume)


def _copier_rapport_vers_google_drive(report_path: Path) -> Path | None:
    if not GOOGLE_DRIVE_REPORT_DIR:
        return None

    destination_dir = Path(GOOGLE_DRIVE_REPORT_DIR).expanduser()
    try:
        destination_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        print(f"  ⚠️ Impossible de créer le dossier Google Drive : {destination_dir} ({exc})")
        return None

    destination_file = destination_dir / report_path.name
    try:
        shutil.copy2(report_path, destination_file)
    except Exception as exc:
        print(f"  ⚠️ Copie Google Drive impossible : {destination_file} ({exc})")
        return None

    return destination_file


def _pretty_fr_number(value: object, digits: int = 1) -> str:
    if value is None or pd.isna(value):
        return "N/D"
    try:
        numeric = float(value)
    except Exception:
        return str(value)

    if digits == 0:
        return f"{int(round(numeric)):,}".replace(",", " ")
    return f"{numeric:,.{digits}f}".replace(",", " ").replace(".", ",")


def _pretty_ratio_pct(value: object, digits: int = 1) -> str:
    if value is None or pd.isna(value):
        return "N/D"
    try:
        return f"{float(value) * 100:.{digits}f}%".replace(".", ",")
    except Exception:
        return "N/D"


def _pretty_delta_pts(value: object, digits: int = 1) -> str:
    if value is None or pd.isna(value):
        return "N/D"
    try:
        return f"{float(value) * 100:+.{digits}f} pts".replace(".", ",")
    except Exception:
        return "N/D"


def _pretty_score(value: object, digits: int = 1) -> str:
    if value is None or pd.isna(value):
        return "N/D"
    try:
        return f"{float(value):.{digits}f} / 100".replace(".", ",")
    except Exception:
        return "N/D"


def _format_estimation_vs_guidance_note(
    estimated_value: object,
    guidance_value: object,
    *,
    prefix: str = "Est.",
    guidance_label: str = "Guidance",
) -> str:
    estimated_text = (
        f"{_pretty_fr_number(estimated_value, 1)} M$"
        if isinstance(estimated_value, numbers.Number)
        else "N/D"
    )
    guidance_text = (
        f"{_pretty_fr_number(guidance_value, 1)} M$"
        if isinstance(guidance_value, numbers.Number)
        else "N/D"
    )
    return f"{prefix} {estimated_text} vs {guidance_label} {guidance_text}"


def _build_quarterly_model_explainer(
    *,
    revenue_prob: object,
    guidance_prob: object,
    revenue_reference: object,
    drivers: list[object],
    risks: list[object],
) -> str:
    primary_driver = str(drivers[0]).lstrip("- ").rstrip(".") if drivers else "la dynamique récente du panel"
    primary_risk = str(risks[0]).lstrip("- ").rstrip(".") if risks else "les limites actuelles de calibration"

    if isinstance(revenue_reference, numbers.Number):
        reference_text = f"{_pretty_fr_number(revenue_reference, 1)} M$"
        revenue_sentence = (
            f"La probabilité de beat revenus ressort à {_pretty_ratio_pct(revenue_prob, 1)} ; "
            f"elle mesure la chance implicite de dépasser la référence interne du trimestre, "
            f"actuellement ancrée sur la guidance management de {reference_text}."
        )
        guidance_sentence = (
            f"La probabilité de guidance raise ressort à {_pretty_ratio_pct(guidance_prob, 1)} et reste surtout portée par {primary_driver}."
            if _parse_float(guidance_prob) is not None and _parse_float(guidance_prob) >= 0.5
            else f"La probabilité de guidance raise ressort à {_pretty_ratio_pct(guidance_prob, 1)} et reste freinée par {primary_risk}."
        )
        return (
            "Notre modèle trimestriel combine la monétisation, l'engagement, la rétention, le churn, "
            "les réactivations et la couverture du panel. "
            + revenue_sentence
            + " "
            + guidance_sentence
        )

    return (
        "Notre modèle trimestriel combine la monétisation, l'engagement, la rétention, le churn, "
        "les réactivations et la couverture du panel. "
        f"La probabilité de beat revenus ressort à {_pretty_ratio_pct(revenue_prob, 1)} et s'appuie encore sur une référence interne de transition. "
        f"La probabilité de guidance raise ressort à {_pretty_ratio_pct(guidance_prob, 1)} et reste pour l'instant surtout influencée par {primary_risk}."
    )


def _build_quarterly_model_explainer(
    *,
    revenue_prob: object,
    guidance_prob: object,
    revenue_reference: object,
    drivers: list[object],
    risks: list[object],
) -> str:
    primary_driver = str(drivers[0]).lstrip("- ").rstrip(".") if drivers else "la dynamique récente du panel"
    primary_risk = str(risks[0]).lstrip("- ").rstrip(".") if risks else "les limites actuelles de calibration"

    base_sentence = (
        "Notre modèle trimestriel combine la monétisation, l'engagement, la rétention, "
        "le churn, les réactivations et la couverture du panel."
    )

    if isinstance(revenue_reference, numbers.Number):
        reference_text = f"{_pretty_fr_number(revenue_reference, 1)} M$"
        revenue_sentence = (
            f"La probabilité implicite de battre les revenus du trimestre ressort à {_pretty_ratio_pct(revenue_prob, 1)} ; "
            f"elle compare notre estimation à la référence interne du trimestre, actuellement ancrée sur la guidance management de {reference_text}."
        )
    else:
        revenue_sentence = (
            f"La probabilité implicite de battre les revenus du trimestre ressort à {_pretty_ratio_pct(revenue_prob, 1)} ; "
            "elle s'appuie encore sur une référence interne de transition, faute de guidance exploitable dans l'historique."
        )

    guidance_sentence = (
        f"La probabilité implicite d'un relèvement de guidance ressort à {_pretty_ratio_pct(guidance_prob, 1)} et reste surtout portée par {primary_driver}."
        if _parse_float(guidance_prob) is not None and _parse_float(guidance_prob) >= 0.5
        else f"La probabilité implicite d'un relèvement de guidance ressort à {_pretty_ratio_pct(guidance_prob, 1)} et reste freinée par {primary_risk}."
    )

    return base_sentence + " " + revenue_sentence + " " + guidance_sentence


def _label_signal_bias(value: object) -> str:
    mapping = {
        "favorable": "Favorable",
        "neutral": "Neutre",
        "unfavorable": "Defavorable",
    }
    return mapping.get(str(value or "").strip().lower(), "N/D")


def _label_confidence(value: object) -> str:
    mapping = {
        "high": "Elevee",
        "medium": "Moyenne",
        "low": "Faible",
    }
    return mapping.get(str(value or "").strip().lower(), "N/D")


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        value = value.replace("**", "")
    text = str(value).replace("\r", "\n").replace("•", "-")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def _truncate_text(text: str, max_chars: int) -> str:
    clean = _normalize_text(text)
    if len(clean) <= max_chars:
        return clean
    shortened = clean[: max_chars - 1].rstrip(" ,;:-")
    return f"{shortened}…"


def _compact_summary_text(text: object, max_sentences: int = 2, max_chars: int = 180) -> str:
    clean = _normalize_text(text)
    if not clean:
        return "-"
    sentences = [part.strip() for part in re.split(r"(?<=[.!?])\s+", clean) if part.strip()]
    if not sentences:
        return clean

    selected: list[str] = []
    current_length = 0
    for sentence in sentences:
        candidate_length = current_length + (1 if selected else 0) + len(sentence)
        if selected and candidate_length > max_chars:
            break
        selected.append(sentence)
        current_length = candidate_length
        if len(selected) >= max_sentences:
            break

    return " ".join(selected) if selected else sentences[0]


def _compact_bullet_text(text: object, max_items: int = 2, max_chars: int = 95) -> str:
    clean = _normalize_text(text)
    if not clean:
        return "-"

    items: list[str] = []
    for raw_line in clean.splitlines():
        candidate = raw_line.lstrip("-* ").strip()
        if candidate:
            items.append(candidate)

    if not items:
        items = [part.strip() for part in re.split(r"(?<=[.!?])\s+", clean) if part.strip()]

    compact_items: list[str] = []
    current_length = 0
    for item in items:
        candidate_length = current_length + (1 if compact_items else 0) + len(item)
        if compact_items and candidate_length > max_chars:
            break
        if item and item not in compact_items:
            compact_items.append(item)
            current_length = candidate_length
        if len(compact_items) >= max_items:
            break

    if not compact_items:
        return "-"
    return "\n".join(f"- {item}" for item in compact_items)


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        value = value.replace("**", "")
    text = str(value).replace("\r", "\n")
    text = text.replace("•", "-").replace("â€¢", "-")
    text = text.replace("…", "...").replace("â€¦", "...")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def _truncate_text(text: str, max_chars: int) -> str:
    clean = _normalize_text(text)
    if len(clean) <= max_chars:
        return clean
    shortened = clean[: max_chars - 3].rstrip(" ,;:-")
    return f"{shortened}..."


def calculer_statistiques() -> dict | None:
    """
    PARTIE 2 : Calcule les statistiques d'engagement pour aujourd'hui
    et extrait les tendances par rapport à la veille.
    """
    print("============================================================")
    print("  PARTIE 2 — CALCUL DES STATISTIQUES")
    print("============================================================")

    if not DAILY_LOG_FILE.exists():
        print("  ❌ Aucun fichier de données trouvé.")
        return None

    try:
        df = _load_daily_log_df()
    except Exception as e:
        print(f"  ❌ Erreur de lecture du CSV : {e}")
        return None

    if df.empty:
        print("  ⚠️ Le fichier est vide.")
        return None

    df = df[~df["Username"].str.contains("Aggregated", na=False)]
    df = df[df["Cohort"] != "Global"]

    # Un seul enregistrement par utilisateur par jour
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    df = df.sort_values(["Date", "Username"]).drop_duplicates(subset=["Date", "Username"], keep="last")

    now_local = now_toronto()
    aujourdhui = now_local.strftime("%Y-%m-%d")
    hier = (now_local - timedelta(days=1)).strftime("%Y-%m-%d")

    df_jour = df[df["Date"] == aujourdhui]
    df_hier = df[df["Date"] == hier]

    stats = {
        "date_jour": aujourdhui,
        "date_hier": hier,
        "moyenne_streak_jour": 0.0,
        "moyenne_streak_hier": None,
        "utilisateurs_actifs": 0,
        "streaks_tombes_zero": 0,
        "reactivations_veille": 0,
        "progression_debutants_vers_standard": 0,
        "taux_progression_debutants_vers_standard": 0.0,
        "nb_profils_jour": len(df_jour),
        "nb_profils_hier": len(df_hier),
        "taux_conversion_plus": None,
        "taux_conversion_max": None,
        "delta_xp_moyen": 0.0,
        "cohortes": {
            "Debutants": {
                "actifs": 0,
                "total": 0,
                "retention": 0.0,
                "churn": 0.0,
                "tombes_zero": 0,
                "transitions_sortantes": 0,
                "transition_vers_standard": 0.0,
            },
            "Standard": {
                "actifs": 0,
                "total": 0,
                "retention": 0.0,
                "churn": 0.0,
                "tombes_zero": 0,
                "transitions_sortantes": 0,
                "transition_vers_standard": 0.0,
            },
            "Super-Actifs": {
                "actifs": 0,
                "total": 0,
                "retention": 0.0,
                "churn": 0.0,
                "tombes_zero": 0,
                "transitions_sortantes": 0,
                "transition_vers_standard": 0.0,
            },
        },
    }

    if not df_jour.empty:
        stats["moyenne_streak_jour"] = float(df_jour["Streak"].mean())
        stats["utilisateurs_actifs"] = int(len(df_jour[df_jour["Streak"] > 0]))
        stats["score_sante_jour"] = (
            (stats["utilisateurs_actifs"] / stats["nb_profils_jour"]) * 100
            if stats["nb_profils_jour"] > 0
            else 0
        )

        # Tant que HasMax n'est pas observe de facon fiable,
        # le taux Super reste un taux premium observable base sur HasPlus.
        if stats["nb_profils_jour"] > 0:
            stats["taux_conversion_max"] = _compute_max_rate_pct(df_jour)
            stats["taux_conversion_plus"] = _compute_super_rate_pct(df_jour)

        if "Cohort" in df_jour.columns:
            for cohorte in stats["cohortes"].keys():
                df_c = df_jour[df_jour["Cohort"] == cohorte]
                stats["cohortes"][cohorte]["total"] = int(len(df_c))
                stats["cohortes"][cohorte]["actifs"] = int(len(df_c[df_c["Streak"] > 0]))

    if not df_hier.empty and not df_jour.empty:
        stats["moyenne_streak_hier"] = float(df_hier["Streak"].mean())

        merged = df_hier.merge(
            df_jour,
            on="Username",
            suffixes=("_hier", "_jour"),
            how="inner",
        )

        if "TotalXP_hier" in merged.columns and "TotalXP_jour" in merged.columns:
            merged["Delta_XP"] = merged["TotalXP_jour"] - merged["TotalXP_hier"]
            merged.loc[merged["Delta_XP"] < 0, "Delta_XP"] = 0
            stats["delta_xp_moyen"] = float(merged["Delta_XP"].mean()) if not merged.empty else 0.0

        tombes_a_zero = merged[
            (merged["Streak_hier"] > 0) & (merged["Streak_jour"] == 0)
        ]
        stats["streaks_tombes_zero"] = int(len(tombes_a_zero))

        reactivated_mask = (merged["Streak_hier"] == 0) & (merged["Streak_jour"] > 0)
        stats["reactivations_veille"] = int(reactivated_mask.sum())

        actifs_hier = int(len(df_hier[df_hier["Streak"] > 0]))
        stats["taux_retention"] = (
            ((actifs_hier - stats["streaks_tombes_zero"]) / actifs_hier * 100)
            if actifs_hier > 0
            else 0
        )
        stats["taux_churn"] = (
            (stats["streaks_tombes_zero"] / actifs_hier * 100)
            if actifs_hier > 0
            else 0
        )

        if "Cohort_hier" in merged.columns:
            for cohorte in stats["cohortes"].keys():
                merged_c = merged[merged["Cohort_hier"] == cohorte]
                tombes_c = merged_c[
                    (merged_c["Streak_hier"] > 0) & (merged_c["Streak_jour"] == 0)
                ]
                actifs_hier_c = (
                    int(len(df_hier[(df_hier["Cohort"] == cohorte) & (df_hier["Streak"] > 0)]))
                    if "Cohort" in df_hier.columns
                    else 0
                )

                stats["cohortes"][cohorte]["tombes_zero"] = int(len(tombes_c))
                if actifs_hier_c > 0:
                    stats["cohortes"][cohorte]["churn"] = (len(tombes_c) / actifs_hier_c) * 100
                    stats["cohortes"][cohorte]["retention"] = (
                        (actifs_hier_c - len(tombes_c)) / actifs_hier_c
                    ) * 100

            if "Cohort_jour" in merged.columns:
                transitions_deb_standard = merged[
                    (merged["Cohort_hier"] == "Debutants")
                    & (merged["Streak_hier"] > 0)
                    & (merged["Streak_jour"] > 0)
                    & (merged["Cohort_jour"] == "Standard")
                ]
                actifs_hier_debutants = (
                    int(len(df_hier[(df_hier["Cohort"] == "Debutants") & (df_hier["Streak"] > 0)]))
                    if "Cohort" in df_hier.columns
                    else 0
                )
                stats["progression_debutants_vers_standard"] = int(len(transitions_deb_standard))
                stats["cohortes"]["Debutants"]["transitions_sortantes"] = int(len(transitions_deb_standard))
                if actifs_hier_debutants > 0:
                    taux_transition = (len(transitions_deb_standard) / actifs_hier_debutants) * 100
                    stats["taux_progression_debutants_vers_standard"] = taux_transition
                    stats["cohortes"]["Debutants"]["transition_vers_standard"] = taux_transition
    else:
        print("  ⚠️ Aucune donnée pour hier — comparaison impossible.\n")
        stats["taux_retention"] = 0
        stats["taux_churn"] = 0
        stats["score_sante_jour"] = stats.get("score_sante_jour", 0)

    print(f"  📊 Statistiques du {aujourdhui} :")
    print(f"     • Série Moyenne (Streak) : {stats['moyenne_streak_jour']:.1f} j")
    if stats["moyenne_streak_hier"] is not None:
        evolution = stats["moyenne_streak_jour"] - stats["moyenne_streak_hier"]
        signe = "+" if evolution > 0 else ""
        print(f"     • Évolution de la série : {signe}{evolution:.1f} j vs hier")
        print(f"     • Intensité d'Apprentissage : +{stats['delta_xp_moyen']:.0f} XP/jour")

    print(f"     • Utilisateurs Actifs : {stats['utilisateurs_actifs']}")
    taux_super = stats.get("taux_conversion_plus")
    print(f"     • Pénétration Super Duolingo : {f'{taux_super:.1f}%' if isinstance(taux_super, numbers.Number) else 'N/D'}")
    print(f"     • Ruptures de Série (abandons) : {stats['streaks_tombes_zero']}")
    print(f"     • Taux d'Abandon Global : {stats.get('taux_churn', 0):.1f}%")
    print(f"     • Reactivations vs veille : {stats.get('reactivations_veille', 0)}")
    print(
        "     • Progression Débutants -> Standard : "
        f"{stats.get('progression_debutants_vers_standard', 0)} "
        f"({stats.get('taux_progression_debutants_vers_standard', 0):.1f}%)"
    )

    print("\n     • Analyse par Segment (Cohorte) :")
    for nom, donnees in stats["cohortes"].items():
        print(
            f"       - {nom:12s} : {donnees['actifs']}/{donnees['total']} actifs | "
            f"Taux d'abandon: {donnees['churn']:.1f}%"
        )

    print(f"\n     • Profils collectés aujourd'hui : {stats['nb_profils_jour']}\n")

    return stats


def sauvegarder_rapport_excel(
    stats: dict,
    ia_report: str = None,
    financial_signals: dict | None = None,
    quarterly_nowcast: dict | None = None,
    dcf_valuation: dict | None = None,
) -> None:
    """
    Exporte les données d'engagement et les statistiques dans un fichier Excel (.xlsx)
    avec un design premium et l'analyse de l'IA.
    """
    print("  📂 Génération du rapport Excel Premium...")

    if not DAILY_LOG_FILE.exists():
        return

    try:
        df = _load_daily_log_df()
        df = df[~df["Username"].str.contains("Aggregated", na=False)]
        df = df[df["Cohort"] != "Global"]
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
        df = df.sort_values(["Date", "Username"]).drop_duplicates(subset=["Date", "Username"], keep="last")

        df_stats = build_summary_today_df(stats)
        df_resume = merge_summary_history(_charger_resume_historique(), df_stats, _normalize_summary_df)

        writer_mode = "a" if RAPPORT_EXCEL_FILE.exists() else "w"
        writer_kwargs = {"if_sheet_exists": "replace"} if writer_mode == "a" else {}

        with pd.ExcelWriter(RAPPORT_EXCEL_FILE, engine="openpyxl", mode=writer_mode, **writer_kwargs) as writer:
            df_resume.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)

            if financial_signals:
                signal_sheet_df = build_financial_signal_sheet_df(financial_signals)
                pd.DataFrame().to_excel(writer, sheet_name=SIGNALS_SHEET, index=False)
                signal_sheet_df.to_excel(writer, sheet_name=SIGNALS_RAW_SHEET, index=False)
            if quarterly_nowcast:
                quarterly_sheet_df = build_quarterly_nowcast_raw_df(quarterly_nowcast)
                pd.DataFrame().to_excel(writer, sheet_name=QUARTERLY_SHEET, index=False)
                quarterly_sheet_df.to_excel(writer, sheet_name=QUARTERLY_RAW_SHEET, index=False)
            if dcf_valuation:
                pd.DataFrame().to_excel(writer, sheet_name=DCF_SHEET, index=False)

            df_glossaire = build_kpi_dictionary_df()
            pd.DataFrame().to_excel(writer, sheet_name=GLOSSAIRE_SHEET, index=False)
            df_glossaire.to_excel(writer, sheet_name=GLOSSAIRE_RAW_SHEET, index=False)

        from openpyxl import load_workbook

        monthly_stats, df_daily_conv = build_monthly_trends_frames(df)

        with pd.ExcelWriter(RAPPORT_EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            if not monthly_stats.empty:
                monthly_stats.to_excel(writer, sheet_name=TRENDS_SHEET, index=False)

            if not df_daily_conv.empty:
                df_daily_conv.to_excel(writer, sheet_name=CHART_DATA_SHEET, index=False)

        wb = load_workbook(RAPPORT_EXCEL_FILE)

        remove_sheets(wb, [*BAD_SHEET_NAMES, AI_SHEET])
        hide_sheets(wb, [SIGNALS_RAW_SHEET, QUARTERLY_RAW_SHEET, GLOSSAIRE_RAW_SHEET])

        style_ctx = build_style_context()

        render_helpers = {
            "label_signal_bias": _label_signal_bias,
            "label_confidence": _label_confidence,
            "pretty_fr_number": _pretty_fr_number,
            "pretty_ratio_pct": _pretty_ratio_pct,
            "pretty_score": _pretty_score,
            "pretty_delta_pts": _pretty_delta_pts,
            "compact_summary_text": _compact_summary_text,
            "compact_bullet_text": _compact_bullet_text,
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            if sheet_name == SIGNALS_SHEET and financial_signals:
                render_financial_nowcast_sheet(ws, financial_signals, ia_report, style_ctx, render_helpers)
                ws.freeze_panes = "A5"
                continue

            if sheet_name == QUARTERLY_SHEET and quarterly_nowcast:
                render_quarterly_nowcast_sheet(ws, quarterly_nowcast, wb, QUARTERLY_RAW_SHEET, style_ctx)
                ws.freeze_panes = "A5"
                continue

            if sheet_name == DCF_SHEET and dcf_valuation:
                render_dcf_valuation_sheet(ws, dcf_valuation, style_ctx)
                ws.freeze_panes = "A5"
                continue

            if sheet_name == GLOSSAIRE_SHEET:
                render_kpi_dictionary_sheet(ws, wb, GLOSSAIRE_RAW_SHEET, style_ctx)
                ws.freeze_panes = "A12"
                continue

            if sheet_name == TRENDS_SHEET:
                try:
                    add_monthly_trends_chart(ws, wb, CHART_DATA_SHEET)
                except Exception as e:
                    print(f"  ⚠️ Erreur ajout graphique : {e}")

            apply_standard_table_style(
                ws,
                sheet_name,
                style_ctx,
                summary_sheet_name=SUMMARY_SHEET,
                glossary_sheet_name=GLOSSAIRE_SHEET,
                chart_data_sheet_name=CHART_DATA_SHEET,
            )

        ordered_sheet_names = [
            SUMMARY_SHEET,
            SIGNALS_SHEET,
            QUARTERLY_SHEET,
            TRENDS_SHEET,
            GLOSSAIRE_SHEET,
            SIGNALS_RAW_SHEET,
            QUARTERLY_RAW_SHEET,
            GLOSSAIRE_RAW_SHEET,
            CHART_DATA_SHEET,
        ]
        reorder_sheets(wb, ordered_sheet_names)

        if DCF_SHEET in wb.sheetnames:
            reorder_sheets(
                wb,
                [
                    SUMMARY_SHEET,
                    SIGNALS_SHEET,
                    QUARTERLY_SHEET,
                    TRENDS_SHEET,
                    DCF_SHEET,
                    GLOSSAIRE_SHEET,
                    SIGNALS_RAW_SHEET,
                    QUARTERLY_RAW_SHEET,
                    GLOSSAIRE_RAW_SHEET,
                    CHART_DATA_SHEET,
                ],
            )

        wb.save(RAPPORT_EXCEL_FILE)
        refresh_trends_dashboard(RAPPORT_EXCEL_FILE)
        wb = load_workbook(RAPPORT_EXCEL_FILE)
        hide_sheets(wb, [SIGNALS_RAW_SHEET, QUARTERLY_RAW_SHEET, GLOSSAIRE_RAW_SHEET, CHART_DATA_SHEET])
        ordered_sheet_names = [
            SUMMARY_SHEET,
            SIGNALS_SHEET,
            QUARTERLY_SHEET,
            TRENDS_SHEET,
            GLOSSAIRE_SHEET,
            SIGNALS_RAW_SHEET,
            QUARTERLY_RAW_SHEET,
            GLOSSAIRE_RAW_SHEET,
            CHART_DATA_SHEET,
        ]
        reorder_sheets(wb, ordered_sheet_names)
        if DCF_SHEET in wb.sheetnames:
            reorder_sheets(
                wb,
                [
                    SUMMARY_SHEET,
                    SIGNALS_SHEET,
                    QUARTERLY_SHEET,
                    TRENDS_SHEET,
                    DCF_SHEET,
                    GLOSSAIRE_SHEET,
                    SIGNALS_RAW_SHEET,
                    QUARTERLY_RAW_SHEET,
                    GLOSSAIRE_RAW_SHEET,
                    CHART_DATA_SHEET,
                ],
            )
        wb.save(RAPPORT_EXCEL_FILE)
        google_drive_file = _copier_rapport_vers_google_drive(RAPPORT_EXCEL_FILE)

        print("  ✅ Rapport Excel Premium sauvegardé dans :")
        print(f"     → {RAPPORT_EXCEL_FILE}")
        if google_drive_file is not None:
            print(f"     → Copie Google Drive : {google_drive_file}")

    except Exception as e:
        import traceback
        print(f"  ❌ Erreur lors de la sauvegarde Excel : {e}")
        traceback.print_exc()
