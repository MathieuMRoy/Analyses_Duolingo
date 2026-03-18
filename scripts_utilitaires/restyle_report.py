from __future__ import annotations

from datetime import datetime
from pathlib import Path
import numbers
import re
import sys

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def _parse_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, numbers.Number):
        return float(value)
    if isinstance(value, str):
        raw = value.strip()
        if not raw or raw.upper() in {"N/A", "NA"}:
            return None
        # Normalize decimal separator and strip non-numeric chars.
        cleaned = raw.replace(",", ".")
        cleaned = cleaned.replace("XP", "").replace("xp", "")
        cleaned = cleaned.replace("%", "")
        cleaned = re.sub(r"[^0-9\.\-\+]", "", cleaned)
        if cleaned in {"", "+", "-"}:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _parse_date(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def _normalize_summary_values(ws) -> None:
    # Convert text values in the summary sheet into numerics for proper formatting.
    if ws.max_row < 2:
        return
    for cell in ws[2]:
        header_value = ws.cell(1, cell.column).value
        header_key = str(header_value).strip().lower() if header_value is not None else ""
        if "date" in header_key:
            date_val = _parse_date(cell.value)
            if date_val:
                cell.value = date_val
            continue
        if any(k in header_key for k in ["taux", "attrition", "abandon", "score", "pénétration", "penetration"]):
            parsed = _parse_float(cell.value)
            if parsed is not None:
                cell.value = parsed
            continue
        if "xp" in header_key:
            parsed = _parse_float(cell.value)
            if parsed is not None:
                cell.value = parsed
            continue
        if any(k in header_key for k in ["évol", "evol", "delta", "Δ"]):
            parsed = _parse_float(cell.value)
            if parsed is not None:
                cell.value = parsed
            continue


def restyle_report(report_path: Path) -> None:
    if not report_path.exists():
        raise FileNotFoundError(report_path)

    wb = load_workbook(report_path)

    # Identify summary sheet for normalization.
    for sheet_name in wb.sheetnames:
        if "Résumé Financier" in sheet_name or "Resume Financier" in sheet_name:
            _normalize_summary_values(wb[sheet_name])
            break

    # --- Palette de Couleurs ---
    DUO_GREEN = "58CC02"    # Vert Duolingo
    DUO_BLUE  = "1CB0F6"    # Bleu Duolingo
    NAVY      = "1F4E78"    # Bleu Marine Pro
    LIGHT_GREY= "F2F2F2"    # Zébrures
    RED_SOFT  = "FFC7CE"    # Alerte
    WHITE     = "FFFFFF"

    # --- Styles de Base ---
    BASE_FONT_NAME = "Calibri"
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_font = Font(name=BASE_FONT_NAME, color=WHITE, bold=True, size=11)
    zebra_fill  = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
    alert_fill  = PatternFill(start_color=RED_SOFT, end_color=RED_SOFT, fill_type="solid")
    base_font   = Font(name=BASE_FONT_NAME, size=11, color="000000")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left", vertical="center", indent=1)

    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.sheet_view.showGridLines = False

        # Skip the AI dashboard sheet layout (keep its custom formatting).
        if "Analyse Stratégique" in sheet_name or "Tendances Mensuelles" in sheet_name or "Données Graphique" in sheet_name:
            continue

        # Headers
        ws.row_dimensions[1].height = 22
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            is_zebra = row_idx % 2 == 0
            for cell in row:
                cell.border = thin_border
                header_value = ws.cell(1, cell.column).value
                header_key = str(header_value).strip().lower() if header_value is not None else ""
                is_delta = ("Δ" in str(header_value)) or ("Î”" in str(header_value)) or ("delta" in header_key) or ("evol" in header_key)
                is_percent = ("%" in str(header_value)) or any(k in header_key for k in ["taux", "attrition", "abandon", "score", "pénétration", "penetration"])
                is_xp = "xp" in header_key
                is_streak = ("série" in header_key) or ("serie" in header_key) or ("streak" in header_key)
                is_panel = any(k in header_key for k in ["panel", "total", "profils", "actifs"])

                # Clean NaN
                if isinstance(cell.value, numbers.Number) and cell.value != cell.value:
                    cell.value = None

                # Alignment + formats
                if isinstance(cell.value, datetime):
                    cell.alignment = center_align
                    cell.font = base_font
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(cell.value, numbers.Number):
                    cell.alignment = center_align
                    cell.font = base_font
                    if is_delta and is_percent:
                        cell.number_format = '+0.0"%" ;-0.0"%" ;0.0"%"'
                    elif is_percent:
                        cell.number_format = '0.0"%"'
                    elif is_delta and is_xp:
                        cell.number_format = '+#,##0" XP";-#,##0" XP";0" XP"'
                    elif is_xp:
                        cell.number_format = '#,##0" XP"'
                    elif is_delta:
                        cell.number_format = "+0.0;-0.0;0.0"
                    elif is_streak:
                        cell.number_format = "0.0"
                    elif is_panel:
                        cell.number_format = "#,##0"

                    if is_delta:
                        if cell.value > 0:
                            cell.font = Font(name=BASE_FONT_NAME, size=11, color="008000", bold=True)
                        elif cell.value < 0:
                            cell.font = Font(name=BASE_FONT_NAME, size=11, color="C00000", bold=True)
                elif cell.value is None and is_delta:
                    cell.value = "N/A"
                    cell.alignment = center_align
                    cell.font = base_font
                else:
                    cell.alignment = left_align
                    cell.font = base_font
                    if "définition" in header_key or "definition" in header_key:
                        cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)

                if is_zebra:
                    cell.fill = zebra_fill

                if "Résumé Financier" in sheet_name and any(k in header_key for k in ["attrition", "abandon", "churn"]):
                    try:
                        if float(cell.value) > 0:
                            cell.fill = alert_fill
                    except Exception:
                        pass

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            first_real_cell = next((cell for cell in col if not isinstance(cell, MergedCell)), None)
            if first_real_cell is None:
                continue
            column = first_real_cell.column_letter
            for cell in col:
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 4)
            max_width = 70 if "Dictionnaire" in sheet_name else 60
            ws.column_dimensions[column].width = min(adjusted_width, max_width)

        # Freeze header + filter
        if ws.max_row > 1:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

    wb.save(report_path)


def main() -> int:
    if len(sys.argv) > 1:
        report_path = Path(sys.argv[1])
    else:
        report_path = Path(r"S:\Analyses_Duolingo\rapports_donnees\rapport_historique.xlsx")
    restyle_report(report_path)
    print(f"OK: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
