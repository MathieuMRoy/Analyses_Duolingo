
import openpyxl
from pathlib import Path

excel_path = Path("s:/Analyses_Duolingo/rapports_donnees/rapport_2026-03-15.xlsx")
if excel_path.exists():
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "📊 Données Graphique" in wb.sheetnames:
        ws = wb["📊 Données Graphique"]
        for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
            print(row)
