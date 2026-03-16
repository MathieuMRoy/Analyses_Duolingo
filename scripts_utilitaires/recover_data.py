
import openpyxl
import pandas as pd
from pathlib import Path

excel_path = Path("s:/Analyses_Duolingo/rapports_donnees/rapport_2026-03-15.xlsx")
target_users_path = Path("s:/Analyses_Duolingo/target_users.csv")
daily_log_path = Path("s:/Analyses_Duolingo/daily_streaks_log.csv")

if not excel_path.exists():
    print(f"Error: {excel_path} not found.")
    exit(1)

print(f"Loading {excel_path}...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

if "📊 Données Graphique" not in wb.sheetnames:
    print("Error: Hidden data sheet not found in Excel.")
    exit(1)

ws = wb["📊 Données Graphique"]
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    # Format according to stats.py: Date, Username, Cohort, Streak, TotalXP, HasPlus
    # Note: stats.py might have written multiple rows per user if there are multiple dates.
    # But usually it's the latest data.
    if any(row):
        data.append(row)

df = pd.DataFrame(data, columns=["Date", "Username", "Cohort", "Streak", "TotalXP", "HasPlus"])

# Recreate target_users.csv (unique Username, Cohort)
target_users = df[["Username", "Cohort"]].drop_duplicates()
target_users.to_csv(target_users_path, index=False)
print(f"Recovered {len(target_users)} users to {target_users_path}")

# Recreate daily_streaks_log.csv
# We only have data for the dates present in the Excel report, but it's better than nothing.
df.to_csv(daily_log_path, index=False)
print(f"Recovered {len(df)} streak log entries to {daily_log_path}")
