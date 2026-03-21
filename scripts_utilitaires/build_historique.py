from __future__ import annotations

from pathlib import Path
import shutil
import pandas as pd
from openpyxl import load_workbook

from duolingo_analyzer.config import RAPPORT_EXCEL_FILE, REPORT_DIR
from duolingo_analyzer.excel_dashboard import refresh_trends_dashboard
from duolingo_analyzer.reporting.sheets.kpi_dictionary_sheet import build_kpi_dictionary_df, render_kpi_dictionary_sheet
from duolingo_analyzer.reporting.styles import build_style_context
from duolingo_analyzer.columns import (
    GLOSSAIRE_RAW_SHEET,
    GLOSSAIRE_SHEET,
    LEGACY_SHEET_NAMES,
    SUMMARY_SHEET,
)
from duolingo_analyzer.report_builder import _load_summary_sheet
from duolingo_analyzer.stats import _normalize_summary_df
from scripts_utilitaires.restyle_report import restyle_report


HISTO_FILE = REPORT_DIR / "rapport_historique.xlsx"


def build_historique() -> None:
    report_files = sorted(REPORT_DIR.glob("rapport_*.xlsx"))
    report_files = [p for p in report_files if p.name != HISTO_FILE.name]

    if not report_files:
        print("Aucun rapport journalier trouvé.")
        return

    base_report = RAPPORT_EXCEL_FILE if RAPPORT_EXCEL_FILE.exists() else report_files[-1]

    shutil.copyfile(base_report, HISTO_FILE)

    frames: list[pd.DataFrame] = []
    for report_path in report_files:
        try:
            df = _load_summary_sheet(report_path)
            if df is not None and not df.empty:
                frames.append(_normalize_summary_df(df))
        except Exception:
            continue

    if not frames:
        print("Aucune feuille de résumé trouvée.")
        return

    df_all = pd.concat(frames, ignore_index=True)
    df_all = _normalize_summary_df(df_all)

    df_glossaire = build_kpi_dictionary_df()

    with pd.ExcelWriter(HISTO_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_all.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=GLOSSAIRE_SHEET, index=False)
        df_glossaire.to_excel(writer, sheet_name=GLOSSAIRE_RAW_SHEET, index=False)

    # Appliquer le style pro (convertit aussi les valeurs texte en nombres)
    restyle_report(HISTO_FILE)
    refresh_trends_dashboard(HISTO_FILE)

    wb = load_workbook(HISTO_FILE)
    for legacy_sheet_name in LEGACY_SHEET_NAMES:
        if legacy_sheet_name in wb.sheetnames:
            wb.remove(wb[legacy_sheet_name])
    if GLOSSAIRE_SHEET in wb.sheetnames:
        render_kpi_dictionary_sheet(wb[GLOSSAIRE_SHEET], wb, GLOSSAIRE_RAW_SHEET, build_style_context())
    if GLOSSAIRE_RAW_SHEET in wb.sheetnames:
        wb[GLOSSAIRE_RAW_SHEET].sheet_state = "hidden"
    wb.save(HISTO_FILE)

    print(f"OK: {HISTO_FILE}")


if __name__ == "__main__":
    build_historique()
