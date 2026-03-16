from __future__ import annotations

from datetime import datetime
from pathlib import Path
import sys

from openpyxl import load_workbook


SUMMARY_SHEET = "📊 Résumé Financier Q1"


def _parse_date(value: object) -> datetime.date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        raw = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    return None


def main() -> int:
    date_str = sys.argv[1] if len(sys.argv) > 1 else "2026-03-13"
    report_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(
        r"S:\Analyses_Duolingo\rapports_donnees\rapport_historique.xlsx"
    )

    if not report_path.exists():
        print(f"Fichier introuvable: {report_path}")
        return 1

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        print("Format de date attendu: YYYY-MM-DD")
        return 1

    wb = load_workbook(report_path)
    if SUMMARY_SHEET not in wb.sheetnames:
        print(f"Onglet introuvable: {SUMMARY_SHEET}")
        return 1

    ws = wb[SUMMARY_SHEET]
    rows_to_delete: list[int] = []
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        cell_date = _parse_date(cell_value)
        if cell_date == target_date:
            rows_to_delete.append(row)

    for row in reversed(rows_to_delete):
        ws.delete_rows(row, 1)

    wb.save(report_path)
    print(f"OK: supprimé {len(rows_to_delete)} ligne(s) pour {date_str}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
