
import openpyxl
import pandas as pd
from pathlib import Path

report_dir = Path("s:/Analyses_Duolingo/rapports_donnees")
daily_log_path = Path("s:/Analyses_Duolingo/daily_streaks_log.csv")

all_history = []

# Process all reports to extract aggregated data
for excel_file in sorted(report_dir.glob("rapport_*.xlsx")):
    print(f"Loading {excel_file}...")
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        if "📊 Données Graphique" in wb.sheetnames:
            ws = wb["📊 Données Graphique"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]: # Date
                    # We create dummy users for each cohort to preserve the trend mean
                    date = row[0]
                    # Row: Date, Global, Debutants, Standard, Super-Actifs
                    # Columns in daily_log: Date, Username, Cohort, Streak, TotalXP, HasPlus
                    
                    # Store these as "trend points" or dummy users
                    # This allows stats.py to still see the history for the charts
                    # Format: Date, Username, Cohort, Streak, TotalXP, HasPlus
                    all_history.append([date, "Aggregated_Global", "Global", row[1], 0, True])
                    all_history.append([date, "Aggregated_Debutants", "Debutants", row[2], 0, True])
                    all_history.append([date, "Aggregated_Standard", "Standard", row[3], 0, True])
                    all_history.append([date, "Aggregated_Super-Actifs", "Super-Actifs", row[4], 0, True])
    except Exception as e:
        print(f"Error processing {excel_file}: {e}")

if all_history:
    df_history = pd.DataFrame(all_history, columns=["Date", "Username", "Cohort", "Streak", "TotalXP", "HasPlus"])
    # Sort and save
    df_history.sort_values("Date").to_csv(daily_log_path, index=False)
    print(f"Successfully recovered {len(df_history)} history points from Excel reports into {daily_log_path}")
else:
    print("No history found in Excel reports.")
