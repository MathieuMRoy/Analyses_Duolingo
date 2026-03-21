"""
Generate a local preview workbook using the current daily log and nowcast layers.

This avoids running the network scraping pipeline and lets us validate workbook
rendering changes locally before pushing to GitHub.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from duolingo_analyzer.financial_signals import generate_financial_signal_package
from duolingo_analyzer.quarterly_nowcast import generate_quarterly_nowcast_package
from duolingo_analyzer import config as config_module
from duolingo_analyzer import report_builder as report_builder_module
from duolingo_analyzer import stats as stats_module


def main() -> None:
    preview_path = config_module.REPORT_DIR / "rapport_preview_nowcast_trimestriel.xlsx"

    original_report_path = config_module.RAPPORT_EXCEL_FILE
    original_drive_dir = config_module.GOOGLE_DRIVE_REPORT_DIR
    config_module.RAPPORT_EXCEL_FILE = preview_path
    config_module.GOOGLE_DRIVE_REPORT_DIR = None

    try:
        statistiques = stats_module.calculer_statistiques()
        if not statistiques:
            raise SystemExit("Impossible de calculer les statistiques locales.")

        date_jour = statistiques.get("date_jour")
        signaux_financiers = generate_financial_signal_package(date_jour)
        nowcast_trimestriel = generate_quarterly_nowcast_package(date_jour)

        # The quarterly sheet already carries its own explanatory text. We use
        # an empty IA payload here so local validation stays deterministic.
        report_builder_module.sauvegarder_rapport_excel(
            stats=statistiques,
            ia_report=None,
            financial_signals=None,
            quarterly_nowcast=nowcast_trimestriel,
            dcf_valuation=None,
        )
        config_module.RAPPORT_EXCEL_FILE = original_report_path
        config_module.GOOGLE_DRIVE_REPORT_DIR = original_drive_dir
    except Exception:
        # Ensure config is reset even if an error occurs during report generation
        config_module.RAPPORT_EXCEL_FILE = original_report_path
        config_module.GOOGLE_DRIVE_REPORT_DIR = original_drive_dir
        raise


    wb = load_workbook(preview_path)
    print("SHEETS:")
    for ws in wb.worksheets:
        print(f"- {ws.title} [{ws.sheet_state}]")

    if "Nowcast Trimestriel" in wb.sheetnames:
        ws = wb["Nowcast Trimestriel"]
        print("NOWCAST_PREVIEW:")
        print(f"A3={ws['A3'].value}")
        print(f"A4={ws['A4'].value}")
        print(f"C4={ws['C4'].value}")
        print(f"E5={ws['E5'].value}")
        print(f"E8={ws['E8'].value}")
        validations = list(ws.data_validations.dataValidation) if ws.data_validations else []
        print(f"VALIDATIONS={len(validations)}")

    print(preview_path)


if __name__ == "__main__":
    main()
